# -*- coding: utf-8 -*-
"""
Created on Tue Sep  6 17:48:54 2022

@author: camiv
"""

import os
import plotly.express as px
import pandas as pd
import plotly.graph_objects as go
import numpy as np
#import streamlit
import streamlit as st
import base64
import pytz
import datetime
from PIL import Image
import openpyxl
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from openpyxl.styles import Font
from openpyxl.styles import PatternFill
from openpyxl.styles.borders import Border, Side
from openpyxl.workbook import Workbook
import string
from openpyxl.styles import Border, Side, PatternFill
import fun

dir_main = "/app/app_cop/"

###############################################################################
#0. Inputs                                                                    #
###############################################################################
#Title
title="Game's report draft"


# Usuarios
d = {'user': ["admin","inspector","fleet1","fleet2","fleet3"], 'pass': ["Admin2022","GOT2020","Arya2022","Sansa2021","Daenerys2020"]}
df_users = pd.DataFrame(data=d)


#Dataframes
#@st.cache(ttl=60*60, max_entries=20, suppress_st_warning=True,allow_output_mutation=True)
#def load_score(dir_main):
#    df = pd.read_csv(os.path.join(dir_main+"score.csv"), sep=";")
#    return df
    
#df = load_score(dir_main)



###############################################################################
#1. User   login                                                              #
###############################################################################
# Título
image = Image.open(os.path.join(dir_main+"photo.jpg"))
st.sidebar.image(image, channels="BGR")
    
st.sidebar.title("Login")

user = st.sidebar.text_input("Username", value='', max_chars=None, key=None, type='default')

passw = st.sidebar.text_input("Password", value='', max_chars=None, key=None, type='password')


try:
  log0 =   df_users.loc[df_users['user']==user,['pass']]==passw
  log = log0.iloc[0,0]
except:
  pass

if len(log0)!=1:
  log= False

if log == False and user!="" and passw!="":
  st.sidebar.error('User or password incorrect')
if log == True:
  st.sidebar.success('Hi {u}, you enter succesfully!'.format(u=user))
  
  st.sidebar.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)


###############################################################################
#A. Fleets                                                                    #
###############################################################################
if user == "fleet1": 
    
###############################################################################
#2. Section selection                                                         #
###############################################################################
  section= ["Day 1", "Day 2", 'Day 3','Day 4','Day 5','Day 6','Day 7','Day 8','Day 9','Day 10']
  
  st.sidebar.title("Section selection")
  st.sidebar.markdown('''Select the day you wish to see.''')
  
  #Selection of daily report 
  pag_input = st.sidebar.selectbox('Select day:', section)

  st.sidebar.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)

###############################################################################
#2.i Day 1                                                                    #
###############################################################################
  if pag_input=="Day 1": 
    #Selection of section in the daily report
   # Title
   st.title(title)  
   st.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)
   st.header("**Day 1: Type in information**")
   st.markdown('''In this subsection the fleet manager types in the score and the profit for every action (this is already
accumulated)''')

   with st.expander("Vessel scores"):
#    if "df" not in st.session_state or score.columns[0]=="Unnamed: 0":
    if "df" not in st.session_state:
        st.session_state.df = pd.DataFrame(columns=["Vessel 1: Arya", 
                                                    "Vessel 2: Cersei", 
                                                    "Vessel 3: Lyanna", 
                                                    "Vessel 4: Daenerys"])


    st.markdown("Add Record")
    
    num_new_rows = 10
    ncol = st.session_state.df.shape[1]  # col count
    rw = -1
    
    with st.form(key="add form", clear_on_submit= True):
        cols = st.columns(ncol)
        rwdta = []
    
        for i in range(ncol):
            rwdta.append(cols[i].text_input(st.session_state.df.columns[i]))
    
        # you can insert code for a list comprehension here to change the data (rwdta) 
        # values into integer / float, if required
    
        if st.form_submit_button("Add"):
            if st.session_state.df.shape[0] == num_new_rows:
                st.error("Add row limit reached. Cant add any more records..")
            else:
                rw = st.session_state.df.shape[0] + 1
                st.info(f"Row: {rw} / {num_new_rows} added")
                st.session_state.df.loc[rw] = rwdta
    
                if st.session_state.df.shape[0] == num_new_rows:
                    st.error("Add row limit reached...")
    df = st.session_state.df
    df.to_csv("fleet 1/score1.csv",encoding='utf-8',index=False)
    st.text(df)
    
   with st.expander("Vessel profits"):
    if "df2" not in st.session_state:
        st.session_state.df2 = pd.DataFrame(columns=["Vessel 1: Arya", 
                                                    "Vessel 2: Cersei", 
                                                    "Vessel 3: Lyanna", 
                                                    "Vessel 4: Daenerys"])
    
    st.markdown("Add Record")
    
    ncol = st.session_state.df2.shape[1]  # col count
    rw = -1
    
    with st.form(key="add profits form", clear_on_submit= True):
        cols = st.columns(ncol)
        rwdta = []
    
        for i in range(ncol):
            rwdta.append(cols[i].text_input(st.session_state.df2.columns[i]))
    
        # you can insert code for a list comprehension here to change the data (rwdta) 
        # values into integer / float, if required
    
        if st.form_submit_button("Add"):
            if st.session_state.df2.shape[0] == num_new_rows:
                st.error("Add row limit reached. Cant add any more records..")
            else:
                rw = st.session_state.df2.shape[0] + 1
                st.info(f"Row: {rw} / {num_new_rows} added")
                st.session_state.df2.loc[rw] = rwdta
    
                if st.session_state.df2.shape[0] == num_new_rows:
                    st.error("Add row limit reached...")
    


    df2 = st.session_state.df2
    df2.to_csv("fleet 1/profits1.csv",encoding='utf-8',index=False)

    st.text(st.session_state.df2)
 
   with st.expander("Vessel fines"):
    if "df3" not in st.session_state:
        st.session_state.df3 = pd.DataFrame(columns=["Vessel 1: Arya", 
                                                    "Vessel 2: Cersei", 
                                                    "Vessel 3: Lyanna", 
                                                    "Vessel 4: Daenerys"])
    
    st.markdown("Add Record")
    
    ncol = st.session_state.df3.shape[1]  # col count
    rw = -1
    
    with st.form(key="add fines form", clear_on_submit= True):
        cols = st.columns(ncol)
        rwdta = []
    
        for i in range(ncol):
            rwdta.append(cols[i].text_input(st.session_state.df3.columns[i]))
    
        # you can insert code for a list comprehension here to change the data (rwdta) 
        # values into integer / float, if required
    
        if st.form_submit_button("Add"):
            if st.session_state.df3.shape[0] == num_new_rows:
                st.error("Add row limit reached. Cant add any more records..")
            else:
                rw = st.session_state.df3.shape[0] + 1
                st.info(f"Row: {rw} / {num_new_rows} added")
                st.session_state.df3.loc[rw] = rwdta
    
                if st.session_state.df3.shape[0] == num_new_rows:
                    st.error("Add row limit reached...")
    

    df3 = st.session_state.df3
    df3.to_csv("fleet 1/fines1.csv",encoding='utf-8',index=False)

    st.text(st.session_state.df3)

###############################################################################
#2.i Overall reports                                                          #
###############################################################################
   st.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)
   if st.button('Create overall reports'): 
    st.header("**Day 1: Overall reports**")
    st.markdown('''In this subsection the fleet manager can see the current accumulated score and profit of day 1 and the daily fleet scores and profits.''')
    st.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)
    #
    df = pd.read_csv(os.path.join(dir_main,'fleet 1/score1.csv'), sep=',')
    df2 = pd.read_csv(os.path.join(dir_main,'fleet 1/profits1.csv'), sep=',')
    df3 = pd.read_csv(os.path.join(dir_main,'fleet 1/fines1.csv'), sep=',')
    d = {'Vessel': df.columns[0:4] , 'Score':  df.iloc[-1] , 'Accumulated Profit':df2.iloc[-1] ,'Fines':df3.iloc[-1] }
    #d = pd.DataFrame()
    #d['Vessel']= df.columns[0:4]
    #d['CDF'] = df.values[::-1]
    #d['Score'] = df.iloc[-1] 
    #d['Accumulated Profit'] = df2.iloc[-1] 
    #['Fines'] = df3.iloc[-1] 
    outcomes = pd.DataFrame(data=d)
    outcomes=outcomes.reset_index(drop=True)
    st.dataframe(outcomes)
    
    wb = openpyxl.Workbook()
    wb.save(os.path.join(dir_main,'fleet 1/reports1.xlsx')) 
    book = load_workbook(os.path.join(dir_main,'fleet 1/reports1.xlsx'))
    del book['Sheet']   

    writer = pd.ExcelWriter(os.path.join(dir_main,'fleet 1/reports1.xlsx'), engine='openpyxl')
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}
    outcomes.to_excel(writer, sheet_name = "Outcomes 1", na_rep = "", index =   False, startrow = 0)  
    writer.save()

    #Plot scores 
    df01 = pd.read_csv(os.path.join(dir_main,'fleet 1/score1.csv'), sep=',')
    df01.index = np.arange(1, len(df01) + 1)
    #Create the daily average 
    df01['Fleet score']=df01.sum(axis=1)/4
    df01['Action'] = df01.index
    df01 = pd.melt(df01, id_vars='Action', value_vars=["Vessel 1: Arya", 
                                                    "Vessel 2: Cersei", 
                                                    "Vessel 3: Lyanna", 
                                                    "Vessel 4: Daenerys", 
                                                    "Fleet score"])
    st.subheader("*Daily fleet and vessel scores*")
    fig01 = px.line(df01, x='Action', y='value', color='variable', markers=True,range_x=[0,10],width=650,height=650)
    fig01.update_layout(yaxis_type='category')
    fig01.layout.xaxis.fixedrange= True
    fig01.layout.yaxis.fixedrange: True
    fig01.update(layout_showlegend=True)
    fig01.update_traces(patch={"line": {"color": "green", "width": 3, "dash": 'dot'}}, selector={"legendgroup": "Fleet score"}) 
    st.plotly_chart(fig01)
  
    #Plot profits  
    df02 = pd.read_csv(os.path.join(dir_main,'fleet 1/profits1.csv'), sep=',')
    df03 = pd.read_csv(os.path.join(dir_main,'fleet 1/fines1.csv'), sep=',')
    df02.index = np.arange(1, len(df02) + 1)
    df03.index = np.arange(1, len(df03) + 1)

    df02['Action'] = df02.index
    df02['Fleet profit']=df02.sum(axis=1)
    aux = df01.loc[df01.variable=="Fleet score",'value'].reset_index(drop=True)
    aux.index = np.arange(1, len(aux) + 1)
    df02['Fleet score'] = aux
    
    df03['Fleet fines']=df03.sum(axis=1)
    df02['Fleet w. score'] = df02['Fleet profit'] - df03['Fleet fines']
    
    df02 = pd.melt(df02, id_vars='Action', value_vars=["Fleet profit",
                                                       "Fleet score",
                                                       "Fleet w. score"])
    df02_1 = df02.loc[df02.variable == "Fleet score",:]
    df02_2 = df02.loc[df02.variable != "Fleet score",:]
    
    # Create figure with secondary y-axis 
    subfig = make_subplots(specs=[[{"secondary_y": True}]])

    # Defining Custom Colors
    colours = {
      "Fleet profit": "#A1A9A3",
      "Fleet w. score": "#DFDF70",
    }

    # create two independent figures with px.line each containing data from multiple columns
    fig = px.line(df02_1, x='Action', y='value',markers=True,range_x=[1,10],width=650,height=650, text="value")
    fig.layout.xaxis.fixedrange= True
    fig.update_traces(line_color='orange', line_width=2)
    fig2 = px.bar(df02_2, y='value', x='Action', color='variable',range_x=[1,10],width=650,height=650,color_discrete_map=colours)
    fig.update_traces(yaxis="y2")
    fig2.update(layout_showlegend=True)
    
    subfig.add_traces(fig2.data + fig.data)
    subfig.layout.xaxis.fixedrange= True
    
    # Set y-axes titles
    subfig.layout.yaxis.title="Kr."
    subfig.layout.yaxis2.title="Score"
    subfig.update_xaxes(range=[0,10], dtick=1)
    st.subheader("*Fleet profit loss by score increase*")
    st.plotly_chart(subfig)
  
    #Save the data for next day 
    cav=open(dir_main+"fleet 1/reports1.xlsx",'rb')
    b64_01 = base64.b64encode(cav.read()).decode('utf-8')
    href = f'<a href="data:file/cav;base64,{b64_01}" download="results1.xlsx">:arrow_down_small: Overall results day 1 </a>'
    st.markdown(href, unsafe_allow_html=True)

###############################################################################
#2.i Day 2                                                                    #
###############################################################################
  if pag_input=="Day 2": 
    #Selection of section in the daily report
   # Title
   st.title(title)  
   st.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)
   st.header("**Day 2: Type in information**")
   st.markdown('''In this subsection the fleet manager types in the score and the profit for every action (this is already
accumulated)''')

   with st.expander("Vessel scores"):
#    if "df" not in st.session_state or score.columns[0]=="Unnamed: 0":
    if "df" not in st.session_state:
        st.session_state.df = pd.DataFrame(columns=["Vessel 1: Arya", 
                                                    "Vessel 2: Cersei", 
                                                    "Vessel 3: Lyanna", 
                                                    "Vessel 4: Daenerys"])


    st.markdown("Add Record")
    
    num_new_rows = 10
    ncol = st.session_state.df.shape[1]  # col count
    rw = -1
    
    with st.form(key="add form", clear_on_submit= True):
        cols = st.columns(ncol)
        rwdta = []
    
        for i in range(ncol):
            rwdta.append(cols[i].text_input(st.session_state.df.columns[i]))
    
        # you can insert code for a list comprehension here to change the data (rwdta) 
        # values into integer / float, if required
    
        if st.form_submit_button("Add"):
            if st.session_state.df.shape[0] == num_new_rows:
                st.error("Add row limit reached. Cant add any more records..")
            else:
                rw = st.session_state.df.shape[0] + 1
                st.info(f"Row: {rw} / {num_new_rows} added")
                st.session_state.df.loc[rw] = rwdta
    
                if st.session_state.df.shape[0] == num_new_rows:
                    st.error("Add row limit reached...")
    df = st.session_state.df
    df.to_csv("fleet 1/score2.csv",encoding='utf-8',index=False)
    st.text(df)
    
   with st.expander("Vessel profits"):
    if "df2" not in st.session_state:
        st.session_state.df2 = pd.DataFrame(columns=["Vessel 1: Arya", 
                                                    "Vessel 2: Cersei", 
                                                    "Vessel 3: Lyanna", 
                                                    "Vessel 4: Daenerys"])
    
    st.markdown("Add Record")
    
    ncol = st.session_state.df2.shape[1]  # col count
    rw = -1
    
    with st.form(key="add profits form", clear_on_submit= True):
        cols = st.columns(ncol)
        rwdta = []
    
        for i in range(ncol):
            rwdta.append(cols[i].text_input(st.session_state.df2.columns[i]))
    
        # you can insert code for a list comprehension here to change the data (rwdta) 
        # values into integer / float, if required
    
        if st.form_submit_button("Add"):
            if st.session_state.df2.shape[0] == num_new_rows:
                st.error("Add row limit reached. Cant add any more records..")
            else:
                rw = st.session_state.df2.shape[0] + 1
                st.info(f"Row: {rw} / {num_new_rows} added")
                st.session_state.df2.loc[rw] = rwdta
    
                if st.session_state.df2.shape[0] == num_new_rows:
                    st.error("Add row limit reached...")
    


    df2 = st.session_state.df2
    df2.to_csv("fleet 1/profits2.csv",encoding='utf-8',index=False)

    st.text(st.session_state.df2)
 
   with st.expander("Vessel fines"):
    if "df3" not in st.session_state:
        st.session_state.df3 = pd.DataFrame(columns=["Vessel 1: Arya", 
                                                    "Vessel 2: Cersei", 
                                                    "Vessel 3: Lyanna", 
                                                    "Vessel 4: Daenerys"])
    
    st.markdown("Add Record")
    
    ncol = st.session_state.df3.shape[1]  # col count
    rw = -1
    
    with st.form(key="add fines form", clear_on_submit= True):
        cols = st.columns(ncol)
        rwdta = []
    
        for i in range(ncol):
            rwdta.append(cols[i].text_input(st.session_state.df3.columns[i]))
    
        # you can insert code for a list comprehension here to change the data (rwdta) 
        # values into integer / float, if required
    
        if st.form_submit_button("Add"):
            if st.session_state.df3.shape[0] == num_new_rows:
                st.error("Add row limit reached. Cant add any more records..")
            else:
                rw = st.session_state.df3.shape[0] + 1
                st.info(f"Row: {rw} / {num_new_rows} added")
                st.session_state.df3.loc[rw] = rwdta
    
                if st.session_state.df3.shape[0] == num_new_rows:
                    st.error("Add row limit reached...")
    

    df3 = st.session_state.df3
    df3.to_csv("fleet 1/fines2.csv",encoding='utf-8',index=False)

    st.text(st.session_state.df3)

###############################################################################
#2.i Overall reports                                                          #
###############################################################################
   st.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)
   if st.button('Create overall reports'): 
    st.header("**Day 2: Overall reports**")
    st.markdown('''In this subsection the fleet manager can see the current accumulated score and profit of day 2 and the daily fleet scores and profits.''')
    st.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)
    #
    df = pd.read_csv(os.path.join(dir_main,'fleet 1/score2.csv'), sep=',')
    df2 = pd.read_csv(os.path.join(dir_main,'fleet 1/profits2.csv'), sep=',')
    df3 = pd.read_csv(os.path.join(dir_main,'fleet 1/fines2.csv'), sep=',')
    df4 = pd.read_excel(os.path.join(dir_main,'fleet 1/results1.xlsx'))
    df4.index = df4.Vessel
    
    d = {'Vessel': df.columns[0:4] , 
         'Score':  df.iloc[-1] , 
         'Score t-1' : df4.Score,
         'Accumulated Profit':df2.iloc[-1] ,
         'Fines':df3.iloc[-1] }
    outcomes = pd.DataFrame(data=d)
    outcomes=outcomes.reset_index(drop=True)
    st.dataframe(outcomes)

    wb = openpyxl.Workbook()
    wb.save(os.path.join(dir_main,'fleet 1/reports2.xlsx')) 
    book = load_workbook(os.path.join(dir_main,'fleet 1/reports2.xlsx'))
    del book['Sheet']   

    writer = pd.ExcelWriter(os.path.join(dir_main,'fleet 1/reports2.xlsx'), engine='openpyxl')
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}
    outcomes.to_excel(writer, sheet_name = "Outcomes 2", na_rep = "", index =   False, startrow = 0)  
    writer.save()
    
    #Plot scores 
    df01 = pd.read_csv(os.path.join(dir_main,'fleet 1/score2.csv'), sep=',')
    df01.index = np.arange(1, len(df01) + 1)
    #Create the daily average 
    df01['Fleet score']=df01.sum(axis=1)/4
    df01['Action'] = df01.index
    df01 = pd.melt(df01, id_vars='Action', value_vars=["Vessel 1: Arya", 
                                                    "Vessel 2: Cersei", 
                                                    "Vessel 3: Lyanna", 
                                                    "Vessel 4: Daenerys", 
                                                    "Fleet score"])
    st.subheader("*Daily fleet and vessel scores*")
    fig01 = px.line(df01, x='Action', y='value', color='variable', markers=True,range_x=[0,10],width=650,height=650)
    fig01.update_layout(yaxis_type='category')
    fig01.layout.xaxis.fixedrange= True
    fig01.layout.yaxis.fixedrange: True
    fig01.update(layout_showlegend=True)
    fig01.update_traces(patch={"line": {"color": "green", "width": 3, "dash": 'dot'}}, selector={"legendgroup": "Fleet score"}) 
    st.plotly_chart(fig01)
  
    #Plot profits  
    df02 = pd.read_csv(os.path.join(dir_main,'fleet 1/profits2.csv'), sep=',')
    df03 = pd.read_csv(os.path.join(dir_main,'fleet 1/fines2.csv'), sep=',')
    df02.index = np.arange(1, len(df02) + 1)
    df03.index = np.arange(1, len(df03) + 1)

    df02['Action'] = df02.index
    df02['Fleet profit']=df02.sum(axis=1)
    aux = df01.loc[df01.variable=="Fleet score",'value'].reset_index(drop=True)
    aux.index = np.arange(1, len(aux) + 1)
    df02['Fleet score'] = aux
    
    df03['Fleet fines']=df03.sum(axis=1)
    df02['Fleet w. score'] = df02['Fleet profit'] - df03['Fleet fines']
    
    df02 = pd.melt(df02, id_vars='Action', value_vars=["Fleet profit",
                                                       "Fleet score",
                                                       "Fleet w. score"])
    df02_1 = df02.loc[df02.variable == "Fleet score",:]
    df02_2 = df02.loc[df02.variable != "Fleet score",:]
    
    # Create figure with secondary y-axis 
    subfig = make_subplots(specs=[[{"secondary_y": True}]])

    # Defining Custom Colors
    colours = {
      "Fleet profit": "#A1A9A3",
      "Fleet w. score": "#DFDF70",
    }

    # create two independent figures with px.line each containing data from multiple columns
    fig = px.line(df02_1, x='Action', y='value',markers=True,range_x=[1,10],width=650,height=650, text="value")
    fig.layout.xaxis.fixedrange= True
    fig.update_traces(line_color='orange', line_width=2)
    fig2 = px.bar(df02_2, y='value', x='Action', color='variable',range_x=[1,10],width=650,height=650,color_discrete_map=colours)
    fig.update_traces(yaxis="y2")
    fig2.update(layout_showlegend=True)
    
    subfig.add_traces(fig2.data + fig.data)
    subfig.layout.xaxis.fixedrange= True
    
    # Set y-axes titles
    subfig.layout.yaxis.title="Kr."
    subfig.layout.yaxis2.title="Score"
    subfig.update_xaxes(range=[0,10], dtick=1)
    st.subheader("*Fleet profit loss by score increase*")
    st.plotly_chart(subfig)
  
     #Save the data for next day 
    cav=open(dir_main+"fleet 1/reports2.xlsx",'rb')
    b64_01 = base64.b64encode(cav.read()).decode('utf-8')
    href = f'<a href="data:file/cav;base64,{b64_01}" download="results2.xlsx">:arrow_down_small: Overall results day 2 </a>'
    st.markdown(href, unsafe_allow_html=True)
 
###############################################################################
#2.i Day 3                                                                    #
###############################################################################
  if pag_input=="Day 3": 
    #Selection of section in the daily report
   # Title
   st.title(title)  
   st.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)
   st.header("**Day 3: Type in information**")
   st.markdown('''In this subsection the fleet manager types in the score and the profit for every action (this is already
accumulated)''')

   with st.expander("Vessel scores"):
#    if "df" not in st.session_state or score.columns[0]=="Unnamed: 0":
    if "df" not in st.session_state:
        st.session_state.df = pd.DataFrame(columns=["Vessel 1: Arya", 
                                                    "Vessel 2: Cersei", 
                                                    "Vessel 3: Lyanna", 
                                                    "Vessel 4: Daenerys"])


    st.markdown("Add Record")
    
    num_new_rows = 10
    ncol = st.session_state.df.shape[1]  # col count
    rw = -1
    
    with st.form(key="add form", clear_on_submit= True):
        cols = st.columns(ncol)
        rwdta = []
    
        for i in range(ncol):
            rwdta.append(cols[i].text_input(st.session_state.df.columns[i]))
    
        # you can insert code for a list comprehension here to change the data (rwdta) 
        # values into integer / float, if required
    
        if st.form_submit_button("Add"):
            if st.session_state.df.shape[0] == num_new_rows:
                st.error("Add row limit reached. Cant add any more records..")
            else:
                rw = st.session_state.df.shape[0] + 1
                st.info(f"Row: {rw} / {num_new_rows} added")
                st.session_state.df.loc[rw] = rwdta
    
                if st.session_state.df.shape[0] == num_new_rows:
                    st.error("Add row limit reached...")
    df = st.session_state.df
    df.to_csv("fleet 1/score3.csv",encoding='utf-8',index=False)
    st.text(df)
    
   with st.expander("Vessel profits"):
    if "df2" not in st.session_state:
        st.session_state.df2 = pd.DataFrame(columns=["Vessel 1: Arya", 
                                                    "Vessel 2: Cersei", 
                                                    "Vessel 3: Lyanna", 
                                                    "Vessel 4: Daenerys"])
    
    st.markdown("Add Record")
    
    ncol = st.session_state.df2.shape[1]  # col count
    rw = -1
    
    with st.form(key="add profits form", clear_on_submit= True):
        cols = st.columns(ncol)
        rwdta = []
    
        for i in range(ncol):
            rwdta.append(cols[i].text_input(st.session_state.df2.columns[i]))
    
        # you can insert code for a list comprehension here to change the data (rwdta) 
        # values into integer / float, if required
    
        if st.form_submit_button("Add"):
            if st.session_state.df2.shape[0] == num_new_rows:
                st.error("Add row limit reached. Cant add any more records..")
            else:
                rw = st.session_state.df2.shape[0] + 1
                st.info(f"Row: {rw} / {num_new_rows} added")
                st.session_state.df2.loc[rw] = rwdta
    
                if st.session_state.df2.shape[0] == num_new_rows:
                    st.error("Add row limit reached...")
    


    df2 = st.session_state.df2
    df2.to_csv("fleet 1/profits3.csv",encoding='utf-8',index=False)

    st.text(st.session_state.df2)
 
   with st.expander("Vessel fines"):
    if "df3" not in st.session_state:
        st.session_state.df3 = pd.DataFrame(columns=["Vessel 1: Arya", 
                                                    "Vessel 2: Cersei", 
                                                    "Vessel 3: Lyanna", 
                                                    "Vessel 4: Daenerys"])
    
    st.markdown("Add Record")
    
    ncol = st.session_state.df3.shape[1]  # col count
    rw = -1
    
    with st.form(key="add fines form", clear_on_submit= True):
        cols = st.columns(ncol)
        rwdta = []
    
        for i in range(ncol):
            rwdta.append(cols[i].text_input(st.session_state.df3.columns[i]))
    
        # you can insert code for a list comprehension here to change the data (rwdta) 
        # values into integer / float, if required
    
        if st.form_submit_button("Add"):
            if st.session_state.df3.shape[0] == num_new_rows:
                st.error("Add row limit reached. Cant add any more records..")
            else:
                rw = st.session_state.df3.shape[0] + 1
                st.info(f"Row: {rw} / {num_new_rows} added")
                st.session_state.df3.loc[rw] = rwdta
    
                if st.session_state.df3.shape[0] == num_new_rows:
                    st.error("Add row limit reached...")
    

    df3 = st.session_state.df3
    df3.to_csv("fleet 1/fines3.csv",encoding='utf-8',index=False)

    st.text(st.session_state.df3)

###############################################################################
#2.i Overall reports                                                          #
###############################################################################
   st.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)
   if st.button('Create overall reports'): 
    st.header("**Day 3: Overall reports**")
    st.markdown('''In this subsection the fleet manager can see the current accumulated score and profit of day 3 and the daily fleet scores and profits.''')
    st.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)
    #
    df = pd.read_csv(os.path.join(dir_main,'fleet 1/score3.csv'), sep=',')
    df2 = pd.read_csv(os.path.join(dir_main,'fleet 1/profits3.csv'), sep=',')
    df3 = pd.read_csv(os.path.join(dir_main,'fleet 1/fines3.csv'), sep=',')
    df4 = pd.read_excel(os.path.join(dir_main,'fleet 1/results2.xlsx'))
    df4.index = df4.Vessel
    df5 = pd.read_excel(os.path.join(dir_main,'fleet 1/results1.xlsx'))
    df5.index = df5.Vessel
    
    d = {'Vessel': df.columns[0:4] , 
         'Score':  df.iloc[-1] , 
         'Score t-1' : df4.Score,
         'Score t-2' : df5.Score,
         'Accumulated Profit':df2.iloc[-1] ,
         'Fines':df3.iloc[-1] }
    outcomes = pd.DataFrame(data=d)
    outcomes=outcomes.reset_index(drop=True)
    st.dataframe(outcomes)

    wb = openpyxl.Workbook()
    wb.save(os.path.join(dir_main,'fleet 1/reports3.xlsx')) 
    book = load_workbook(os.path.join(dir_main,'fleet 1/reports3.xlsx'))
    del book['Sheet']   

    writer = pd.ExcelWriter(os.path.join(dir_main,'fleet 1/reports3.xlsx'), engine='openpyxl')
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}
    outcomes.to_excel(writer, sheet_name = "Outcomes 3", na_rep = "", index =   False, startrow = 0)  
    writer.save()
    
    #Plot scores 
    df01 = pd.read_csv(os.path.join(dir_main,'fleet 1/score3.csv'), sep=',')
    df01.index = np.arange(1, len(df01) + 1)
    #Create the daily average 
    df01['Fleet score']=df01.sum(axis=1)/4
    df01['Action'] = df01.index
    df01 = pd.melt(df01, id_vars='Action', value_vars=["Vessel 1: Arya", 
                                                    "Vessel 2: Cersei", 
                                                    "Vessel 3: Lyanna", 
                                                    "Vessel 4: Daenerys", 
                                                    "Fleet score"])
    st.subheader("*Daily fleet and vessel scores*")
    fig01 = px.line(df01, x='Action', y='value', color='variable', markers=True,range_x=[0,10],width=650,height=650)
    fig01.update_layout(yaxis_type='category')
    fig01.layout.xaxis.fixedrange= True
    fig01.layout.yaxis.fixedrange: True
    fig01.update(layout_showlegend=True)
    fig01.update_traces(patch={"line": {"color": "green", "width": 3, "dash": 'dot'}}, selector={"legendgroup": "Fleet score"}) 
    st.plotly_chart(fig01)
  
    #Plot profits  
    df02 = pd.read_csv(os.path.join(dir_main,'fleet 1/profits3.csv'), sep=',')
    df03 = pd.read_csv(os.path.join(dir_main,'fleet 1/fines3.csv'), sep=',')
    df02.index = np.arange(1, len(df02) + 1)
    df03.index = np.arange(1, len(df03) + 1)

    df02['Action'] = df02.index
    df02['Fleet profit']=df02.sum(axis=1)
    aux = df01.loc[df01.variable=="Fleet score",'value'].reset_index(drop=True)
    aux.index = np.arange(1, len(aux) + 1)
    df02['Fleet score'] = aux
    
    df03['Fleet fines']=df03.sum(axis=1)
    df02['Fleet w. score'] = df02['Fleet profit'] - df03['Fleet fines']
    
    df02 = pd.melt(df02, id_vars='Action', value_vars=["Fleet profit",
                                                       "Fleet score",
                                                       "Fleet w. score"])
    df02_1 = df02.loc[df02.variable == "Fleet score",:]
    df02_2 = df02.loc[df02.variable != "Fleet score",:]
    
    # Create figure with secondary y-axis 
    subfig = make_subplots(specs=[[{"secondary_y": True}]])

    # Defining Custom Colors
    colours = {
      "Fleet profit": "#A1A9A3",
      "Fleet w. score": "#DFDF70",
    }

    # create two independent figures with px.line each containing data from multiple columns
    fig = px.line(df02_1, x='Action', y='value',markers=True,range_x=[1,10],width=650,height=650, text="value")
    fig.layout.xaxis.fixedrange= True
    fig.update_traces(line_color='orange', line_width=2)
    fig2 = px.bar(df02_2, y='value', x='Action', color='variable',range_x=[1,10],width=650,height=650,color_discrete_map=colours)
    fig.update_traces(yaxis="y2")
    fig2.update(layout_showlegend=True)
    
    subfig.add_traces(fig2.data + fig.data)
    subfig.layout.xaxis.fixedrange= True
    
    # Set y-axes titles
    subfig.layout.yaxis.title="Kr."
    subfig.layout.yaxis2.title="Score"
    subfig.update_xaxes(range=[0,10], dtick=1)
    st.subheader("*Fleet profit loss by score increase*")
    st.plotly_chart(subfig)
  
     #Save the data for next day 
    cav=open(dir_main+"fleet 1/reports3.xlsx",'rb')
    b64_01 = base64.b64encode(cav.read()).decode('utf-8')
    href = f'<a href="data:file/cav;base64,{b64_01}" download="results3.xlsx">:arrow_down_small: Overall results day 3 </a>'
    st.markdown(href, unsafe_allow_html=True)
  
###############################################################################
#2.i Day 4                                                                    #
###############################################################################
  if pag_input=="Day 4": 
    #Selection of section in the daily report
   # Title
   st.title(title)  
   st.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)
   st.header("**Day 4: Type in information**")
   st.markdown('''In this subsection the fleet manager types in the score and the profit for every action (this is already
accumulated)''')

   with st.expander("Vessel scores"):
#    if "df" not in st.session_state or score.columns[0]=="Unnamed: 0":
    if "df" not in st.session_state:
        st.session_state.df = pd.DataFrame(columns=["Vessel 1: Arya", 
                                                    "Vessel 2: Cersei", 
                                                    "Vessel 3: Lyanna", 
                                                    "Vessel 4: Daenerys"])


    st.markdown("Add Record")
    
    num_new_rows = 10
    ncol = st.session_state.df.shape[1]  # col count
    rw = -1
    
    with st.form(key="add form", clear_on_submit= True):
        cols = st.columns(ncol)
        rwdta = []
    
        for i in range(ncol):
            rwdta.append(cols[i].text_input(st.session_state.df.columns[i]))
    
        # you can insert code for a list comprehension here to change the data (rwdta) 
        # values into integer / float, if required
    
        if st.form_submit_button("Add"):
            if st.session_state.df.shape[0] == num_new_rows:
                st.error("Add row limit reached. Cant add any more records..")
            else:
                rw = st.session_state.df.shape[0] + 1
                st.info(f"Row: {rw} / {num_new_rows} added")
                st.session_state.df.loc[rw] = rwdta
    
                if st.session_state.df.shape[0] == num_new_rows:
                    st.error("Add row limit reached...")
    df = st.session_state.df
    df.to_csv("fleet 1/score4.csv",encoding='utf-8',index=False)
    st.text(df)
    
   with st.expander("Vessel profits"):
    if "df2" not in st.session_state:
        st.session_state.df2 = pd.DataFrame(columns=["Vessel 1: Arya", 
                                                    "Vessel 2: Cersei", 
                                                    "Vessel 3: Lyanna", 
                                                    "Vessel 4: Daenerys"])
    
    st.markdown("Add Record")
    
    ncol = st.session_state.df2.shape[1]  # col count
    rw = -1
    
    with st.form(key="add profits form", clear_on_submit= True):
        cols = st.columns(ncol)
        rwdta = []
    
        for i in range(ncol):
            rwdta.append(cols[i].text_input(st.session_state.df2.columns[i]))
    
        # you can insert code for a list comprehension here to change the data (rwdta) 
        # values into integer / float, if required
    
        if st.form_submit_button("Add"):
            if st.session_state.df2.shape[0] == num_new_rows:
                st.error("Add row limit reached. Cant add any more records..")
            else:
                rw = st.session_state.df2.shape[0] + 1
                st.info(f"Row: {rw} / {num_new_rows} added")
                st.session_state.df2.loc[rw] = rwdta
    
                if st.session_state.df2.shape[0] == num_new_rows:
                    st.error("Add row limit reached...")
    


    df2 = st.session_state.df2
    df2.to_csv("fleet 1/profits4.csv",encoding='utf-8',index=False)

    st.text(st.session_state.df2)
 
   with st.expander("Vessel fines"):
    if "df3" not in st.session_state:
        st.session_state.df3 = pd.DataFrame(columns=["Vessel 1: Arya", 
                                                    "Vessel 2: Cersei", 
                                                    "Vessel 3: Lyanna", 
                                                    "Vessel 4: Daenerys"])
    
    st.markdown("Add Record")
    
    ncol = st.session_state.df3.shape[1]  # col count
    rw = -1
    
    with st.form(key="add fines form", clear_on_submit= True):
        cols = st.columns(ncol)
        rwdta = []
    
        for i in range(ncol):
            rwdta.append(cols[i].text_input(st.session_state.df3.columns[i]))
    
        # you can insert code for a list comprehension here to change the data (rwdta) 
        # values into integer / float, if required
    
        if st.form_submit_button("Add"):
            if st.session_state.df3.shape[0] == num_new_rows:
                st.error("Add row limit reached. Cant add any more records..")
            else:
                rw = st.session_state.df3.shape[0] + 1
                st.info(f"Row: {rw} / {num_new_rows} added")
                st.session_state.df3.loc[rw] = rwdta
    
                if st.session_state.df3.shape[0] == num_new_rows:
                    st.error("Add row limit reached...")
    

    df3 = st.session_state.df3
    df3.to_csv("fleet 1/fines4.csv",encoding='utf-8',index=False)

    st.text(st.session_state.df3)

###############################################################################
#2.i Overall reports                                                          #
###############################################################################
   st.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)
   if st.button('Create overall reports'): 
    st.header("**Day 4: Overall reports**")
    st.markdown('''In this subsection the fleet manager can see the current accumulated score and profit of day 4 and the daily fleet scores and profits.''')
    st.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)
    #
    df = pd.read_csv(os.path.join(dir_main,'fleet 1/score4.csv'), sep=',')
    df2 = pd.read_csv(os.path.join(dir_main,'fleet 1/profits4.csv'), sep=',')
    df3 = pd.read_csv(os.path.join(dir_main,'fleet 1/fines4.csv'), sep=',')
    df4 = pd.read_excel(os.path.join(dir_main,'fleet 1/results2.xlsx'))
    df4.index = df4.Vessel
    df5 = pd.read_excel(os.path.join(dir_main,'fleet 1/results3.xlsx'))
    df5.index = df5.Vessel
    
    d = {'Vessel': df.columns[0:4] , 
         'Score':  df.iloc[-1] , 
         'Score t-1' : df4.Score,
         'Score t-2' : df5.Score,
         'Accumulated Profit':df2.iloc[-1] ,
         'Fines':df3.iloc[-1] }
    outcomes = pd.DataFrame(data=d)
    outcomes=outcomes.reset_index(drop=True)
    st.dataframe(outcomes)

    wb = openpyxl.Workbook()
    wb.save(os.path.join(dir_main,'fleet 1/reports4.xlsx')) 
    book = load_workbook(os.path.join(dir_main,'fleet 1/reports4.xlsx'))
    del book['Sheet']   

    writer = pd.ExcelWriter(os.path.join(dir_main,'fleet 1/reports4.xlsx'), engine='openpyxl')
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}
    outcomes.to_excel(writer, sheet_name = "Outcomes 4", na_rep = "", index =   False, startrow = 0)  
    writer.save()
    
    #Plot scores 
    df01 = pd.read_csv(os.path.join(dir_main,'fleet 1/score4.csv'), sep=',')
    df01.index = np.arange(1, len(df01) + 1)
    #Create the daily average 
    df01['Fleet score']=df01.sum(axis=1)/4
    df01['Action'] = df01.index
    df01 = pd.melt(df01, id_vars='Action', value_vars=["Vessel 1: Arya", 
                                                    "Vessel 2: Cersei", 
                                                    "Vessel 3: Lyanna", 
                                                    "Vessel 4: Daenerys", 
                                                    "Fleet score"])
    st.subheader("*Daily fleet and vessel scores*")
    fig01 = px.line(df01, x='Action', y='value', color='variable', markers=True,range_x=[0,10],width=650,height=650)
    fig01.update_layout(yaxis_type='category')
    fig01.layout.xaxis.fixedrange= True
    fig01.layout.yaxis.fixedrange: True
    fig01.update(layout_showlegend=True)
    fig01.update_traces(patch={"line": {"color": "green", "width": 3, "dash": 'dot'}}, selector={"legendgroup": "Fleet score"}) 
    st.plotly_chart(fig01)
  
    #Plot profits  
    df02 = pd.read_csv(os.path.join(dir_main,'fleet 1/profits4.csv'), sep=',')
    df03 = pd.read_csv(os.path.join(dir_main,'fleet 1/fines4.csv'), sep=',')
    df02.index = np.arange(1, len(df02) + 1)
    df03.index = np.arange(1, len(df03) + 1)

    df02['Action'] = df02.index
    df02['Fleet profit']=df02.sum(axis=1)
    aux = df01.loc[df01.variable=="Fleet score",'value'].reset_index(drop=True)
    aux.index = np.arange(1, len(aux) + 1)
    df02['Fleet score'] = aux
    
    df03['Fleet fines']=df03.sum(axis=1)
    df02['Fleet w. score'] = df02['Fleet profit'] - df03['Fleet fines']
    
    df02 = pd.melt(df02, id_vars='Action', value_vars=["Fleet profit",
                                                       "Fleet score",
                                                       "Fleet w. score"])
    df02_1 = df02.loc[df02.variable == "Fleet score",:]
    df02_2 = df02.loc[df02.variable != "Fleet score",:]
    
    # Create figure with secondary y-axis 
    subfig = make_subplots(specs=[[{"secondary_y": True}]])

    # Defining Custom Colors
    colours = {
      "Fleet profit": "#A1A9A3",
      "Fleet w. score": "#DFDF70",
    }

    # create two independent figures with px.line each containing data from multiple columns
    fig = px.line(df02_1, x='Action', y='value',markers=True,range_x=[1,10],width=650,height=650, text="value")
    fig.layout.xaxis.fixedrange= True
    fig.update_traces(line_color='orange', line_width=2)
    fig2 = px.bar(df02_2, y='value', x='Action', color='variable',range_x=[1,10],width=650,height=650,color_discrete_map=colours)
    fig.update_traces(yaxis="y2")
    fig2.update(layout_showlegend=True)
    
    subfig.add_traces(fig2.data + fig.data)
    subfig.layout.xaxis.fixedrange= True
    
    # Set y-axes titles
    subfig.layout.yaxis.title="Kr."
    subfig.layout.yaxis2.title="Score"
    subfig.update_xaxes(range=[0,10], dtick=1)
    st.subheader("*Fleet profit loss by score increase*")
    st.plotly_chart(subfig)
  
     #Save the data for next day 
    cav=open(dir_main+"fleet 1/reports4.xlsx",'rb')
    b64_01 = base64.b64encode(cav.read()).decode('utf-8')
    href = f'<a href="data:file/cav;base64,{b64_01}" download="results4.xlsx">:arrow_down_small: Overall results day 4 </a>'
    st.markdown(href, unsafe_allow_html=True)
  
###############################################################################
#2.i Day 5                                                                    #
###############################################################################
  if pag_input=="Day 5": 
    #Selection of section in the daily report
   # Title
   st.title(title)  
   st.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)
   st.header("**Day 5: Type in information**")
   st.markdown('''In this subsection the fleet manager types in the score and the profit for every action (this is already
accumulated)''')

   with st.expander("Vessel scores"):
#    if "df" not in st.session_state or score.columns[0]=="Unnamed: 0":
    if "df" not in st.session_state:
        st.session_state.df = pd.DataFrame(columns=["Vessel 1: Arya", 
                                                    "Vessel 2: Cersei", 
                                                    "Vessel 3: Lyanna", 
                                                    "Vessel 4: Daenerys"])


    st.markdown("Add Record")
    
    num_new_rows = 10
    ncol = st.session_state.df.shape[1]  # col count
    rw = -1
    
    with st.form(key="add form", clear_on_submit= True):
        cols = st.columns(ncol)
        rwdta = []
    
        for i in range(ncol):
            rwdta.append(cols[i].text_input(st.session_state.df.columns[i]))
    
        # you can insert code for a list comprehension here to change the data (rwdta) 
        # values into integer / float, if required
    
        if st.form_submit_button("Add"):
            if st.session_state.df.shape[0] == num_new_rows:
                st.error("Add row limit reached. Cant add any more records..")
            else:
                rw = st.session_state.df.shape[0] + 1
                st.info(f"Row: {rw} / {num_new_rows} added")
                st.session_state.df.loc[rw] = rwdta
    
                if st.session_state.df.shape[0] == num_new_rows:
                    st.error("Add row limit reached...")
    df = st.session_state.df
    df.to_csv("fleet 1/score5.csv",encoding='utf-8',index=False)
    st.text(df)
    
   with st.expander("Vessel profits"):
    if "df2" not in st.session_state:
        st.session_state.df2 = pd.DataFrame(columns=["Vessel 1: Arya", 
                                                    "Vessel 2: Cersei", 
                                                    "Vessel 3: Lyanna", 
                                                    "Vessel 4: Daenerys"])
    
    st.markdown("Add Record")
    
    ncol = st.session_state.df2.shape[1]  # col count
    rw = -1
    
    with st.form(key="add profits form", clear_on_submit= True):
        cols = st.columns(ncol)
        rwdta = []
    
        for i in range(ncol):
            rwdta.append(cols[i].text_input(st.session_state.df2.columns[i]))
    
        # you can insert code for a list comprehension here to change the data (rwdta) 
        # values into integer / float, if required
    
        if st.form_submit_button("Add"):
            if st.session_state.df2.shape[0] == num_new_rows:
                st.error("Add row limit reached. Cant add any more records..")
            else:
                rw = st.session_state.df2.shape[0] + 1
                st.info(f"Row: {rw} / {num_new_rows} added")
                st.session_state.df2.loc[rw] = rwdta
    
                if st.session_state.df2.shape[0] == num_new_rows:
                    st.error("Add row limit reached...")
    


    df2 = st.session_state.df2
    df2.to_csv("fleet 1/profits5.csv",encoding='utf-8',index=False)

    st.text(st.session_state.df2)
 
   with st.expander("Vessel fines"):
    if "df3" not in st.session_state:
        st.session_state.df3 = pd.DataFrame(columns=["Vessel 1: Arya", 
                                                    "Vessel 2: Cersei", 
                                                    "Vessel 3: Lyanna", 
                                                    "Vessel 4: Daenerys"])
    
    st.markdown("Add Record")
    
    ncol = st.session_state.df3.shape[1]  # col count
    rw = -1
    
    with st.form(key="add fines form", clear_on_submit= True):
        cols = st.columns(ncol)
        rwdta = []
    
        for i in range(ncol):
            rwdta.append(cols[i].text_input(st.session_state.df3.columns[i]))
    
        # you can insert code for a list comprehension here to change the data (rwdta) 
        # values into integer / float, if required
    
        if st.form_submit_button("Add"):
            if st.session_state.df3.shape[0] == num_new_rows:
                st.error("Add row limit reached. Cant add any more records..")
            else:
                rw = st.session_state.df3.shape[0] + 1
                st.info(f"Row: {rw} / {num_new_rows} added")
                st.session_state.df3.loc[rw] = rwdta
    
                if st.session_state.df3.shape[0] == num_new_rows:
                    st.error("Add row limit reached...")
    

    df3 = st.session_state.df3
    df3.to_csv("fleet 1/fines5.csv",encoding='utf-8',index=False)

    st.text(st.session_state.df3)

###############################################################################
#2.i Overall reports                                                          #
###############################################################################
   st.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)
   if st.button('Create overall reports'): 
    st.header("**Day 5: Overall reports**")
    st.markdown('''In this subsection the fleet manager can see the current accumulated score and profit of day 5 and the daily fleet scores and profits.''')
    st.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)
    #
    df = pd.read_csv(os.path.join(dir_main,'fleet 1/score5.csv'), sep=',')
    df2 = pd.read_csv(os.path.join(dir_main,'fleet 1/profits5.csv'), sep=',')
    df3 = pd.read_csv(os.path.join(dir_main,'fleet 1/fines5.csv'), sep=',')
    df4 = pd.read_excel(os.path.join(dir_main,'fleet 1/results3.xlsx'))
    df4.index = df4.Vessel
    df5 = pd.read_excel(os.path.join(dir_main,'fleet 1/results4.xlsx'))
    df5.index = df5.Vessel
    
    d = {'Vessel': df.columns[0:4] , 
         'Score':  df.iloc[-1] , 
         'Score t-1' : df4.Score,
         'Score t-2' : df5.Score,
         'Accumulated Profit':df2.iloc[-1] ,
         'Fines':df3.iloc[-1] }
    outcomes = pd.DataFrame(data=d)
    outcomes=outcomes.reset_index(drop=True)
    st.dataframe(outcomes)

    wb = openpyxl.Workbook()
    wb.save(os.path.join(dir_main,'fleet 1/reports5.xlsx')) 
    book = load_workbook(os.path.join(dir_main,'fleet 1/reports5.xlsx'))
    del book['Sheet']   

    writer = pd.ExcelWriter(os.path.join(dir_main,'fleet 1/reports5.xlsx'), engine='openpyxl')
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}
    outcomes.to_excel(writer, sheet_name = "Outcomes 5", na_rep = "", index =   False, startrow = 0)  
    writer.save()
    
    #Plot scores 
    df01 = pd.read_csv(os.path.join(dir_main,'fleet 1/score5.csv'), sep=',')
    df01.index = np.arange(1, len(df01) + 1)
    #Create the daily average 
    df01['Fleet score']=df01.sum(axis=1)/4
    df01['Action'] = df01.index
    df01 = pd.melt(df01, id_vars='Action', value_vars=["Vessel 1: Arya", 
                                                    "Vessel 2: Cersei", 
                                                    "Vessel 3: Lyanna", 
                                                    "Vessel 4: Daenerys", 
                                                    "Fleet score"])
    st.subheader("*Daily fleet and vessel scores*")
    fig01 = px.line(df01, x='Action', y='value', color='variable', markers=True,range_x=[0,10],width=650,height=650)
    fig01.update_layout(yaxis_type='category')
    fig01.layout.xaxis.fixedrange= True
    fig01.layout.yaxis.fixedrange: True
    fig01.update(layout_showlegend=True)
    fig01.update_traces(patch={"line": {"color": "green", "width": 3, "dash": 'dot'}}, selector={"legendgroup": "Fleet score"}) 
    st.plotly_chart(fig01)
  
    #Plot profits  
    df02 = pd.read_csv(os.path.join(dir_main,'fleet 1/profits5.csv'), sep=',')
    df03 = pd.read_csv(os.path.join(dir_main,'fleet 1/fines5.csv'), sep=',')
    df02.index = np.arange(1, len(df02) + 1)
    df03.index = np.arange(1, len(df03) + 1)

    df02['Action'] = df02.index
    df02['Fleet profit']=df02.sum(axis=1)
    aux = df01.loc[df01.variable=="Fleet score",'value'].reset_index(drop=True)
    aux.index = np.arange(1, len(aux) + 1)
    df02['Fleet score'] = aux
    
    df03['Fleet fines']=df03.sum(axis=1)
    df02['Fleet w. score'] = df02['Fleet profit'] - df03['Fleet fines']
    
    df02 = pd.melt(df02, id_vars='Action', value_vars=["Fleet profit",
                                                       "Fleet score",
                                                       "Fleet w. score"])
    df02_1 = df02.loc[df02.variable == "Fleet score",:]
    df02_2 = df02.loc[df02.variable != "Fleet score",:]
    
    # Create figure with secondary y-axis 
    subfig = make_subplots(specs=[[{"secondary_y": True}]])

    # Defining Custom Colors
    colours = {
      "Fleet profit": "#A1A9A3",
      "Fleet w. score": "#DFDF70",
    }

    # create two independent figures with px.line each containing data from multiple columns
    fig = px.line(df02_1, x='Action', y='value',markers=True,range_x=[1,10],width=650,height=650, text="value")
    fig.layout.xaxis.fixedrange= True
    fig.update_traces(line_color='orange', line_width=2)
    fig2 = px.bar(df02_2, y='value', x='Action', color='variable',range_x=[1,10],width=650,height=650,color_discrete_map=colours)
    fig.update_traces(yaxis="y2")
    fig2.update(layout_showlegend=True)
    
    subfig.add_traces(fig2.data + fig.data)
    subfig.layout.xaxis.fixedrange= True
    
    # Set y-axes titles
    subfig.layout.yaxis.title="Kr."
    subfig.layout.yaxis2.title="Score"
    subfig.update_xaxes(range=[0,10], dtick=1)
    st.subheader("*Fleet profit loss by score increase*")
    st.plotly_chart(subfig)
  
     #Save the data for next day 
    cav=open(dir_main+"fleet 1/reports5.xlsx",'rb')
    b64_01 = base64.b64encode(cav.read()).decode('utf-8')
    href = f'<a href="data:file/cav;base64,{b64_01}" download="results5.xlsx">:arrow_down_small: Overall results day 5 </a>'
    st.markdown(href, unsafe_allow_html=True)
 
###############################################################################
#2.i Day 6                                                                    #
###############################################################################
  if pag_input=="Day 6": 
    #Selection of section in the daily report
   # Title
   st.title(title)  
   st.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)
   st.header("**Day 6: Type in information**")
   st.markdown('''In this subsection the fleet manager types in the score and the profit for every action (this is already
accumulated)''')

   with st.expander("Vessel scores"):
#    if "df" not in st.session_state or score.columns[0]=="Unnamed: 0":
    if "df" not in st.session_state:
        st.session_state.df = pd.DataFrame(columns=["Vessel 1: Arya", 
                                                    "Vessel 2: Cersei", 
                                                    "Vessel 3: Lyanna", 
                                                    "Vessel 4: Daenerys"])


    st.markdown("Add Record")
    
    num_new_rows = 10
    ncol = st.session_state.df.shape[1]  # col count
    rw = -1
    
    with st.form(key="add form", clear_on_submit= True):
        cols = st.columns(ncol)
        rwdta = []
    
        for i in range(ncol):
            rwdta.append(cols[i].text_input(st.session_state.df.columns[i]))
    
        # you can insert code for a list comprehension here to change the data (rwdta) 
        # values into integer / float, if required
    
        if st.form_submit_button("Add"):
            if st.session_state.df.shape[0] == num_new_rows:
                st.error("Add row limit reached. Cant add any more records..")
            else:
                rw = st.session_state.df.shape[0] + 1
                st.info(f"Row: {rw} / {num_new_rows} added")
                st.session_state.df.loc[rw] = rwdta
    
                if st.session_state.df.shape[0] == num_new_rows:
                    st.error("Add row limit reached...")
    df = st.session_state.df
    df.to_csv("fleet 1/score6.csv",encoding='utf-8',index=False)
    st.text(df)
    
   with st.expander("Vessel profits"):
    if "df2" not in st.session_state:
        st.session_state.df2 = pd.DataFrame(columns=["Vessel 1: Arya", 
                                                    "Vessel 2: Cersei", 
                                                    "Vessel 3: Lyanna", 
                                                    "Vessel 4: Daenerys"])
    
    st.markdown("Add Record")
    
    ncol = st.session_state.df2.shape[1]  # col count
    rw = -1
    
    with st.form(key="add profits form", clear_on_submit= True):
        cols = st.columns(ncol)
        rwdta = []
    
        for i in range(ncol):
            rwdta.append(cols[i].text_input(st.session_state.df2.columns[i]))
    
        # you can insert code for a list comprehension here to change the data (rwdta) 
        # values into integer / float, if required
    
        if st.form_submit_button("Add"):
            if st.session_state.df2.shape[0] == num_new_rows:
                st.error("Add row limit reached. Cant add any more records..")
            else:
                rw = st.session_state.df2.shape[0] + 1
                st.info(f"Row: {rw} / {num_new_rows} added")
                st.session_state.df2.loc[rw] = rwdta
    
                if st.session_state.df2.shape[0] == num_new_rows:
                    st.error("Add row limit reached...")
    


    df2 = st.session_state.df2
    df2.to_csv("fleet 1/profits6.csv",encoding='utf-8',index=False)

    st.text(st.session_state.df2)
 
   with st.expander("Vessel fines"):
    if "df3" not in st.session_state:
        st.session_state.df3 = pd.DataFrame(columns=["Vessel 1: Arya", 
                                                    "Vessel 2: Cersei", 
                                                    "Vessel 3: Lyanna", 
                                                    "Vessel 4: Daenerys"])
    
    st.markdown("Add Record")
    
    ncol = st.session_state.df3.shape[1]  # col count
    rw = -1
    
    with st.form(key="add fines form", clear_on_submit= True):
        cols = st.columns(ncol)
        rwdta = []
    
        for i in range(ncol):
            rwdta.append(cols[i].text_input(st.session_state.df3.columns[i]))
    
        # you can insert code for a list comprehension here to change the data (rwdta) 
        # values into integer / float, if required
    
        if st.form_submit_button("Add"):
            if st.session_state.df3.shape[0] == num_new_rows:
                st.error("Add row limit reached. Cant add any more records..")
            else:
                rw = st.session_state.df3.shape[0] + 1
                st.info(f"Row: {rw} / {num_new_rows} added")
                st.session_state.df3.loc[rw] = rwdta
    
                if st.session_state.df3.shape[0] == num_new_rows:
                    st.error("Add row limit reached...")
    

    df3 = st.session_state.df3
    df3.to_csv("fleet 1/fines6.csv",encoding='utf-8',index=False)

    st.text(st.session_state.df3)

###############################################################################
#2.i Overall reports                                                          #
###############################################################################
   st.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)
   if st.button('Create overall reports'): 
    st.header("**Day 6: Overall reports**")
    st.markdown('''In this subsection the fleet manager can see the current accumulated score and profit of day 6 and the daily fleet scores and profits.''')
    st.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)
    #
    df = pd.read_csv(os.path.join(dir_main,'fleet 1/score6.csv'), sep=',')
    df2 = pd.read_csv(os.path.join(dir_main,'fleet 1/profits6.csv'), sep=',')
    df3 = pd.read_csv(os.path.join(dir_main,'fleet 1/fines6.csv'), sep=',')
    df4 = pd.read_excel(os.path.join(dir_main,'fleet 1/results4.xlsx'))
    df4.index = df4.Vessel
    df5 = pd.read_excel(os.path.join(dir_main,'fleet 1/results5.xlsx'))
    df5.index = df5.Vessel
    
    d = {'Vessel': df.columns[0:4] , 
         'Score':  df.iloc[-1] , 
         'Score t-1' : df4.Score,
         'Score t-2' : df5.Score,
         'Accumulated Profit':df2.iloc[-1] ,
         'Fines':df3.iloc[-1] }
    outcomes = pd.DataFrame(data=d)
    outcomes=outcomes.reset_index(drop=True)
    st.dataframe(outcomes)

    wb = openpyxl.Workbook()
    wb.save(os.path.join(dir_main,'fleet 1/reports6.xlsx')) 
    book = load_workbook(os.path.join(dir_main,'fleet 1/reports6.xlsx'))
    del book['Sheet']   

    writer = pd.ExcelWriter(os.path.join(dir_main,'fleet 1/reports6.xlsx'), engine='openpyxl')
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}
    outcomes.to_excel(writer, sheet_name = "Outcomes 6", na_rep = "", index =   False, startrow = 0)  
    writer.save()
    
    #Plot scores 
    df01 = pd.read_csv(os.path.join(dir_main,'fleet 1/score6.csv'), sep=',')
    df01.index = np.arange(1, len(df01) + 1)
    #Create the daily average 
    df01['Fleet score']=df01.sum(axis=1)/4
    df01['Action'] = df01.index
    df01 = pd.melt(df01, id_vars='Action', value_vars=["Vessel 1: Arya", 
                                                    "Vessel 2: Cersei", 
                                                    "Vessel 3: Lyanna", 
                                                    "Vessel 4: Daenerys", 
                                                    "Fleet score"])
    st.subheader("*Daily fleet and vessel scores*")
    fig01 = px.line(df01, x='Action', y='value', color='variable', markers=True,range_x=[0,10],width=650,height=650)
    fig01.update_layout(yaxis_type='category')
    fig01.layout.xaxis.fixedrange= True
    fig01.layout.yaxis.fixedrange: True
    fig01.update(layout_showlegend=True)
    fig01.update_traces(patch={"line": {"color": "green", "width": 3, "dash": 'dot'}}, selector={"legendgroup": "Fleet score"}) 
    st.plotly_chart(fig01)
  
    #Plot profits  
    df02 = pd.read_csv(os.path.join(dir_main,'fleet 1/profits6.csv'), sep=',')
    df03 = pd.read_csv(os.path.join(dir_main,'fleet 1/fines6.csv'), sep=',')
    df02.index = np.arange(1, len(df02) + 1)
    df03.index = np.arange(1, len(df03) + 1)

    df02['Action'] = df02.index
    df02['Fleet profit']=df02.sum(axis=1)
    aux = df01.loc[df01.variable=="Fleet score",'value'].reset_index(drop=True)
    aux.index = np.arange(1, len(aux) + 1)
    df02['Fleet score'] = aux
    
    df03['Fleet fines']=df03.sum(axis=1)
    df02['Fleet w. score'] = df02['Fleet profit'] - df03['Fleet fines']
    
    df02 = pd.melt(df02, id_vars='Action', value_vars=["Fleet profit",
                                                       "Fleet score",
                                                       "Fleet w. score"])
    df02_1 = df02.loc[df02.variable == "Fleet score",:]
    df02_2 = df02.loc[df02.variable != "Fleet score",:]
    
    # Create figure with secondary y-axis 
    subfig = make_subplots(specs=[[{"secondary_y": True}]])

    # Defining Custom Colors
    colours = {
      "Fleet profit": "#A1A9A3",
      "Fleet w. score": "#DFDF70",
    }

    # create two independent figures with px.line each containing data from multiple columns
    fig = px.line(df02_1, x='Action', y='value',markers=True,range_x=[1,10],width=650,height=650, text="value")
    fig.layout.xaxis.fixedrange= True
    fig.update_traces(line_color='orange', line_width=2)
    fig2 = px.bar(df02_2, y='value', x='Action', color='variable',range_x=[1,10],width=650,height=650,color_discrete_map=colours)
    fig.update_traces(yaxis="y2")
    fig2.update(layout_showlegend=True)
    
    subfig.add_traces(fig2.data + fig.data)
    subfig.layout.xaxis.fixedrange= True
    
    # Set y-axes titles
    subfig.layout.yaxis.title="Kr."
    subfig.layout.yaxis2.title="Score"
    subfig.update_xaxes(range=[0,10], dtick=1)
    st.subheader("*Fleet profit loss by score increase*")
    st.plotly_chart(subfig)
  
     #Save the data for next day 
    cav=open(dir_main+"fleet 1/reports6.xlsx",'rb')
    b64_01 = base64.b64encode(cav.read()).decode('utf-8')
    href = f'<a href="data:file/cav;base64,{b64_01}" download="results6.xlsx">:arrow_down_small: Overall results day 6 </a>'
    st.markdown(href, unsafe_allow_html=True)
  
###############################################################################
#2.i Day 7                                                                    #
###############################################################################
  if pag_input=="Day 7": 
    #Selection of section in the daily report
   # Title
   st.title(title)  
   st.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)
   st.header("**Day 7: Type in information**")
   st.markdown('''In this subsection the fleet manager types in the score and the profit for every action (this is already
accumulated)''')

   with st.expander("Vessel scores"):
#    if "df" not in st.session_state or score.columns[0]=="Unnamed: 0":
    if "df" not in st.session_state:
        st.session_state.df = pd.DataFrame(columns=["Vessel 1: Arya", 
                                                    "Vessel 2: Cersei", 
                                                    "Vessel 3: Lyanna", 
                                                    "Vessel 4: Daenerys"])


    st.markdown("Add Record")
    
    num_new_rows = 10
    ncol = st.session_state.df.shape[1]  # col count
    rw = -1
    
    with st.form(key="add form", clear_on_submit= True):
        cols = st.columns(ncol)
        rwdta = []
    
        for i in range(ncol):
            rwdta.append(cols[i].text_input(st.session_state.df.columns[i]))
    
        # you can insert code for a list comprehension here to change the data (rwdta) 
        # values into integer / float, if required
    
        if st.form_submit_button("Add"):
            if st.session_state.df.shape[0] == num_new_rows:
                st.error("Add row limit reached. Cant add any more records..")
            else:
                rw = st.session_state.df.shape[0] + 1
                st.info(f"Row: {rw} / {num_new_rows} added")
                st.session_state.df.loc[rw] = rwdta
    
                if st.session_state.df.shape[0] == num_new_rows:
                    st.error("Add row limit reached...")
    df = st.session_state.df
    df.to_csv("fleet 1/score7.csv",encoding='utf-8',index=False)
    st.text(df)
    
   with st.expander("Vessel profits"):
    if "df2" not in st.session_state:
        st.session_state.df2 = pd.DataFrame(columns=["Vessel 1: Arya", 
                                                    "Vessel 2: Cersei", 
                                                    "Vessel 3: Lyanna", 
                                                    "Vessel 4: Daenerys"])
    
    st.markdown("Add Record")
    
    ncol = st.session_state.df2.shape[1]  # col count
    rw = -1
    
    with st.form(key="add profits form", clear_on_submit= True):
        cols = st.columns(ncol)
        rwdta = []
    
        for i in range(ncol):
            rwdta.append(cols[i].text_input(st.session_state.df2.columns[i]))
    
        # you can insert code for a list comprehension here to change the data (rwdta) 
        # values into integer / float, if required
    
        if st.form_submit_button("Add"):
            if st.session_state.df2.shape[0] == num_new_rows:
                st.error("Add row limit reached. Cant add any more records..")
            else:
                rw = st.session_state.df2.shape[0] + 1
                st.info(f"Row: {rw} / {num_new_rows} added")
                st.session_state.df2.loc[rw] = rwdta
    
                if st.session_state.df2.shape[0] == num_new_rows:
                    st.error("Add row limit reached...")
    


    df2 = st.session_state.df2
    df2.to_csv("fleet 1/profits7.csv",encoding='utf-8',index=False)

    st.text(st.session_state.df2)
 
   with st.expander("Vessel fines"):
    if "df3" not in st.session_state:
        st.session_state.df3 = pd.DataFrame(columns=["Vessel 1: Arya", 
                                                    "Vessel 2: Cersei", 
                                                    "Vessel 3: Lyanna", 
                                                    "Vessel 4: Daenerys"])
    
    st.markdown("Add Record")
    
    ncol = st.session_state.df3.shape[1]  # col count
    rw = -1
    
    with st.form(key="add fines form", clear_on_submit= True):
        cols = st.columns(ncol)
        rwdta = []
    
        for i in range(ncol):
            rwdta.append(cols[i].text_input(st.session_state.df3.columns[i]))
    
        # you can insert code for a list comprehension here to change the data (rwdta) 
        # values into integer / float, if required
    
        if st.form_submit_button("Add"):
            if st.session_state.df3.shape[0] == num_new_rows:
                st.error("Add row limit reached. Cant add any more records..")
            else:
                rw = st.session_state.df3.shape[0] + 1
                st.info(f"Row: {rw} / {num_new_rows} added")
                st.session_state.df3.loc[rw] = rwdta
    
                if st.session_state.df3.shape[0] == num_new_rows:
                    st.error("Add row limit reached...")
    

    df3 = st.session_state.df3
    df3.to_csv("fleet 1/fines7.csv",encoding='utf-8',index=False)

    st.text(st.session_state.df3)

###############################################################################
#2.i Overall reports                                                          #
###############################################################################
   st.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)
   if st.button('Create overall reports'): 
    st.header("**Day 7: Overall reports**")
    st.markdown('''In this subsection the fleet manager can see the current accumulated score and profit of day 7 and the daily fleet scores and profits.''')
    st.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)
    #
    df = pd.read_csv(os.path.join(dir_main,'fleet 1/score7.csv'), sep=',')
    df2 = pd.read_csv(os.path.join(dir_main,'fleet 1/profits7.csv'), sep=',')
    df3 = pd.read_csv(os.path.join(dir_main,'fleet 1/fines7.csv'), sep=',')
    df4 = pd.read_excel(os.path.join(dir_main,'fleet 1/results5.xlsx'))
    df4.index = df4.Vessel
    df5 = pd.read_excel(os.path.join(dir_main,'fleet 1/results6.xlsx'))
    df5.index = df5.Vessel
    
    d = {'Vessel': df.columns[0:4] , 
         'Score':  df.iloc[-1] , 
         'Score t-1' : df4.Score,
         'Score t-2' : df5.Score,
         'Accumulated Profit':df2.iloc[-1] ,
         'Fines':df3.iloc[-1] }
    outcomes = pd.DataFrame(data=d)
    outcomes=outcomes.reset_index(drop=True)
    st.dataframe(outcomes)

    wb = openpyxl.Workbook()
    wb.save(os.path.join(dir_main,'fleet 1/reports7.xlsx')) 
    book = load_workbook(os.path.join(dir_main,'fleet 1/reports7.xlsx'))
    del book['Sheet']   

    writer = pd.ExcelWriter(os.path.join(dir_main,'fleet 1/reports7.xlsx'), engine='openpyxl')
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}
    outcomes.to_excel(writer, sheet_name = "Outcomes 7", na_rep = "", index =   False, startrow = 0)  
    writer.save()
    
    #Plot scores 
    df01 = pd.read_csv(os.path.join(dir_main,'fleet 1/score7.csv'), sep=',')
    df01.index = np.arange(1, len(df01) + 1)
    #Create the daily average 
    df01['Fleet score']=df01.sum(axis=1)/4
    df01['Action'] = df01.index
    df01 = pd.melt(df01, id_vars='Action', value_vars=["Vessel 1: Arya", 
                                                    "Vessel 2: Cersei", 
                                                    "Vessel 3: Lyanna", 
                                                    "Vessel 4: Daenerys", 
                                                    "Fleet score"])
    st.subheader("*Daily fleet and vessel scores*")
    fig01 = px.line(df01, x='Action', y='value', color='variable', markers=True,range_x=[0,10],width=650,height=650)
    fig01.update_layout(yaxis_type='category')
    fig01.layout.xaxis.fixedrange= True
    fig01.layout.yaxis.fixedrange: True
    fig01.update(layout_showlegend=True)
    fig01.update_traces(patch={"line": {"color": "green", "width": 3, "dash": 'dot'}}, selector={"legendgroup": "Fleet score"}) 
    st.plotly_chart(fig01)
  
    #Plot profits  
    df02 = pd.read_csv(os.path.join(dir_main,'fleet 1/profits7.csv'), sep=',')
    df03 = pd.read_csv(os.path.join(dir_main,'fleet 1/fines7.csv'), sep=',')
    df02.index = np.arange(1, len(df02) + 1)
    df03.index = np.arange(1, len(df03) + 1)

    df02['Action'] = df02.index
    df02['Fleet profit']=df02.sum(axis=1)
    aux = df01.loc[df01.variable=="Fleet score",'value'].reset_index(drop=True)
    aux.index = np.arange(1, len(aux) + 1)
    df02['Fleet score'] = aux
    
    df03['Fleet fines']=df03.sum(axis=1)
    df02['Fleet w. score'] = df02['Fleet profit'] - df03['Fleet fines']
    
    df02 = pd.melt(df02, id_vars='Action', value_vars=["Fleet profit",
                                                       "Fleet score",
                                                       "Fleet w. score"])
    df02_1 = df02.loc[df02.variable == "Fleet score",:]
    df02_2 = df02.loc[df02.variable != "Fleet score",:]
    
    # Create figure with secondary y-axis 
    subfig = make_subplots(specs=[[{"secondary_y": True}]])

    # Defining Custom Colors
    colours = {
      "Fleet profit": "#A1A9A3",
      "Fleet w. score": "#DFDF70",
    }

    # create two independent figures with px.line each containing data from multiple columns
    fig = px.line(df02_1, x='Action', y='value',markers=True,range_x=[1,10],width=650,height=650, text="value")
    fig.layout.xaxis.fixedrange= True
    fig.update_traces(line_color='orange', line_width=2)
    fig2 = px.bar(df02_2, y='value', x='Action', color='variable',range_x=[1,10],width=650,height=650,color_discrete_map=colours)
    fig.update_traces(yaxis="y2")
    fig2.update(layout_showlegend=True)
    
    subfig.add_traces(fig2.data + fig.data)
    subfig.layout.xaxis.fixedrange= True
    
    # Set y-axes titles
    subfig.layout.yaxis.title="Kr."
    subfig.layout.yaxis2.title="Score"
    subfig.update_xaxes(range=[0,10], dtick=1)
    st.subheader("*Fleet profit loss by score increase*")
    st.plotly_chart(subfig)
  
     #Save the data for next day 
    cav=open(dir_main+"fleet 1/reports7.xlsx",'rb')
    b64_01 = base64.b64encode(cav.read()).decode('utf-8')
    href = f'<a href="data:file/cav;base64,{b64_01}" download="results7.xlsx">:arrow_down_small: Overall results day 7 </a>'
    st.markdown(href, unsafe_allow_html=True)
  
###############################################################################
#2.i Day 8                                                                    #
###############################################################################
  if pag_input=="Day 8": 
    #Selection of section in the daily report
   # Title
   st.title(title)  
   st.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)
   st.header("**Day 8: Type in information**")
   st.markdown('''In this subsection the fleet manager types in the score and the profit for every action (this is already
accumulated)''')

   with st.expander("Vessel scores"):
#    if "df" not in st.session_state or score.columns[0]=="Unnamed: 0":
    if "df" not in st.session_state:
        st.session_state.df = pd.DataFrame(columns=["Vessel 1: Arya", 
                                                    "Vessel 2: Cersei", 
                                                    "Vessel 3: Lyanna", 
                                                    "Vessel 4: Daenerys"])


    st.markdown("Add Record")
    
    num_new_rows = 10
    ncol = st.session_state.df.shape[1]  # col count
    rw = -1
    
    with st.form(key="add form", clear_on_submit= True):
        cols = st.columns(ncol)
        rwdta = []
    
        for i in range(ncol):
            rwdta.append(cols[i].text_input(st.session_state.df.columns[i]))
    
        # you can insert code for a list comprehension here to change the data (rwdta) 
        # values into integer / float, if required
    
        if st.form_submit_button("Add"):
            if st.session_state.df.shape[0] == num_new_rows:
                st.error("Add row limit reached. Cant add any more records..")
            else:
                rw = st.session_state.df.shape[0] + 1
                st.info(f"Row: {rw} / {num_new_rows} added")
                st.session_state.df.loc[rw] = rwdta
    
                if st.session_state.df.shape[0] == num_new_rows:
                    st.error("Add row limit reached...")
    df = st.session_state.df
    df.to_csv("fleet 1/score8.csv",encoding='utf-8',index=False)
    st.text(df)
    
   with st.expander("Vessel profits"):
    if "df2" not in st.session_state:
        st.session_state.df2 = pd.DataFrame(columns=["Vessel 1: Arya", 
                                                    "Vessel 2: Cersei", 
                                                    "Vessel 3: Lyanna", 
                                                    "Vessel 4: Daenerys"])
    
    st.markdown("Add Record")
    
    ncol = st.session_state.df2.shape[1]  # col count
    rw = -1
    
    with st.form(key="add profits form", clear_on_submit= True):
        cols = st.columns(ncol)
        rwdta = []
    
        for i in range(ncol):
            rwdta.append(cols[i].text_input(st.session_state.df2.columns[i]))
    
        # you can insert code for a list comprehension here to change the data (rwdta) 
        # values into integer / float, if required
    
        if st.form_submit_button("Add"):
            if st.session_state.df2.shape[0] == num_new_rows:
                st.error("Add row limit reached. Cant add any more records..")
            else:
                rw = st.session_state.df2.shape[0] + 1
                st.info(f"Row: {rw} / {num_new_rows} added")
                st.session_state.df2.loc[rw] = rwdta
    
                if st.session_state.df2.shape[0] == num_new_rows:
                    st.error("Add row limit reached...")
    


    df2 = st.session_state.df2
    df2.to_csv("fleet 1/profits8.csv",encoding='utf-8',index=False)

    st.text(st.session_state.df2)
 
   with st.expander("Vessel fines"):
    if "df3" not in st.session_state:
        st.session_state.df3 = pd.DataFrame(columns=["Vessel 1: Arya", 
                                                    "Vessel 2: Cersei", 
                                                    "Vessel 3: Lyanna", 
                                                    "Vessel 4: Daenerys"])
    
    st.markdown("Add Record")
    
    ncol = st.session_state.df3.shape[1]  # col count
    rw = -1
    
    with st.form(key="add fines form", clear_on_submit= True):
        cols = st.columns(ncol)
        rwdta = []
    
        for i in range(ncol):
            rwdta.append(cols[i].text_input(st.session_state.df3.columns[i]))
    
        # you can insert code for a list comprehension here to change the data (rwdta) 
        # values into integer / float, if required
    
        if st.form_submit_button("Add"):
            if st.session_state.df3.shape[0] == num_new_rows:
                st.error("Add row limit reached. Cant add any more records..")
            else:
                rw = st.session_state.df3.shape[0] + 1
                st.info(f"Row: {rw} / {num_new_rows} added")
                st.session_state.df3.loc[rw] = rwdta
    
                if st.session_state.df3.shape[0] == num_new_rows:
                    st.error("Add row limit reached...")
    

    df3 = st.session_state.df3
    df3.to_csv("fleet 1/fines8.csv",encoding='utf-8',index=False)

    st.text(st.session_state.df3)

###############################################################################
#2.i Overall reports                                                          #
###############################################################################
   st.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)
   if st.button('Create overall reports'): 
    st.header("**Day 8: Overall reports**")
    st.markdown('''In this subsection the fleet manager can see the current accumulated score and profit of day 8 and the daily fleet scores and profits.''')
    st.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)
    #
    df = pd.read_csv(os.path.join(dir_main,'fleet 1/score8.csv'), sep=',')
    df2 = pd.read_csv(os.path.join(dir_main,'fleet 1/profits8.csv'), sep=',')
    df3 = pd.read_csv(os.path.join(dir_main,'fleet 1/fines8.csv'), sep=',')
    df4 = pd.read_excel(os.path.join(dir_main,'fleet 1/results6.xlsx'))
    df4.index = df4.Vessel
    df5 = pd.read_excel(os.path.join(dir_main,'fleet 1/results7.xlsx'))
    df5.index = df5.Vessel
    
    d = {'Vessel': df.columns[0:4] , 
         'Score':  df.iloc[-1] , 
         'Score t-1' : df4.Score,
         'Score t-2' : df5.Score,
         'Accumulated Profit':df2.iloc[-1] ,
         'Fines':df3.iloc[-1] }
    outcomes = pd.DataFrame(data=d)
    outcomes=outcomes.reset_index(drop=True)
    st.dataframe(outcomes)

    wb = openpyxl.Workbook()
    wb.save(os.path.join(dir_main,'fleet 1/reports8.xlsx')) 
    book = load_workbook(os.path.join(dir_main,'fleet 1/reports8.xlsx'))
    del book['Sheet']   

    writer = pd.ExcelWriter(os.path.join(dir_main,'fleet 1/reports8.xlsx'), engine='openpyxl')
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}
    outcomes.to_excel(writer, sheet_name = "Outcomes 8", na_rep = "", index =   False, startrow = 0)  
    writer.save()
    
    #Plot scores 
    df01 = pd.read_csv(os.path.join(dir_main,'fleet 1/score8.csv'), sep=',')
    df01.index = np.arange(1, len(df01) + 1)
    #Create the daily average 
    df01['Fleet score']=df01.sum(axis=1)/4
    df01['Action'] = df01.index
    df01 = pd.melt(df01, id_vars='Action', value_vars=["Vessel 1: Arya", 
                                                    "Vessel 2: Cersei", 
                                                    "Vessel 3: Lyanna", 
                                                    "Vessel 4: Daenerys", 
                                                    "Fleet score"])
    st.subheader("*Daily fleet and vessel scores*")
    fig01 = px.line(df01, x='Action', y='value', color='variable', markers=True,range_x=[0,10],width=650,height=650)
    fig01.update_layout(yaxis_type='category')
    fig01.layout.xaxis.fixedrange= True
    fig01.layout.yaxis.fixedrange: True
    fig01.update(layout_showlegend=True)
    fig01.update_traces(patch={"line": {"color": "green", "width": 3, "dash": 'dot'}}, selector={"legendgroup": "Fleet score"}) 
    st.plotly_chart(fig01)
  
    #Plot profits  
    df02 = pd.read_csv(os.path.join(dir_main,'fleet 1/profits8.csv'), sep=',')
    df03 = pd.read_csv(os.path.join(dir_main,'fleet 1/fines8.csv'), sep=',')
    df02.index = np.arange(1, len(df02) + 1)
    df03.index = np.arange(1, len(df03) + 1)

    df02['Action'] = df02.index
    df02['Fleet profit']=df02.sum(axis=1)
    aux = df01.loc[df01.variable=="Fleet score",'value'].reset_index(drop=True)
    aux.index = np.arange(1, len(aux) + 1)
    df02['Fleet score'] = aux
    
    df03['Fleet fines']=df03.sum(axis=1)
    df02['Fleet w. score'] = df02['Fleet profit'] - df03['Fleet fines']
    
    df02 = pd.melt(df02, id_vars='Action', value_vars=["Fleet profit",
                                                       "Fleet score",
                                                       "Fleet w. score"])
    df02_1 = df02.loc[df02.variable == "Fleet score",:]
    df02_2 = df02.loc[df02.variable != "Fleet score",:]
    
    # Create figure with secondary y-axis 
    subfig = make_subplots(specs=[[{"secondary_y": True}]])

    # Defining Custom Colors
    colours = {
      "Fleet profit": "#A1A9A3",
      "Fleet w. score": "#DFDF70",
    }

    # create two independent figures with px.line each containing data from multiple columns
    fig = px.line(df02_1, x='Action', y='value',markers=True,range_x=[1,10],width=650,height=650, text="value")
    fig.layout.xaxis.fixedrange= True
    fig.update_traces(line_color='orange', line_width=2)
    fig2 = px.bar(df02_2, y='value', x='Action', color='variable',range_x=[1,10],width=650,height=650,color_discrete_map=colours)
    fig.update_traces(yaxis="y2")
    fig2.update(layout_showlegend=True)
    
    subfig.add_traces(fig2.data + fig.data)
    subfig.layout.xaxis.fixedrange= True
    
    # Set y-axes titles
    subfig.layout.yaxis.title="Kr."
    subfig.layout.yaxis2.title="Score"
    subfig.update_xaxes(range=[0,10], dtick=1)
    st.subheader("*Fleet profit loss by score increase*")
    st.plotly_chart(subfig)
  
     #Save the data for next day 
    cav=open(dir_main+"fleet 1/reports8.xlsx",'rb')
    b64_01 = base64.b64encode(cav.read()).decode('utf-8')
    href = f'<a href="data:file/cav;base64,{b64_01}" download="results8.xlsx">:arrow_down_small: Overall results day 8 </a>'
    st.markdown(href, unsafe_allow_html=True)
  
###############################################################################
#2.i Day 4                                                                    #
###############################################################################
  if pag_input=="Day 9": 
    #Selection of section in the daily report
   # Title
   st.title(title)  
   st.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)
   st.header("**Day 9: Type in information**")
   st.markdown('''In this subsection the fleet manager types in the score and the profit for every action (this is already
accumulated)''')

   with st.expander("Vessel scores"):
#    if "df" not in st.session_state or score.columns[0]=="Unnamed: 0":
    if "df" not in st.session_state:
        st.session_state.df = pd.DataFrame(columns=["Vessel 1: Arya", 
                                                    "Vessel 2: Cersei", 
                                                    "Vessel 3: Lyanna", 
                                                    "Vessel 4: Daenerys"])


    st.markdown("Add Record")
    
    num_new_rows = 10
    ncol = st.session_state.df.shape[1]  # col count
    rw = -1
    
    with st.form(key="add form", clear_on_submit= True):
        cols = st.columns(ncol)
        rwdta = []
    
        for i in range(ncol):
            rwdta.append(cols[i].text_input(st.session_state.df.columns[i]))
    
        # you can insert code for a list comprehension here to change the data (rwdta) 
        # values into integer / float, if required
    
        if st.form_submit_button("Add"):
            if st.session_state.df.shape[0] == num_new_rows:
                st.error("Add row limit reached. Cant add any more records..")
            else:
                rw = st.session_state.df.shape[0] + 1
                st.info(f"Row: {rw} / {num_new_rows} added")
                st.session_state.df.loc[rw] = rwdta
    
                if st.session_state.df.shape[0] == num_new_rows:
                    st.error("Add row limit reached...")
    df = st.session_state.df
    df.to_csv("fleet 1/score9.csv",encoding='utf-8',index=False)
    st.text(df)
    
   with st.expander("Vessel profits"):
    if "df2" not in st.session_state:
        st.session_state.df2 = pd.DataFrame(columns=["Vessel 1: Arya", 
                                                    "Vessel 2: Cersei", 
                                                    "Vessel 3: Lyanna", 
                                                    "Vessel 4: Daenerys"])
    
    st.markdown("Add Record")
    
    ncol = st.session_state.df2.shape[1]  # col count
    rw = -1
    
    with st.form(key="add profits form", clear_on_submit= True):
        cols = st.columns(ncol)
        rwdta = []
    
        for i in range(ncol):
            rwdta.append(cols[i].text_input(st.session_state.df2.columns[i]))
    
        # you can insert code for a list comprehension here to change the data (rwdta) 
        # values into integer / float, if required
    
        if st.form_submit_button("Add"):
            if st.session_state.df2.shape[0] == num_new_rows:
                st.error("Add row limit reached. Cant add any more records..")
            else:
                rw = st.session_state.df2.shape[0] + 1
                st.info(f"Row: {rw} / {num_new_rows} added")
                st.session_state.df2.loc[rw] = rwdta
    
                if st.session_state.df2.shape[0] == num_new_rows:
                    st.error("Add row limit reached...")
    


    df2 = st.session_state.df2
    df2.to_csv("fleet 1/profits9.csv",encoding='utf-8',index=False)

    st.text(st.session_state.df2)
 
   with st.expander("Vessel fines"):
    if "df3" not in st.session_state:
        st.session_state.df3 = pd.DataFrame(columns=["Vessel 1: Arya", 
                                                    "Vessel 2: Cersei", 
                                                    "Vessel 3: Lyanna", 
                                                    "Vessel 4: Daenerys"])
    
    st.markdown("Add Record")
    
    ncol = st.session_state.df3.shape[1]  # col count
    rw = -1
    
    with st.form(key="add fines form", clear_on_submit= True):
        cols = st.columns(ncol)
        rwdta = []
    
        for i in range(ncol):
            rwdta.append(cols[i].text_input(st.session_state.df3.columns[i]))
    
        # you can insert code for a list comprehension here to change the data (rwdta) 
        # values into integer / float, if required
    
        if st.form_submit_button("Add"):
            if st.session_state.df3.shape[0] == num_new_rows:
                st.error("Add row limit reached. Cant add any more records..")
            else:
                rw = st.session_state.df3.shape[0] + 1
                st.info(f"Row: {rw} / {num_new_rows} added")
                st.session_state.df3.loc[rw] = rwdta
    
                if st.session_state.df3.shape[0] == num_new_rows:
                    st.error("Add row limit reached...")
    

    df3 = st.session_state.df3
    df3.to_csv("fleet 1/fines9.csv",encoding='utf-8',index=False)

    st.text(st.session_state.df3)

###############################################################################
#2.i Overall reports                                                          #
###############################################################################
   st.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)
   if st.button('Create overall reports'): 
    st.header("**Day 9: Overall reports**")
    st.markdown('''In this subsection the fleet manager can see the current accumulated score and profit of day 9 and the daily fleet scores and profits.''')
    st.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)
    #
    df = pd.read_csv(os.path.join(dir_main,'fleet 1/score9.csv'), sep=',')
    df2 = pd.read_csv(os.path.join(dir_main,'fleet 1/profits9.csv'), sep=',')
    df3 = pd.read_csv(os.path.join(dir_main,'fleet 1/fines9.csv'), sep=',')
    df4 = pd.read_excel(os.path.join(dir_main,'fleet 1/results7.xlsx'))
    df4.index = df4.Vessel
    df5 = pd.read_excel(os.path.join(dir_main,'fleet 1/results8.xlsx'))
    df5.index = df5.Vessel
    
    d = {'Vessel': df.columns[0:4] , 
         'Score':  df.iloc[-1] , 
         'Score t-1' : df4.Score,
         'Score t-2' : df5.Score,
         'Accumulated Profit':df2.iloc[-1] ,
         'Fines':df3.iloc[-1] }
    outcomes = pd.DataFrame(data=d)
    outcomes=outcomes.reset_index(drop=True)
    st.dataframe(outcomes)

    wb = openpyxl.Workbook()
    wb.save(os.path.join(dir_main,'fleet 1/reports9.xlsx')) 
    book = load_workbook(os.path.join(dir_main,'fleet 1/reports9.xlsx'))
    del book['Sheet']   

    writer = pd.ExcelWriter(os.path.join(dir_main,'fleet 1/reports9.xlsx'), engine='openpyxl')
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}
    outcomes.to_excel(writer, sheet_name = "Outcomes 9", na_rep = "", index =   False, startrow = 0)  
    writer.save()
    
    #Plot scores 
    df01 = pd.read_csv(os.path.join(dir_main,'fleet 1/score9.csv'), sep=',')
    df01.index = np.arange(1, len(df01) + 1)
    #Create the daily average 
    df01['Fleet score']=df01.sum(axis=1)/4
    df01['Action'] = df01.index
    df01 = pd.melt(df01, id_vars='Action', value_vars=["Vessel 1: Arya", 
                                                    "Vessel 2: Cersei", 
                                                    "Vessel 3: Lyanna", 
                                                    "Vessel 4: Daenerys", 
                                                    "Fleet score"])
    st.subheader("*Daily fleet and vessel scores*")
    fig01 = px.line(df01, x='Action', y='value', color='variable', markers=True,range_x=[0,10],width=650,height=650)
    fig01.update_layout(yaxis_type='category')
    fig01.layout.xaxis.fixedrange= True
    fig01.layout.yaxis.fixedrange: True
    fig01.update(layout_showlegend=True)
    fig01.update_traces(patch={"line": {"color": "green", "width": 3, "dash": 'dot'}}, selector={"legendgroup": "Fleet score"}) 
    st.plotly_chart(fig01)
  
    #Plot profits  
    df02 = pd.read_csv(os.path.join(dir_main,'fleet 1/profits9.csv'), sep=',')
    df03 = pd.read_csv(os.path.join(dir_main,'fleet 1/fines9.csv'), sep=',')
    df02.index = np.arange(1, len(df02) + 1)
    df03.index = np.arange(1, len(df03) + 1)

    df02['Action'] = df02.index
    df02['Fleet profit']=df02.sum(axis=1)
    aux = df01.loc[df01.variable=="Fleet score",'value'].reset_index(drop=True)
    aux.index = np.arange(1, len(aux) + 1)
    df02['Fleet score'] = aux
    
    df03['Fleet fines']=df03.sum(axis=1)
    df02['Fleet w. score'] = df02['Fleet profit'] - df03['Fleet fines']
    
    df02 = pd.melt(df02, id_vars='Action', value_vars=["Fleet profit",
                                                       "Fleet score",
                                                       "Fleet w. score"])
    df02_1 = df02.loc[df02.variable == "Fleet score",:]
    df02_2 = df02.loc[df02.variable != "Fleet score",:]
    
    # Create figure with secondary y-axis 
    subfig = make_subplots(specs=[[{"secondary_y": True}]])

    # Defining Custom Colors
    colours = {
      "Fleet profit": "#A1A9A3",
      "Fleet w. score": "#DFDF70",
    }

    # create two independent figures with px.line each containing data from multiple columns
    fig = px.line(df02_1, x='Action', y='value',markers=True,range_x=[1,10],width=650,height=650, text="value")
    fig.layout.xaxis.fixedrange= True
    fig.update_traces(line_color='orange', line_width=2)
    fig2 = px.bar(df02_2, y='value', x='Action', color='variable',range_x=[1,10],width=650,height=650,color_discrete_map=colours)
    fig.update_traces(yaxis="y2")
    fig2.update(layout_showlegend=True)
    
    subfig.add_traces(fig2.data + fig.data)
    subfig.layout.xaxis.fixedrange= True
    
    # Set y-axes titles
    subfig.layout.yaxis.title="Kr."
    subfig.layout.yaxis2.title="Score"
    subfig.update_xaxes(range=[0,10], dtick=1)
    st.subheader("*Fleet profit loss by score increase*")
    st.plotly_chart(subfig)
  
     #Save the data for next day 
    cav=open(dir_main+"fleet 1/reports9.xlsx",'rb')
    b64_01 = base64.b64encode(cav.read()).decode('utf-8')
    href = f'<a href="data:file/cav;base64,{b64_01}" download="results9.xlsx">:arrow_down_small: Overall results day 9 </a>'
    st.markdown(href, unsafe_allow_html=True)
  
  
###############################################################################
#2.i Day 10                                                                   #
###############################################################################
  if pag_input=="Day 10": 
    #Selection of section in the daily report
   # Title
   st.title(title)  
   st.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)
   st.header("**Day 10: Type in information**")
   st.markdown('''In this subsection the fleet manager types in the score and the profit for every action (this is already
accumulated)''')

   with st.expander("Vessel scores"):
#    if "df" not in st.session_state or score.columns[0]=="Unnamed: 0":
    if "df" not in st.session_state:
        st.session_state.df = pd.DataFrame(columns=["Vessel 1: Arya", 
                                                    "Vessel 2: Cersei", 
                                                    "Vessel 3: Lyanna", 
                                                    "Vessel 4: Daenerys"])


    st.markdown("Add Record")
    
    num_new_rows = 10
    ncol = st.session_state.df.shape[1]  # col count
    rw = -1
    
    with st.form(key="add form", clear_on_submit= True):
        cols = st.columns(ncol)
        rwdta = []
    
        for i in range(ncol):
            rwdta.append(cols[i].text_input(st.session_state.df.columns[i]))
    
        # you can insert code for a list comprehension here to change the data (rwdta) 
        # values into integer / float, if required
    
        if st.form_submit_button("Add"):
            if st.session_state.df.shape[0] == num_new_rows:
                st.error("Add row limit reached. Cant add any more records..")
            else:
                rw = st.session_state.df.shape[0] + 1
                st.info(f"Row: {rw} / {num_new_rows} added")
                st.session_state.df.loc[rw] = rwdta
    
                if st.session_state.df.shape[0] == num_new_rows:
                    st.error("Add row limit reached...")
    df = st.session_state.df
    df.to_csv("fleet 1/score10.csv",encoding='utf-8',index=False)
    st.text(df)
    
   with st.expander("Vessel profits"):
    if "df2" not in st.session_state:
        st.session_state.df2 = pd.DataFrame(columns=["Vessel 1: Arya", 
                                                    "Vessel 2: Cersei", 
                                                    "Vessel 3: Lyanna", 
                                                    "Vessel 4: Daenerys"])
    
    st.markdown("Add Record")
    
    ncol = st.session_state.df2.shape[1]  # col count
    rw = -1
    
    with st.form(key="add profits form", clear_on_submit= True):
        cols = st.columns(ncol)
        rwdta = []
    
        for i in range(ncol):
            rwdta.append(cols[i].text_input(st.session_state.df2.columns[i]))
    
        # you can insert code for a list comprehension here to change the data (rwdta) 
        # values into integer / float, if required
    
        if st.form_submit_button("Add"):
            if st.session_state.df2.shape[0] == num_new_rows:
                st.error("Add row limit reached. Cant add any more records..")
            else:
                rw = st.session_state.df2.shape[0] + 1
                st.info(f"Row: {rw} / {num_new_rows} added")
                st.session_state.df2.loc[rw] = rwdta
    
                if st.session_state.df2.shape[0] == num_new_rows:
                    st.error("Add row limit reached...")
    


    df2 = st.session_state.df2
    df2.to_csv("fleet 1/profits10.csv",encoding='utf-8',index=False)

    st.text(st.session_state.df2)
 
   with st.expander("Vessel fines"):
    if "df3" not in st.session_state:
        st.session_state.df3 = pd.DataFrame(columns=["Vessel 1: Arya", 
                                                    "Vessel 2: Cersei", 
                                                    "Vessel 3: Lyanna", 
                                                    "Vessel 4: Daenerys"])
    
    st.markdown("Add Record")
    
    ncol = st.session_state.df3.shape[1]  # col count
    rw = -1
    
    with st.form(key="add fines form", clear_on_submit= True):
        cols = st.columns(ncol)
        rwdta = []
    
        for i in range(ncol):
            rwdta.append(cols[i].text_input(st.session_state.df3.columns[i]))
    
        # you can insert code for a list comprehension here to change the data (rwdta) 
        # values into integer / float, if required
    
        if st.form_submit_button("Add"):
            if st.session_state.df3.shape[0] == num_new_rows:
                st.error("Add row limit reached. Cant add any more records..")
            else:
                rw = st.session_state.df3.shape[0] + 1
                st.info(f"Row: {rw} / {num_new_rows} added")
                st.session_state.df3.loc[rw] = rwdta
    
                if st.session_state.df3.shape[0] == num_new_rows:
                    st.error("Add row limit reached...")
    

    df3 = st.session_state.df3
    df3.to_csv("fleet 1/fines10.csv",encoding='utf-8',index=False)

    st.text(st.session_state.df3)

###############################################################################
#2.i Overall reports                                                          #
###############################################################################
   st.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)
   if st.button('Create overall reports'): 
    st.header("**Day 10: Overall reports**")
    st.markdown('''In this subsection the fleet manager can see the current accumulated score and profit of day 10 and the daily fleet scores and profits.''')
    st.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)
    #
    df = pd.read_csv(os.path.join(dir_main,'fleet 1/score10.csv'), sep=',')
    df2 = pd.read_csv(os.path.join(dir_main,'fleet 1/profits10.csv'), sep=',')
    df3 = pd.read_csv(os.path.join(dir_main,'fleet 1/fines10.csv'), sep=',')
    df4 = pd.read_excel(os.path.join(dir_main,'fleet 1/results8.xlsx'))
    df4.index = df4.Vessel
    df5 = pd.read_excel(os.path.join(dir_main,'fleet 1/results9.xlsx'))
    df5.index = df5.Vessel
    
    d = {'Vessel': df.columns[0:4] , 
         'Score':  df.iloc[-1] , 
         'Score t-1' : df4.Score,
         'Score t-2' : df5.Score,
         'Accumulated Profit':df2.iloc[-1] ,
         'Fines':df3.iloc[-1] }
    outcomes = pd.DataFrame(data=d)
    outcomes=outcomes.reset_index(drop=True)
    st.dataframe(outcomes)

    wb = openpyxl.Workbook()
    wb.save(os.path.join(dir_main,'fleet 1/reports10.xlsx')) 
    book = load_workbook(os.path.join(dir_main,'fleet 1/reports10.xlsx'))
    del book['Sheet']   

    writer = pd.ExcelWriter(os.path.join(dir_main,'fleet 1/reports10.xlsx'), engine='openpyxl')
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}
    outcomes.to_excel(writer, sheet_name = "Outcomes 10", na_rep = "", index =   False, startrow = 0)  
    writer.save()
    
    #Plot scores 
    df01 = pd.read_csv(os.path.join(dir_main,'fleet 1/score10.csv'), sep=',')
    df01.index = np.arange(1, len(df01) + 1)
    #Create the daily average 
    df01['Fleet score']=df01.sum(axis=1)/4
    df01['Action'] = df01.index
    df01 = pd.melt(df01, id_vars='Action', value_vars=["Vessel 1: Arya", 
                                                    "Vessel 2: Cersei", 
                                                    "Vessel 3: Lyanna", 
                                                    "Vessel 4: Daenerys", 
                                                    "Fleet score"])
    st.subheader("*Daily fleet and vessel scores*")
    fig01 = px.line(df01, x='Action', y='value', color='variable', markers=True,range_x=[0,10],width=650,height=650)
    fig01.update_layout(yaxis_type='category')
    fig01.layout.xaxis.fixedrange= True
    fig01.layout.yaxis.fixedrange: True
    fig01.update(layout_showlegend=True)
    fig01.update_traces(patch={"line": {"color": "green", "width": 3, "dash": 'dot'}}, selector={"legendgroup": "Fleet score"}) 
    st.plotly_chart(fig01)
  
    #Plot profits  
    df02 = pd.read_csv(os.path.join(dir_main,'fleet 1/profits10.csv'), sep=',')
    df03 = pd.read_csv(os.path.join(dir_main,'fleet 1/fines10.csv'), sep=',')
    df02.index = np.arange(1, len(df02) + 1)
    df03.index = np.arange(1, len(df03) + 1)

    df02['Action'] = df02.index
    df02['Fleet profit']=df02.sum(axis=1)
    aux = df01.loc[df01.variable=="Fleet score",'value'].reset_index(drop=True)
    aux.index = np.arange(1, len(aux) + 1)
    df02['Fleet score'] = aux
    
    df03['Fleet fines']=df03.sum(axis=1)
    df02['Fleet w. score'] = df02['Fleet profit'] - df03['Fleet fines']
    
    df02 = pd.melt(df02, id_vars='Action', value_vars=["Fleet profit",
                                                       "Fleet score",
                                                       "Fleet w. score"])
    df02_1 = df02.loc[df02.variable == "Fleet score",:]
    df02_2 = df02.loc[df02.variable != "Fleet score",:]
    
    # Create figure with secondary y-axis 
    subfig = make_subplots(specs=[[{"secondary_y": True}]])

    # Defining Custom Colors
    colours = {
      "Fleet profit": "#A1A9A3",
      "Fleet w. score": "#DFDF70",
    }

    # create two independent figures with px.line each containing data from multiple columns
    fig = px.line(df02_1, x='Action', y='value',markers=True,range_x=[1,10],width=650,height=650, text="value")
    fig.layout.xaxis.fixedrange= True
    fig.update_traces(line_color='orange', line_width=2)
    fig2 = px.bar(df02_2, y='value', x='Action', color='variable',range_x=[1,10],width=650,height=650,color_discrete_map=colours)
    fig.update_traces(yaxis="y2")
    fig2.update(layout_showlegend=True)
    
    subfig.add_traces(fig2.data + fig.data)
    subfig.layout.xaxis.fixedrange= True
    
    # Set y-axes titles
    subfig.layout.yaxis.title="Kr."
    subfig.layout.yaxis2.title="Score"
    subfig.update_xaxes(range=[0,10], dtick=1)
    st.subheader("*Fleet profit loss by score increase*")
    st.plotly_chart(subfig)
  
     #Save the data for next day 
    cav=open(dir_main+"fleet 1/reports10.xlsx",'rb')
    b64_01 = base64.b64encode(cav.read()).decode('utf-8')
    href = f'<a href="data:file/cav;base64,{b64_01}" download="results10.xlsx">:arrow_down_small: Overall results day 10 </a>'
    st.markdown(href, unsafe_allow_html=True)
  
  
###############################################################################
#A. Fleets                                                                    #
###############################################################################
if user == "fleet2": 
    
###############################################################################
#2. Section selection                                                         #
###############################################################################
  section= ["Day 1", "Day 2", 'Day 3','Day 4','Day 5','Day 6','Day 7','Day 8','Day 9','Day 10']
  
  st.sidebar.title("Section selection")
  st.sidebar.markdown('''Select the day you wish to see.''')
  
  #Selection of daily report 
  pag_input = st.sidebar.selectbox('Select day:', section)

  st.sidebar.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)

###############################################################################
#2.i Day 1                                                                    #
###############################################################################
  if pag_input=="Day 1": 
    #Selection of section in the daily report
   # Title
   st.title(title)  
   st.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)
   st.header("**Day 1: Type in information**")
   st.markdown('''In this subsection the fleet manager types in the score and the profit for every action (this is already
accumulated)''')

   with st.expander("Vessel scores"):
#    if "df" not in st.session_state or score.columns[0]=="Unnamed: 0":
    if "df" not in st.session_state:
        st.session_state.df = pd.DataFrame(columns=["Vessel 1: Sansa", 
                                                    "Vessel 2: Olenna", 
                                                    "Vessel 3: Missandei"])


    st.markdown("Add Record")
    
    num_new_rows = 10
    ncol = st.session_state.df.shape[1]  # col count
    rw = -1
    
    with st.form(key="add form", clear_on_submit= True):
        cols = st.columns(ncol)
        rwdta = []
    
        for i in range(ncol):
            rwdta.append(cols[i].text_input(st.session_state.df.columns[i]))
    
        # you can insert code for a list comprehension here to change the data (rwdta) 
        # values into integer / float, if required
    
        if st.form_submit_button("Add"):
            if st.session_state.df.shape[0] == num_new_rows:
                st.error("Add row limit reached. Cant add any more records..")
            else:
                rw = st.session_state.df.shape[0] + 1
                st.info(f"Row: {rw} / {num_new_rows} added")
                st.session_state.df.loc[rw] = rwdta
    
                if st.session_state.df.shape[0] == num_new_rows:
                    st.error("Add row limit reached...")
    df = st.session_state.df
    df.to_csv("fleet 2/score1.csv",encoding='utf-8',index=False)
    st.text(df)
    
   with st.expander("Vessel profits"):
    if "df2" not in st.session_state:
        st.session_state.df2 = pd.DataFrame(columns=["Vessel 1: Sansa", 
                                                    "Vessel 2: Olenna", 
                                                    "Vessel 3: Missandei"])
    
    st.markdown("Add Record")
    
    ncol = st.session_state.df2.shape[1]  # col count
    rw = -1
    
    with st.form(key="add profits form", clear_on_submit= True):
        cols = st.columns(ncol)
        rwdta = []
    
        for i in range(ncol):
            rwdta.append(cols[i].text_input(st.session_state.df2.columns[i]))
    
        # you can insert code for a list comprehension here to change the data (rwdta) 
        # values into integer / float, if required
    
        if st.form_submit_button("Add"):
            if st.session_state.df2.shape[0] == num_new_rows:
                st.error("Add row limit reached. Cant add any more records..")
            else:
                rw = st.session_state.df2.shape[0] + 1
                st.info(f"Row: {rw} / {num_new_rows} added")
                st.session_state.df2.loc[rw] = rwdta
    
                if st.session_state.df2.shape[0] == num_new_rows:
                    st.error("Add row limit reached...")
    


    df2 = st.session_state.df2
    df2.to_csv("fleet 2/profits1.csv",encoding='utf-8',index=False)

    st.text(st.session_state.df2)
 
   with st.expander("Vessel fines"):
    if "df3" not in st.session_state:
        st.session_state.df3 = pd.DataFrame(columns=["Vessel 1: Sansa", 
                                                    "Vessel 2: Olenna", 
                                                    "Vessel 3: Missandei"])
    
    st.markdown("Add Record")
    
    ncol = st.session_state.df3.shape[1]  # col count
    rw = -1
    
    with st.form(key="add fines form", clear_on_submit= True):
        cols = st.columns(ncol)
        rwdta = []
    
        for i in range(ncol):
            rwdta.append(cols[i].text_input(st.session_state.df3.columns[i]))
    
        # you can insert code for a list comprehension here to change the data (rwdta) 
        # values into integer / float, if required
    
        if st.form_submit_button("Add"):
            if st.session_state.df3.shape[0] == num_new_rows:
                st.error("Add row limit reached. Cant add any more records..")
            else:
                rw = st.session_state.df3.shape[0] + 1
                st.info(f"Row: {rw} / {num_new_rows} added")
                st.session_state.df3.loc[rw] = rwdta
    
                if st.session_state.df3.shape[0] == num_new_rows:
                    st.error("Add row limit reached...")
    

    df3 = st.session_state.df3
    df3.to_csv("fleet 2/fines1.csv",encoding='utf-8',index=False)

    st.text(st.session_state.df3)

###############################################################################
#2.i Overall reports                                                          #
###############################################################################
   st.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)
   if st.button('Create overall reports'): 
    st.header("**Day 1: Overall reports**")
    st.markdown('''In this subsection the fleet manager can see the current accumulted score and profit of day 1 and the daily fleet scores and profits.''')
    st.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)
    #
    df = pd.read_csv(os.path.join(dir_main,'fleet 2/score1.csv'), sep=',')
    df2 = pd.read_csv(os.path.join(dir_main,'fleet 2/profits1.csv'), sep=',')
    df3 = pd.read_csv(os.path.join(dir_main,'fleet 2/fines1.csv'), sep=',')
    d = {'Vessel': df.columns[0:3] , 'Score':  df.iloc[-1] , 'Accumulated Profit':df2.iloc[-1] ,'Fines':df3.iloc[-1] }
    outcomes = pd.DataFrame(data=d)
    outcomes=outcomes.reset_index(drop=True)
    st.dataframe(outcomes)
    
    wb = openpyxl.Workbook()
    wb.save(os.path.join(dir_main,'fleet 2/reports1.xlsx')) 
    book = load_workbook(os.path.join(dir_main,'fleet 2/reports1.xlsx'))
    del book['Sheet']   

    writer = pd.ExcelWriter(os.path.join(dir_main,'fleet 2/reports1.xlsx'), engine='openpyxl')
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}
    outcomes.to_excel(writer, sheet_name = "Outcomes 1", na_rep = "", index =   False, startrow = 0)  
    writer.save()

    #Plot scores 
    df01 = pd.read_csv(os.path.join(dir_main,'fleet 2/score1.csv'), sep=',')
    df01.index = np.arange(1, len(df01) + 1)
    #Create the daily average 
    df01['Fleet score']=df01.sum(axis=1)/3
    df01['Action'] = df01.index
    df01 = pd.melt(df01, id_vars='Action', value_vars=["Vessel 1: Sansa", 
                                                    "Vessel 2: Olenna", 
                                                    "Vessel 3: Missandei", 
                                                    "Fleet score"])
    st.subheader("*Daily fleet and vessel scores*")
    fig01 = px.line(df01, x='Action', y='value', color='variable', markers=True,range_x=[0,10],width=650,height=650)
    fig01.update_layout(yaxis_type='category')
    fig01.layout.xaxis.fixedrange= True
    fig01.layout.yaxis.fixedrange: True
    fig01.update(layout_showlegend=True)
    fig01.update_traces(patch={"line": {"color": "green", "width": 3, "dash": 'dot'}}, selector={"legendgroup": "Fleet score"}) 
    st.plotly_chart(fig01)
  
    #Plot profits  
    df02 = pd.read_csv(os.path.join(dir_main,'fleet 2/profits1.csv'), sep=',')
    df03 = pd.read_csv(os.path.join(dir_main,'fleet 2/fines1.csv'), sep=',')
    df02.index = np.arange(1, len(df02) + 1)
    df03.index = np.arange(1, len(df03) + 1)

    df02['Action'] = df02.index
    df02['Fleet profit']=df02.sum(axis=1)
    aux = df01.loc[df01.variable=="Fleet score",'value'].reset_index(drop=True)
    aux.index = np.arange(1, len(aux) + 1)
    df02['Fleet score'] = aux
    
    df03['Fleet fines']=df03.sum(axis=1)
    df02['Fleet w. score'] = df02['Fleet profit'] - df03['Fleet fines']
    
    df02 = pd.melt(df02, id_vars='Action', value_vars=["Fleet profit",
                                                       "Fleet score",
                                                       "Fleet w. score"])
    df02_1 = df02.loc[df02.variable == "Fleet score",:]
    df02_2 = df02.loc[df02.variable != "Fleet score",:]
    
    # Create figure with secondary y-axis 
    subfig = make_subplots(specs=[[{"secondary_y": True}]])

    # Defining Custom Colors
    colours = {
      "Fleet profit": "#A1A9A3",
      "Fleet w. score": "#DFDF70",
    }

    # create two independent figures with px.line each containing data from multiple columns
    fig = px.line(df02_1, x='Action', y='value',markers=True,range_x=[1,10],width=650,height=650, text="value")
    fig.layout.xaxis.fixedrange= True
    fig.update_traces(line_color='orange', line_width=2)
    fig2 = px.bar(df02_2, y='value', x='Action', color='variable',range_x=[1,10],width=650,height=650,color_discrete_map=colours)
    fig.update_traces(yaxis="y2")
    fig2.update(layout_showlegend=True)
    
    subfig.add_traces(fig2.data + fig.data)
    subfig.layout.xaxis.fixedrange= True
    
    # Set y-axes titles
    subfig.layout.yaxis.title="Kr."
    subfig.layout.yaxis2.title="Score"
    subfig.update_xaxes(range=[0,10], dtick=1)
    st.subheader("*Fleet profit loss by score increase*")
    st.plotly_chart(subfig)
  
    #Save the data for next day 
    cav=open(dir_main+"fleet 2/reports1.xlsx",'rb')
    b64_01 = base64.b64encode(cav.read()).decode('utf-8')
    href = f'<a href="data:file/cav;base64,{b64_01}" download="results1.xlsx">:arrow_down_small: Overall results day 1 </a>'
    st.markdown(href, unsafe_allow_html=True)

###############################################################################
#2.i Day 2                                                                    #
###############################################################################
  if pag_input=="Day 2": 
    #Selection of section in the daily report
   # Title
   st.title(title)  
   st.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)
   st.header("**Day 2: Type in information**")
   st.markdown('''In this subsection the fleet manager types in the score and the profit for every action (this is already
accumulated)''')

   with st.expander("Vessel scores"):
#    if "df" not in st.session_state or score.columns[0]=="Unnamed: 0":
    if "df" not in st.session_state:
        st.session_state.df = pd.DataFrame(columns=["Vessel 1: Sansa", 
                                                    "Vessel 2: Olenna", 
                                                    "Vessel 3: Missandei"])


    st.markdown("Add Record")
    
    num_new_rows = 10
    ncol = st.session_state.df.shape[1]  # col count
    rw = -1
    
    with st.form(key="add form", clear_on_submit= True):
        cols = st.columns(ncol)
        rwdta = []
    
        for i in range(ncol):
            rwdta.append(cols[i].text_input(st.session_state.df.columns[i]))
    
        # you can insert code for a list comprehension here to change the data (rwdta) 
        # values into integer / float, if required
    
        if st.form_submit_button("Add"):
            if st.session_state.df.shape[0] == num_new_rows:
                st.error("Add row limit reached. Cant add any more records..")
            else:
                rw = st.session_state.df.shape[0] + 1
                st.info(f"Row: {rw} / {num_new_rows} added")
                st.session_state.df.loc[rw] = rwdta
    
                if st.session_state.df.shape[0] == num_new_rows:
                    st.error("Add row limit reached...")
    df = st.session_state.df
    df.to_csv("fleet 2/score2.csv",encoding='utf-8',index=False)
    st.text(df)
    
   with st.expander("Vessel profits"):
    if "df2" not in st.session_state:
        st.session_state.df2 = pd.DataFrame(columns=["Vessel 1: Sansa", 
                                                    "Vessel 2: Olenna", 
                                                    "Vessel 3: Missandei"])
    
    st.markdown("Add Record")
    
    ncol = st.session_state.df2.shape[1]  # col count
    rw = -1
    
    with st.form(key="add profits form", clear_on_submit= True):
        cols = st.columns(ncol)
        rwdta = []
    
        for i in range(ncol):
            rwdta.append(cols[i].text_input(st.session_state.df2.columns[i]))
    
        # you can insert code for a list comprehension here to change the data (rwdta) 
        # values into integer / float, if required
    
        if st.form_submit_button("Add"):
            if st.session_state.df2.shape[0] == num_new_rows:
                st.error("Add row limit reached. Cant add any more records..")
            else:
                rw = st.session_state.df2.shape[0] + 1
                st.info(f"Row: {rw} / {num_new_rows} added")
                st.session_state.df2.loc[rw] = rwdta
    
                if st.session_state.df2.shape[0] == num_new_rows:
                    st.error("Add row limit reached...")
    


    df2 = st.session_state.df2
    df2.to_csv("fleet 2/profits2.csv",encoding='utf-8',index=False)

    st.text(st.session_state.df2)
 
   with st.expander("Vessel fines"):
    if "df3" not in st.session_state:
        st.session_state.df3 = pd.DataFrame(columns=["Vessel 1: Sansa", 
                                                    "Vessel 2: Olenna", 
                                                    "Vessel 3: Missandei"])
    
    st.markdown("Add Record")
    
    ncol = st.session_state.df3.shape[1]  # col count
    rw = -1
    
    with st.form(key="add fines form", clear_on_submit= True):
        cols = st.columns(ncol)
        rwdta = []
    
        for i in range(ncol):
            rwdta.append(cols[i].text_input(st.session_state.df3.columns[i]))
    
        # you can insert code for a list comprehension here to change the data (rwdta) 
        # values into integer / float, if required
    
        if st.form_submit_button("Add"):
            if st.session_state.df3.shape[0] == num_new_rows:
                st.error("Add row limit reached. Cant add any more records..")
            else:
                rw = st.session_state.df3.shape[0] + 1
                st.info(f"Row: {rw} / {num_new_rows} added")
                st.session_state.df3.loc[rw] = rwdta
    
                if st.session_state.df3.shape[0] == num_new_rows:
                    st.error("Add row limit reached...")
    

    df3 = st.session_state.df3
    df3.to_csv("fleet 2/fines2.csv",encoding='utf-8',index=False)

    st.text(st.session_state.df3)

###############################################################################
#2.i Overall reports                                                          #
###############################################################################
   st.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)
   if st.button('Create overall reports'): 
    st.header("**Day 2: Overall reports**")
    st.markdown('''In this subsection the fleet manager can see the current accumulated score and profit of day 2 and the daily fleet scores and profits.''')
    st.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)
    #
    df = pd.read_csv(os.path.join(dir_main,'fleet 2/score2.csv'), sep=',')
    df2 = pd.read_csv(os.path.join(dir_main,'fleet 2/profits2.csv'), sep=',')
    df3 = pd.read_csv(os.path.join(dir_main,'fleet 2/fines2.csv'), sep=',')
    df4 = pd.read_excel(os.path.join(dir_main,'fleet 2/results1.xlsx'))
    df4.index = df4.Vessel
    
    d = {'Vessel': df.columns[0:3] , 
         'Score':  df.iloc[-1] , 
         'Score t-1' : df4.Score,
         'Accumulated Profit':df2.iloc[-1] ,
         'Fines':df3.iloc[-1] }
    outcomes = pd.DataFrame(data=d)
    outcomes=outcomes.reset_index(drop=True)
    st.dataframe(outcomes)

    wb = openpyxl.Workbook()
    wb.save(os.path.join(dir_main,'fleet 2/reports2.xlsx')) 
    book = load_workbook(os.path.join(dir_main,'fleet 2/reports2.xlsx'))
    del book['Sheet']   

    writer = pd.ExcelWriter(os.path.join(dir_main,'fleet 2/reports2.xlsx'), engine='openpyxl')
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}
    outcomes.to_excel(writer, sheet_name = "Outcomes 2", na_rep = "", index =   False, startrow = 0)  
    writer.save()
    
    #Plot scores 
    df01 = pd.read_csv(os.path.join(dir_main,'fleet 2/score2.csv'), sep=',')
    df01.index = np.arange(1, len(df01) + 1)
    #Create the daily average 
    df01['Fleet score']=df01.sum(axis=1)/3
    df01['Action'] = df01.index
    df01 = pd.melt(df01, id_vars='Action', value_vars=["Vessel 1: Sansa", 
                                                    "Vessel 2: Olenna", 
                                                    "Vessel 3: Missandei", 
                                                    "Fleet score"])
    st.subheader("*Daily fleet and vessel scores*")
    fig01 = px.line(df01, x='Action', y='value', color='variable', markers=True,range_x=[0,10],width=650,height=650)
    fig01.update_layout(yaxis_type='category')
    fig01.layout.xaxis.fixedrange= True
    fig01.layout.yaxis.fixedrange: True
    fig01.update(layout_showlegend=True)
    fig01.update_traces(patch={"line": {"color": "green", "width": 3, "dash": 'dot'}}, selector={"legendgroup": "Fleet score"}) 
    st.plotly_chart(fig01)
  
    #Plot profits  
    df02 = pd.read_csv(os.path.join(dir_main,'fleet 2/profits2.csv'), sep=',')
    df03 = pd.read_csv(os.path.join(dir_main,'fleet 2/fines2.csv'), sep=',')
    df02.index = np.arange(1, len(df02) + 1)
    df03.index = np.arange(1, len(df03) + 1)

    df02['Action'] = df02.index
    df02['Fleet profit']=df02.sum(axis=1)
    aux = df01.loc[df01.variable=="Fleet score",'value'].reset_index(drop=True)
    aux.index = np.arange(1, len(aux) + 1)
    df02['Fleet score'] = aux
    
    df03['Fleet fines']=df03.sum(axis=1)
    df02['Fleet w. score'] = df02['Fleet profit'] - df03['Fleet fines']
    
    df02 = pd.melt(df02, id_vars='Action', value_vars=["Fleet profit",
                                                       "Fleet score",
                                                       "Fleet w. score"])
    df02_1 = df02.loc[df02.variable == "Fleet score",:]
    df02_2 = df02.loc[df02.variable != "Fleet score",:]
    
    # Create figure with secondary y-axis 
    subfig = make_subplots(specs=[[{"secondary_y": True}]])

    # Defining Custom Colors
    colours = {
      "Fleet profit": "#A1A9A3",
      "Fleet w. score": "#DFDF70",
    }

    # create two independent figures with px.line each containing data from multiple columns
    fig = px.line(df02_1, x='Action', y='value',markers=True,range_x=[1,10],width=650,height=650, text="value")
    fig.layout.xaxis.fixedrange= True
    fig.update_traces(line_color='orange', line_width=2)
    fig2 = px.bar(df02_2, y='value', x='Action', color='variable',range_x=[1,10],width=650,height=650,color_discrete_map=colours)
    fig.update_traces(yaxis="y2")
    fig2.update(layout_showlegend=True)
    
    subfig.add_traces(fig2.data + fig.data)
    subfig.layout.xaxis.fixedrange= True
    
    # Set y-axes titles
    subfig.layout.yaxis.title="Kr."
    subfig.layout.yaxis2.title="Score"
    subfig.update_xaxes(range=[0,10], dtick=1)
    st.subheader("*Fleet profit loss by score increase*")
    st.plotly_chart(subfig)
  
     #Save the data for next day 
    cav=open(dir_main+"fleet 2/reports2.xlsx",'rb')
    b64_01 = base64.b64encode(cav.read()).decode('utf-8')
    href = f'<a href="data:file/cav;base64,{b64_01}" download="results2.xlsx">:arrow_down_small: Overall results day 2 </a>'
    st.markdown(href, unsafe_allow_html=True)
 
###############################################################################
#2.i Day 3                                                                    #
###############################################################################
  if pag_input=="Day 3": 
    #Selection of section in the daily report
   # Title
   st.title(title)  
   st.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)
   st.header("**Day 3: Type in information**")
   st.markdown('''In this subsection the fleet manager types in the score and the profit for every action (this is already
accumulated)''')

   with st.expander("Vessel scores"):
#    if "df" not in st.session_state or score.columns[0]=="Unnamed: 0":
    if "df" not in st.session_state:
        st.session_state.df = pd.DataFrame(columns=["Vessel 1: Sansa", 
                                                    "Vessel 2: Olenna", 
                                                    "Vessel 3: Missandei"])


    st.markdown("Add Record")
    
    num_new_rows = 10
    ncol = st.session_state.df.shape[1]  # col count
    rw = -1
    
    with st.form(key="add form", clear_on_submit= True):
        cols = st.columns(ncol)
        rwdta = []
    
        for i in range(ncol):
            rwdta.append(cols[i].text_input(st.session_state.df.columns[i]))
    
        # you can insert code for a list comprehension here to change the data (rwdta) 
        # values into integer / float, if required
    
        if st.form_submit_button("Add"):
            if st.session_state.df.shape[0] == num_new_rows:
                st.error("Add row limit reached. Cant add any more records..")
            else:
                rw = st.session_state.df.shape[0] + 1
                st.info(f"Row: {rw} / {num_new_rows} added")
                st.session_state.df.loc[rw] = rwdta
    
                if st.session_state.df.shape[0] == num_new_rows:
                    st.error("Add row limit reached...")
    df = st.session_state.df
    df.to_csv("fleet 2/score3.csv",encoding='utf-8',index=False)
    st.text(df)
    
   with st.expander("Vessel profits"):
    if "df2" not in st.session_state:
        st.session_state.df2 = pd.DataFrame(columns=["Vessel 1: Sansa", 
                                                    "Vessel 2: Olenna", 
                                                    "Vessel 3: Missandei"])
    
    st.markdown("Add Record")
    
    ncol = st.session_state.df2.shape[1]  # col count
    rw = -1
    
    with st.form(key="add profits form", clear_on_submit= True):
        cols = st.columns(ncol)
        rwdta = []
    
        for i in range(ncol):
            rwdta.append(cols[i].text_input(st.session_state.df2.columns[i]))
    
        # you can insert code for a list comprehension here to change the data (rwdta) 
        # values into integer / float, if required
    
        if st.form_submit_button("Add"):
            if st.session_state.df2.shape[0] == num_new_rows:
                st.error("Add row limit reached. Cant add any more records..")
            else:
                rw = st.session_state.df2.shape[0] + 1
                st.info(f"Row: {rw} / {num_new_rows} added")
                st.session_state.df2.loc[rw] = rwdta
    
                if st.session_state.df2.shape[0] == num_new_rows:
                    st.error("Add row limit reached...")
    


    df2 = st.session_state.df2
    df2.to_csv("fleet 2/profits3.csv",encoding='utf-8',index=False)

    st.text(st.session_state.df2)
 
   with st.expander("Vessel fines"):
    if "df3" not in st.session_state:
        st.session_state.df3 = pd.DataFrame(columns=["Vessel 1: Sansa", 
                                                    "Vessel 2: Olenna", 
                                                    "Vessel 3: Missandei"])
    
    st.markdown("Add Record")
    
    ncol = st.session_state.df3.shape[1]  # col count
    rw = -1
    
    with st.form(key="add fines form", clear_on_submit= True):
        cols = st.columns(ncol)
        rwdta = []
    
        for i in range(ncol):
            rwdta.append(cols[i].text_input(st.session_state.df3.columns[i]))
    
        # you can insert code for a list comprehension here to change the data (rwdta) 
        # values into integer / float, if required
    
        if st.form_submit_button("Add"):
            if st.session_state.df3.shape[0] == num_new_rows:
                st.error("Add row limit reached. Cant add any more records..")
            else:
                rw = st.session_state.df3.shape[0] + 1
                st.info(f"Row: {rw} / {num_new_rows} added")
                st.session_state.df3.loc[rw] = rwdta
    
                if st.session_state.df3.shape[0] == num_new_rows:
                    st.error("Add row limit reached...")
    

    df3 = st.session_state.df3
    df3.to_csv("fleet 2/fines3.csv",encoding='utf-8',index=False)

    st.text(st.session_state.df3)

###############################################################################
#2.i Overall reports                                                          #
###############################################################################
   st.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)
   if st.button('Create overall reports'): 
    st.header("**Day 3: Overall reports**")
    st.markdown('''In this subsection the fleet manager can see the current accumulated score and profit of day 3 and the daily fleet scores and profits.''')
    st.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)
    #
    df = pd.read_csv(os.path.join(dir_main,'fleet 2/score3.csv'), sep=',')
    df2 = pd.read_csv(os.path.join(dir_main,'fleet 2/profits3.csv'), sep=',')
    df3 = pd.read_csv(os.path.join(dir_main,'fleet 2/fines3.csv'), sep=',')
    df4 = pd.read_excel(os.path.join(dir_main,'fleet 2/results2.xlsx'))
    df4.index = df4.Vessel
    df5 = pd.read_excel(os.path.join(dir_main,'fleet 2/results1.xlsx'))
    df5.index = df5.Vessel
    
    d = {'Vessel': df.columns[0:3] , 
         'Score':  df.iloc[-1] , 
         'Score t-1' : df4.Score,
         'Score t-2' : df5.Score,
         'Accumulated Profit':df2.iloc[-1] ,
         'Fines':df3.iloc[-1] }
    outcomes = pd.DataFrame(data=d)
    outcomes=outcomes.reset_index(drop=True)
    st.dataframe(outcomes)

    wb = openpyxl.Workbook()
    wb.save(os.path.join(dir_main,'fleet 2/reports3.xlsx')) 
    book = load_workbook(os.path.join(dir_main,'fleet 2/reports3.xlsx'))
    del book['Sheet']   

    writer = pd.ExcelWriter(os.path.join(dir_main,'fleet 2/reports3.xlsx'), engine='openpyxl')
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}
    outcomes.to_excel(writer, sheet_name = "Outcomes 3", na_rep = "", index =   False, startrow = 0)  
    writer.save()
    
    #Plot scores 
    df01 = pd.read_csv(os.path.join(dir_main,'fleet 2/score3.csv'), sep=',')
    df01.index = np.arange(1, len(df01) + 1)
    #Create the daily average 
    df01['Fleet score']=df01.sum(axis=1)/3
    df01['Action'] = df01.index
    df01 = pd.melt(df01, id_vars='Action', value_vars=["Vessel 1: Sansa", 
                                                    "Vessel 2: Olenna", 
                                                    "Vessel 3: Missandei", 
                                                    "Fleet score"])
    st.subheader("*Daily fleet and vessel scores*")
    fig01 = px.line(df01, x='Action', y='value', color='variable', markers=True,range_x=[0,10],width=650,height=650)
    fig01.update_layout(yaxis_type='category')
    fig01.layout.xaxis.fixedrange= True
    fig01.layout.yaxis.fixedrange: True
    fig01.update(layout_showlegend=True)
    fig01.update_traces(patch={"line": {"color": "green", "width": 3, "dash": 'dot'}}, selector={"legendgroup": "Fleet score"}) 
    st.plotly_chart(fig01)
  
    #Plot profits  
    df02 = pd.read_csv(os.path.join(dir_main,'fleet 2/profits3.csv'), sep=',')
    df03 = pd.read_csv(os.path.join(dir_main,'fleet 2/fines3.csv'), sep=',')
    df02.index = np.arange(1, len(df02) + 1)
    df03.index = np.arange(1, len(df03) + 1)

    df02['Action'] = df02.index
    df02['Fleet profit']=df02.sum(axis=1)
    aux = df01.loc[df01.variable=="Fleet score",'value'].reset_index(drop=True)
    aux.index = np.arange(1, len(aux) + 1)
    df02['Fleet score'] = aux
    
    df03['Fleet fines']=df03.sum(axis=1)
    df02['Fleet w. score'] = df02['Fleet profit'] - df03['Fleet fines']
    
    df02 = pd.melt(df02, id_vars='Action', value_vars=["Fleet profit",
                                                       "Fleet score",
                                                       "Fleet w. score"])
    df02_1 = df02.loc[df02.variable == "Fleet score",:]
    df02_2 = df02.loc[df02.variable != "Fleet score",:]
    
    # Create figure with secondary y-axis 
    subfig = make_subplots(specs=[[{"secondary_y": True}]])

    # Defining Custom Colors
    colours = {
      "Fleet profit": "#A1A9A3",
      "Fleet w. score": "#DFDF70",
    }

    # create two independent figures with px.line each containing data from multiple columns
    fig = px.line(df02_1, x='Action', y='value',markers=True,range_x=[1,10],width=650,height=650, text="value")
    fig.layout.xaxis.fixedrange= True
    fig.update_traces(line_color='orange', line_width=2)
    fig2 = px.bar(df02_2, y='value', x='Action', color='variable',range_x=[1,10],width=650,height=650,color_discrete_map=colours)
    fig.update_traces(yaxis="y2")
    fig2.update(layout_showlegend=True)
    
    subfig.add_traces(fig2.data + fig.data)
    subfig.layout.xaxis.fixedrange= True
    
    # Set y-axes titles
    subfig.layout.yaxis.title="Kr."
    subfig.layout.yaxis2.title="Score"
    subfig.update_xaxes(range=[0,10], dtick=1)
    st.subheader("*Fleet profit loss by score increase*")
    st.plotly_chart(subfig)
  
     #Save the data for next day 
    cav=open(dir_main+"fleet 2/reports3.xlsx",'rb')
    b64_01 = base64.b64encode(cav.read()).decode('utf-8')
    href = f'<a href="data:file/cav;base64,{b64_01}" download="results3.xlsx">:arrow_down_small: Overall results day 3 </a>'
    st.markdown(href, unsafe_allow_html=True)
  
###############################################################################
#2.i Day 4                                                                    #
###############################################################################
  if pag_input=="Day 4": 
    #Selection of section in the daily report
   # Title
   st.title(title)  
   st.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)
   st.header("**Day 4: Type in information**")
   st.markdown('''In this subsection the fleet manager types in the score and the profit for every action (this is already
accumulated)''')

   with st.expander("Vessel scores"):
#    if "df" not in st.session_state or score.columns[0]=="Unnamed: 0":
    if "df" not in st.session_state:
        st.session_state.df = pd.DataFrame(columns=["Vessel 1: Sansa", 
                                                    "Vessel 2: Olenna", 
                                                    "Vessel 3: Missandei"])


    st.markdown("Add Record")
    
    num_new_rows = 10
    ncol = st.session_state.df.shape[1]  # col count
    rw = -1
    
    with st.form(key="add form", clear_on_submit= True):
        cols = st.columns(ncol)
        rwdta = []
    
        for i in range(ncol):
            rwdta.append(cols[i].text_input(st.session_state.df.columns[i]))
    
        # you can insert code for a list comprehension here to change the data (rwdta) 
        # values into integer / float, if required
    
        if st.form_submit_button("Add"):
            if st.session_state.df.shape[0] == num_new_rows:
                st.error("Add row limit reached. Cant add any more records..")
            else:
                rw = st.session_state.df.shape[0] + 1
                st.info(f"Row: {rw} / {num_new_rows} added")
                st.session_state.df.loc[rw] = rwdta
    
                if st.session_state.df.shape[0] == num_new_rows:
                    st.error("Add row limit reached...")
    df = st.session_state.df
    df.to_csv("fleet 2/score4.csv",encoding='utf-8',index=False)
    st.text(df)
    
   with st.expander("Vessel profits"):
    if "df2" not in st.session_state:
        st.session_state.df2 = pd.DataFrame(columns=["Vessel 1: Sansa", 
                                                    "Vessel 2: Olenna", 
                                                    "Vessel 3: Missandei"])
    
    st.markdown("Add Record")
    
    ncol = st.session_state.df2.shape[1]  # col count
    rw = -1
    
    with st.form(key="add profits form", clear_on_submit= True):
        cols = st.columns(ncol)
        rwdta = []
    
        for i in range(ncol):
            rwdta.append(cols[i].text_input(st.session_state.df2.columns[i]))
    
        # you can insert code for a list comprehension here to change the data (rwdta) 
        # values into integer / float, if required
    
        if st.form_submit_button("Add"):
            if st.session_state.df2.shape[0] == num_new_rows:
                st.error("Add row limit reached. Cant add any more records..")
            else:
                rw = st.session_state.df2.shape[0] + 1
                st.info(f"Row: {rw} / {num_new_rows} added")
                st.session_state.df2.loc[rw] = rwdta
    
                if st.session_state.df2.shape[0] == num_new_rows:
                    st.error("Add row limit reached...")
    


    df2 = st.session_state.df2
    df2.to_csv("fleet 2/profits4.csv",encoding='utf-8',index=False)

    st.text(st.session_state.df2)
 
   with st.expander("Vessel fines"):
    if "df3" not in st.session_state:
        st.session_state.df3 = pd.DataFrame(columns=["Vessel 1: Sansa", 
                                                    "Vessel 2: Olenna", 
                                                    "Vessel 3: Missandei"])
    
    st.markdown("Add Record")
    
    ncol = st.session_state.df3.shape[1]  # col count
    rw = -1
    
    with st.form(key="add fines form", clear_on_submit= True):
        cols = st.columns(ncol)
        rwdta = []
    
        for i in range(ncol):
            rwdta.append(cols[i].text_input(st.session_state.df3.columns[i]))
    
        # you can insert code for a list comprehension here to change the data (rwdta) 
        # values into integer / float, if required
    
        if st.form_submit_button("Add"):
            if st.session_state.df3.shape[0] == num_new_rows:
                st.error("Add row limit reached. Cant add any more records..")
            else:
                rw = st.session_state.df3.shape[0] + 1
                st.info(f"Row: {rw} / {num_new_rows} added")
                st.session_state.df3.loc[rw] = rwdta
    
                if st.session_state.df3.shape[0] == num_new_rows:
                    st.error("Add row limit reached...")
    

    df3 = st.session_state.df3
    df3.to_csv("fleet 2/fines4.csv",encoding='utf-8',index=False)

    st.text(st.session_state.df3)

###############################################################################
#2.i Overall reports                                                          #
###############################################################################
   st.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)
   if st.button('Create overall reports'): 
    st.header("**Day 4: Overall reports**")
    st.markdown('''In this subsection the fleet manager can see the current accumulated score and profit of day 4 and the daily fleet scores and profits.''')
    st.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)
    #
    df = pd.read_csv(os.path.join(dir_main,'fleet 2/score4.csv'), sep=',')
    df2 = pd.read_csv(os.path.join(dir_main,'fleet 2/profits4.csv'), sep=',')
    df3 = pd.read_csv(os.path.join(dir_main,'fleet 2/fines4.csv'), sep=',')
    df4 = pd.read_excel(os.path.join(dir_main,'fleet 2/results2.xlsx'))
    df4.index = df4.Vessel
    df5 = pd.read_excel(os.path.join(dir_main,'fleet 2/results3.xlsx'))
    df5.index = df5.Vessel
    
    d = {'Vessel': df.columns[0:3] , 
         'Score':  df.iloc[-1] , 
         'Score t-1' : df4.Score,
         'Score t-2' : df5.Score,
         'Accumulated Profit':df2.iloc[-1] ,
         'Fines':df3.iloc[-1] }
    outcomes = pd.DataFrame(data=d)
    outcomes=outcomes.reset_index(drop=True)
    st.dataframe(outcomes)

    wb = openpyxl.Workbook()
    wb.save(os.path.join(dir_main,'fleet 2/reports4.xlsx')) 
    book = load_workbook(os.path.join(dir_main,'fleet 2/reports4.xlsx'))
    del book['Sheet']   

    writer = pd.ExcelWriter(os.path.join(dir_main,'fleet 2/reports4.xlsx'), engine='openpyxl')
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}
    outcomes.to_excel(writer, sheet_name = "Outcomes 4", na_rep = "", index =   False, startrow = 0)  
    writer.save()
    
    #Plot scores 
    df01 = pd.read_csv(os.path.join(dir_main,'fleet 2/score4.csv'), sep=',')
    df01.index = np.arange(1, len(df01) + 1)
    #Create the daily average 
    df01['Fleet score']=df01.sum(axis=1)/3
    df01['Action'] = df01.index
    df01 = pd.melt(df01, id_vars='Action', value_vars=["Vessel 1: Sansa", 
                                                    "Vessel 2: Olenna", 
                                                    "Vessel 3: Missandei", 
                                                    "Fleet score"])
    st.subheader("*Daily fleet and vessel scores*")
    fig01 = px.line(df01, x='Action', y='value', color='variable', markers=True,range_x=[0,10],width=650,height=650)
    fig01.update_layout(yaxis_type='category')
    fig01.layout.xaxis.fixedrange= True
    fig01.layout.yaxis.fixedrange: True
    fig01.update(layout_showlegend=True)
    fig01.update_traces(patch={"line": {"color": "green", "width": 3, "dash": 'dot'}}, selector={"legendgroup": "Fleet score"}) 
    st.plotly_chart(fig01)
  
    #Plot profits  
    df02 = pd.read_csv(os.path.join(dir_main,'fleet 2/profits4.csv'), sep=',')
    df03 = pd.read_csv(os.path.join(dir_main,'fleet 2/fines4.csv'), sep=',')
    df02.index = np.arange(1, len(df02) + 1)
    df03.index = np.arange(1, len(df03) + 1)

    df02['Action'] = df02.index
    df02['Fleet profit']=df02.sum(axis=1)
    aux = df01.loc[df01.variable=="Fleet score",'value'].reset_index(drop=True)
    aux.index = np.arange(1, len(aux) + 1)
    df02['Fleet score'] = aux
    
    df03['Fleet fines']=df03.sum(axis=1)
    df02['Fleet w. score'] = df02['Fleet profit'] - df03['Fleet fines']
    
    df02 = pd.melt(df02, id_vars='Action', value_vars=["Fleet profit",
                                                       "Fleet score",
                                                       "Fleet w. score"])
    df02_1 = df02.loc[df02.variable == "Fleet score",:]
    df02_2 = df02.loc[df02.variable != "Fleet score",:]
    
    # Create figure with secondary y-axis 
    subfig = make_subplots(specs=[[{"secondary_y": True}]])

    # Defining Custom Colors
    colours = {
      "Fleet profit": "#A1A9A3",
      "Fleet w. score": "#DFDF70",
    }

    # create two independent figures with px.line each containing data from multiple columns
    fig = px.line(df02_1, x='Action', y='value',markers=True,range_x=[1,10],width=650,height=650, text="value")
    fig.layout.xaxis.fixedrange= True
    fig.update_traces(line_color='orange', line_width=2)
    fig2 = px.bar(df02_2, y='value', x='Action', color='variable',range_x=[1,10],width=650,height=650,color_discrete_map=colours)
    fig.update_traces(yaxis="y2")
    fig2.update(layout_showlegend=True)
    
    subfig.add_traces(fig2.data + fig.data)
    subfig.layout.xaxis.fixedrange= True
    
    # Set y-axes titles
    subfig.layout.yaxis.title="Kr."
    subfig.layout.yaxis2.title="Score"
    subfig.update_xaxes(range=[0,10], dtick=1)
    st.subheader("*Fleet profit loss by score increase*")
    st.plotly_chart(subfig)
  
     #Save the data for next day 
    cav=open(dir_main+"fleet 2/reports4.xlsx",'rb')
    b64_01 = base64.b64encode(cav.read()).decode('utf-8')
    href = f'<a href="data:file/cav;base64,{b64_01}" download="results4.xlsx">:arrow_down_small: Overall results day 4 </a>'
    st.markdown(href, unsafe_allow_html=True)
  
###############################################################################
#2.i Day 5                                                                    #
###############################################################################
  if pag_input=="Day 5": 
    #Selection of section in the daily report
   # Title
   st.title(title)  
   st.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)
   st.header("**Day 5: Type in information**")
   st.markdown('''In this subsection the fleet manager types in the score and the profit for every action (this is already
accumulated)''')

   with st.expander("Vessel scores"):
#    if "df" not in st.session_state or score.columns[0]=="Unnamed: 0":
    if "df" not in st.session_state:
        st.session_state.df = pd.DataFrame(columns=["Vessel 1: Sansa", 
                                                    "Vessel 2: Olenna", 
                                                    "Vessel 3: Missandei"])


    st.markdown("Add Record")
    
    num_new_rows = 10
    ncol = st.session_state.df.shape[1]  # col count
    rw = -1
    
    with st.form(key="add form", clear_on_submit= True):
        cols = st.columns(ncol)
        rwdta = []
    
        for i in range(ncol):
            rwdta.append(cols[i].text_input(st.session_state.df.columns[i]))
    
        # you can insert code for a list comprehension here to change the data (rwdta) 
        # values into integer / float, if required
    
        if st.form_submit_button("Add"):
            if st.session_state.df.shape[0] == num_new_rows:
                st.error("Add row limit reached. Cant add any more records..")
            else:
                rw = st.session_state.df.shape[0] + 1
                st.info(f"Row: {rw} / {num_new_rows} added")
                st.session_state.df.loc[rw] = rwdta
    
                if st.session_state.df.shape[0] == num_new_rows:
                    st.error("Add row limit reached...")
    df = st.session_state.df
    df.to_csv("fleet 2/score5.csv",encoding='utf-8',index=False)
    st.text(df)
    
   with st.expander("Vessel profits"):
    if "df2" not in st.session_state:
        st.session_state.df2 = pd.DataFrame(columns=["Vessel 1: Sansa", 
                                                    "Vessel 2: Olenna", 
                                                    "Vessel 3: Missandei"])
    
    st.markdown("Add Record")
    
    ncol = st.session_state.df2.shape[1]  # col count
    rw = -1
    
    with st.form(key="add profits form", clear_on_submit= True):
        cols = st.columns(ncol)
        rwdta = []
    
        for i in range(ncol):
            rwdta.append(cols[i].text_input(st.session_state.df2.columns[i]))
    
        # you can insert code for a list comprehension here to change the data (rwdta) 
        # values into integer / float, if required
    
        if st.form_submit_button("Add"):
            if st.session_state.df2.shape[0] == num_new_rows:
                st.error("Add row limit reached. Cant add any more records..")
            else:
                rw = st.session_state.df2.shape[0] + 1
                st.info(f"Row: {rw} / {num_new_rows} added")
                st.session_state.df2.loc[rw] = rwdta
    
                if st.session_state.df2.shape[0] == num_new_rows:
                    st.error("Add row limit reached...")
    


    df2 = st.session_state.df2
    df2.to_csv("fleet 2/profits5.csv",encoding='utf-8',index=False)

    st.text(st.session_state.df2)
 
   with st.expander("Vessel fines"):
    if "df3" not in st.session_state:
        st.session_state.df3 = pd.DataFrame(columns=["Vessel 1: Sansa", 
                                                    "Vessel 2: Olenna", 
                                                    "Vessel 3: Missandei"])
    
    st.markdown("Add Record")
    
    ncol = st.session_state.df3.shape[1]  # col count
    rw = -1
    
    with st.form(key="add fines form", clear_on_submit= True):
        cols = st.columns(ncol)
        rwdta = []
    
        for i in range(ncol):
            rwdta.append(cols[i].text_input(st.session_state.df3.columns[i]))
    
        # you can insert code for a list comprehension here to change the data (rwdta) 
        # values into integer / float, if required
    
        if st.form_submit_button("Add"):
            if st.session_state.df3.shape[0] == num_new_rows:
                st.error("Add row limit reached. Cant add any more records..")
            else:
                rw = st.session_state.df3.shape[0] + 1
                st.info(f"Row: {rw} / {num_new_rows} added")
                st.session_state.df3.loc[rw] = rwdta
    
                if st.session_state.df3.shape[0] == num_new_rows:
                    st.error("Add row limit reached...")
    

    df3 = st.session_state.df3
    df3.to_csv("fleet 2/fines5.csv",encoding='utf-8',index=False)

    st.text(st.session_state.df3)

###############################################################################
#2.i Overall reports                                                          #
###############################################################################
   st.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)
   if st.button('Create overall reports'): 
    st.header("**Day 5: Overall reports**")
    st.markdown('''In this subsection the fleet manager can see the current accumulated score and profit of day 5 and the daily fleet scores and profits.''')
    st.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)
    #
    df = pd.read_csv(os.path.join(dir_main,'fleet 2/score5.csv'), sep=',')
    df2 = pd.read_csv(os.path.join(dir_main,'fleet 2/profits5.csv'), sep=',')
    df3 = pd.read_csv(os.path.join(dir_main,'fleet 2/fines5.csv'), sep=',')
    df4 = pd.read_excel(os.path.join(dir_main,'fleet 2/results3.xlsx'))
    df4.index = df4.Vessel
    df5 = pd.read_excel(os.path.join(dir_main,'fleet 2/results4.xlsx'))
    df5.index = df5.Vessel
    
    d = {'Vessel': df.columns[0:3] , 
         'Score':  df.iloc[-1] , 
         'Score t-1' : df4.Score,
         'Score t-2' : df5.Score,
         'Accumulated Profit':df2.iloc[-1] ,
         'Fines':df3.iloc[-1] }
    outcomes = pd.DataFrame(data=d)
    outcomes=outcomes.reset_index(drop=True)
    st.dataframe(outcomes)

    wb = openpyxl.Workbook()
    wb.save(os.path.join(dir_main,'fleet 2/reports5.xlsx')) 
    book = load_workbook(os.path.join(dir_main,'fleet 2/reports5.xlsx'))
    del book['Sheet']   

    writer = pd.ExcelWriter(os.path.join(dir_main,'fleet 2/reports5.xlsx'), engine='openpyxl')
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}
    outcomes.to_excel(writer, sheet_name = "Outcomes 5", na_rep = "", index =   False, startrow = 0)  
    writer.save()
    
    #Plot scores 
    df01 = pd.read_csv(os.path.join(dir_main,'fleet 2/score5.csv'), sep=',')
    df01.index = np.arange(1, len(df01) + 1)
    #Create the daily average 
    df01['Fleet score']=df01.sum(axis=1)/3
    df01['Action'] = df01.index
    df01 = pd.melt(df01, id_vars='Action', value_vars=["Vessel 1: Sansa", 
                                                    "Vessel 2: Olenna", 
                                                    "Vessel 3: Missandei", 
                                                    "Fleet score"])
    st.subheader("*Daily fleet and vessel scores*")
    fig01 = px.line(df01, x='Action', y='value', color='variable', markers=True,range_x=[0,10],width=650,height=650)
    fig01.update_layout(yaxis_type='category')
    fig01.layout.xaxis.fixedrange= True
    fig01.layout.yaxis.fixedrange: True
    fig01.update(layout_showlegend=True)
    fig01.update_traces(patch={"line": {"color": "green", "width": 3, "dash": 'dot'}}, selector={"legendgroup": "Fleet score"}) 
    st.plotly_chart(fig01)
  
    #Plot profits  
    df02 = pd.read_csv(os.path.join(dir_main,'fleet 2/profits5.csv'), sep=',')
    df03 = pd.read_csv(os.path.join(dir_main,'fleet 2/fines5.csv'), sep=',')
    df02.index = np.arange(1, len(df02) + 1)
    df03.index = np.arange(1, len(df03) + 1)

    df02['Action'] = df02.index
    df02['Fleet profit']=df02.sum(axis=1)
    aux = df01.loc[df01.variable=="Fleet score",'value'].reset_index(drop=True)
    aux.index = np.arange(1, len(aux) + 1)
    df02['Fleet score'] = aux
    
    df03['Fleet fines']=df03.sum(axis=1)
    df02['Fleet w. score'] = df02['Fleet profit'] - df03['Fleet fines']
    
    df02 = pd.melt(df02, id_vars='Action', value_vars=["Fleet profit",
                                                       "Fleet score",
                                                       "Fleet w. score"])
    df02_1 = df02.loc[df02.variable == "Fleet score",:]
    df02_2 = df02.loc[df02.variable != "Fleet score",:]
    
    # Create figure with secondary y-axis 
    subfig = make_subplots(specs=[[{"secondary_y": True}]])

    # Defining Custom Colors
    colours = {
      "Fleet profit": "#A1A9A3",
      "Fleet w. score": "#DFDF70",
    }

    # create two independent figures with px.line each containing data from multiple columns
    fig = px.line(df02_1, x='Action', y='value',markers=True,range_x=[1,10],width=650,height=650, text="value")
    fig.layout.xaxis.fixedrange= True
    fig.update_traces(line_color='orange', line_width=2)
    fig2 = px.bar(df02_2, y='value', x='Action', color='variable',range_x=[1,10],width=650,height=650,color_discrete_map=colours)
    fig.update_traces(yaxis="y2")
    fig2.update(layout_showlegend=True)
    
    subfig.add_traces(fig2.data + fig.data)
    subfig.layout.xaxis.fixedrange= True
    
    # Set y-axes titles
    subfig.layout.yaxis.title="Kr."
    subfig.layout.yaxis2.title="Score"
    subfig.update_xaxes(range=[0,10], dtick=1)
    st.subheader("*Fleet profit loss by score increase*")
    st.plotly_chart(subfig)
  
     #Save the data for next day 
    cav=open(dir_main+"fleet 2/reports5.xlsx",'rb')
    b64_01 = base64.b64encode(cav.read()).decode('utf-8')
    href = f'<a href="data:file/cav;base64,{b64_01}" download="results5.xlsx">:arrow_down_small: Overall results day 5 </a>'
    st.markdown(href, unsafe_allow_html=True)
 
###############################################################################
#2.i Day 6                                                                    #
###############################################################################
  if pag_input=="Day 6": 
    #Selection of section in the daily report
   # Title
   st.title(title)  
   st.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)
   st.header("**Day 6: Type in information**")
   st.markdown('''In this subsection the fleet manager types in the score and the profit for every action (this is already
accumulated)''')

   with st.expander("Vessel scores"):
#    if "df" not in st.session_state or score.columns[0]=="Unnamed: 0":
    if "df" not in st.session_state:
        st.session_state.df = pd.DataFrame(columns=["Vessel 1: Sansa", 
                                                    "Vessel 2: Olenna", 
                                                    "Vessel 3: Missandei"])


    st.markdown("Add Record")
    
    num_new_rows = 10
    ncol = st.session_state.df.shape[1]  # col count
    rw = -1
    
    with st.form(key="add form", clear_on_submit= True):
        cols = st.columns(ncol)
        rwdta = []
    
        for i in range(ncol):
            rwdta.append(cols[i].text_input(st.session_state.df.columns[i]))
    
        # you can insert code for a list comprehension here to change the data (rwdta) 
        # values into integer / float, if required
    
        if st.form_submit_button("Add"):
            if st.session_state.df.shape[0] == num_new_rows:
                st.error("Add row limit reached. Cant add any more records..")
            else:
                rw = st.session_state.df.shape[0] + 1
                st.info(f"Row: {rw} / {num_new_rows} added")
                st.session_state.df.loc[rw] = rwdta
    
                if st.session_state.df.shape[0] == num_new_rows:
                    st.error("Add row limit reached...")
    df = st.session_state.df
    df.to_csv("fleet 2/score6.csv",encoding='utf-8',index=False)
    st.text(df)
    
   with st.expander("Vessel profits"):
    if "df2" not in st.session_state:
        st.session_state.df2 = pd.DataFrame(columns=["Vessel 1: Sansa", 
                                                    "Vessel 2: Olenna", 
                                                    "Vessel 3: Missandei"])
    
    st.markdown("Add Record")
    
    ncol = st.session_state.df2.shape[1]  # col count
    rw = -1
    
    with st.form(key="add profits form", clear_on_submit= True):
        cols = st.columns(ncol)
        rwdta = []
    
        for i in range(ncol):
            rwdta.append(cols[i].text_input(st.session_state.df2.columns[i]))
    
        # you can insert code for a list comprehension here to change the data (rwdta) 
        # values into integer / float, if required
    
        if st.form_submit_button("Add"):
            if st.session_state.df2.shape[0] == num_new_rows:
                st.error("Add row limit reached. Cant add any more records..")
            else:
                rw = st.session_state.df2.shape[0] + 1
                st.info(f"Row: {rw} / {num_new_rows} added")
                st.session_state.df2.loc[rw] = rwdta
    
                if st.session_state.df2.shape[0] == num_new_rows:
                    st.error("Add row limit reached...")
    


    df2 = st.session_state.df2
    df2.to_csv("fleet 2/profits6.csv",encoding='utf-8',index=False)

    st.text(st.session_state.df2)
 
   with st.expander("Vessel fines"):
    if "df3" not in st.session_state:
        st.session_state.df3 = pd.DataFrame(columns=["Vessel 1: Sansa", 
                                                    "Vessel 2: Olenna", 
                                                    "Vessel 3: Missandei"])
    
    st.markdown("Add Record")
    
    ncol = st.session_state.df3.shape[1]  # col count
    rw = -1
    
    with st.form(key="add fines form", clear_on_submit= True):
        cols = st.columns(ncol)
        rwdta = []
    
        for i in range(ncol):
            rwdta.append(cols[i].text_input(st.session_state.df3.columns[i]))
    
        # you can insert code for a list comprehension here to change the data (rwdta) 
        # values into integer / float, if required
    
        if st.form_submit_button("Add"):
            if st.session_state.df3.shape[0] == num_new_rows:
                st.error("Add row limit reached. Cant add any more records..")
            else:
                rw = st.session_state.df3.shape[0] + 1
                st.info(f"Row: {rw} / {num_new_rows} added")
                st.session_state.df3.loc[rw] = rwdta
    
                if st.session_state.df3.shape[0] == num_new_rows:
                    st.error("Add row limit reached...")
    

    df3 = st.session_state.df3
    df3.to_csv("fleet 2/fines6.csv",encoding='utf-8',index=False)

    st.text(st.session_state.df3)

###############################################################################
#2.i Overall reports                                                          #
###############################################################################
   st.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)
   if st.button('Create overall reports'): 
    st.header("**Day 6: Overall reports**")
    st.markdown('''In this subsection the fleet manager can see the current accumulated score and profit of day 6 and the daily fleet scores and profits.''')
    st.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)
    #
    df = pd.read_csv(os.path.join(dir_main,'fleet 2/score6.csv'), sep=',')
    df2 = pd.read_csv(os.path.join(dir_main,'fleet 2/profits6.csv'), sep=',')
    df3 = pd.read_csv(os.path.join(dir_main,'fleet 2/fines6.csv'), sep=',')
    df4 = pd.read_excel(os.path.join(dir_main,'fleet 2/results4.xlsx'))
    df4.index = df4.Vessel
    df5 = pd.read_excel(os.path.join(dir_main,'fleet 2/results5.xlsx'))
    df5.index = df5.Vessel
    
    d = {'Vessel': df.columns[0:3] , 
         'Score':  df.iloc[-1] , 
         'Score t-1' : df4.Score,
         'Score t-2' : df5.Score,
         'Accumulated Profit':df2.iloc[-1] ,
         'Fines':df3.iloc[-1] }
    outcomes = pd.DataFrame(data=d)
    outcomes=outcomes.reset_index(drop=True)
    st.dataframe(outcomes)

    wb = openpyxl.Workbook()
    wb.save(os.path.join(dir_main,'fleet 2/reports6.xlsx')) 
    book = load_workbook(os.path.join(dir_main,'fleet 2/reports6.xlsx'))
    del book['Sheet']   

    writer = pd.ExcelWriter(os.path.join(dir_main,'fleet 2/reports6.xlsx'), engine='openpyxl')
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}
    outcomes.to_excel(writer, sheet_name = "Outcomes 6", na_rep = "", index =   False, startrow = 0)  
    writer.save()
    
    #Plot scores 
    df01 = pd.read_csv(os.path.join(dir_main,'fleet 2/score6.csv'), sep=',')
    df01.index = np.arange(1, len(df01) + 1)
    #Create the daily average 
    df01['Fleet score']=df01.sum(axis=1)/3
    df01['Action'] = df01.index
    df01 = pd.melt(df01, id_vars='Action', value_vars=["Vessel 1: Sansa", 
                                                    "Vessel 2: Olenna", 
                                                    "Vessel 3: Missandei", 
                                                    "Fleet score"])
    st.subheader("*Daily fleet and vessel scores*")
    fig01 = px.line(df01, x='Action', y='value', color='variable', markers=True,range_x=[0,10],width=650,height=650)
    fig01.update_layout(yaxis_type='category')
    fig01.layout.xaxis.fixedrange= True
    fig01.layout.yaxis.fixedrange: True
    fig01.update(layout_showlegend=True)
    fig01.update_traces(patch={"line": {"color": "green", "width": 3, "dash": 'dot'}}, selector={"legendgroup": "Fleet score"}) 
    st.plotly_chart(fig01)
  
    #Plot profits  
    df02 = pd.read_csv(os.path.join(dir_main,'fleet 2/profits6.csv'), sep=',')
    df03 = pd.read_csv(os.path.join(dir_main,'fleet 2/fines6.csv'), sep=',')
    df02.index = np.arange(1, len(df02) + 1)
    df03.index = np.arange(1, len(df03) + 1)

    df02['Action'] = df02.index
    df02['Fleet profit']=df02.sum(axis=1)
    aux = df01.loc[df01.variable=="Fleet score",'value'].reset_index(drop=True)
    aux.index = np.arange(1, len(aux) + 1)
    df02['Fleet score'] = aux
    
    df03['Fleet fines']=df03.sum(axis=1)
    df02['Fleet w. score'] = df02['Fleet profit'] - df03['Fleet fines']
    
    df02 = pd.melt(df02, id_vars='Action', value_vars=["Fleet profit",
                                                       "Fleet score",
                                                       "Fleet w. score"])
    df02_1 = df02.loc[df02.variable == "Fleet score",:]
    df02_2 = df02.loc[df02.variable != "Fleet score",:]
    
    # Create figure with secondary y-axis 
    subfig = make_subplots(specs=[[{"secondary_y": True}]])

    # Defining Custom Colors
    colours = {
      "Fleet profit": "#A1A9A3",
      "Fleet w. score": "#DFDF70",
    }

    # create two independent figures with px.line each containing data from multiple columns
    fig = px.line(df02_1, x='Action', y='value',markers=True,range_x=[1,10],width=650,height=650, text="value")
    fig.layout.xaxis.fixedrange= True
    fig.update_traces(line_color='orange', line_width=2)
    fig2 = px.bar(df02_2, y='value', x='Action', color='variable',range_x=[1,10],width=650,height=650,color_discrete_map=colours)
    fig.update_traces(yaxis="y2")
    fig2.update(layout_showlegend=True)
    
    subfig.add_traces(fig2.data + fig.data)
    subfig.layout.xaxis.fixedrange= True
    
    # Set y-axes titles
    subfig.layout.yaxis.title="Kr."
    subfig.layout.yaxis2.title="Score"
    subfig.update_xaxes(range=[0,10], dtick=1)
    st.subheader("*Fleet profit loss by score increase*")
    st.plotly_chart(subfig)
  
     #Save the data for next day 
    cav=open(dir_main+"fleet 2/reports6.xlsx",'rb')
    b64_01 = base64.b64encode(cav.read()).decode('utf-8')
    href = f'<a href="data:file/cav;base64,{b64_01}" download="results6.xlsx">:arrow_down_small: Overall results day 6 </a>'
    st.markdown(href, unsafe_allow_html=True)
  
###############################################################################
#2.i Day 7                                                                    #
###############################################################################
  if pag_input=="Day 7": 
    #Selection of section in the daily report
   # Title
   st.title(title)  
   st.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)
   st.header("**Day 7: Type in information**")
   st.markdown('''In this subsection the fleet manager types in the score and the profit for every action (this is already
accumulated)''')

   with st.expander("Vessel scores"):
#    if "df" not in st.session_state or score.columns[0]=="Unnamed: 0":
    if "df" not in st.session_state:
        st.session_state.df = pd.DataFrame(columns=["Vessel 1: Sansa", 
                                                    "Vessel 2: Olenna", 
                                                    "Vessel 3: Missandei"])


    st.markdown("Add Record")
    
    num_new_rows = 10
    ncol = st.session_state.df.shape[1]  # col count
    rw = -1
    
    with st.form(key="add form", clear_on_submit= True):
        cols = st.columns(ncol)
        rwdta = []
    
        for i in range(ncol):
            rwdta.append(cols[i].text_input(st.session_state.df.columns[i]))
    
        # you can insert code for a list comprehension here to change the data (rwdta) 
        # values into integer / float, if required
    
        if st.form_submit_button("Add"):
            if st.session_state.df.shape[0] == num_new_rows:
                st.error("Add row limit reached. Cant add any more records..")
            else:
                rw = st.session_state.df.shape[0] + 1
                st.info(f"Row: {rw} / {num_new_rows} added")
                st.session_state.df.loc[rw] = rwdta
    
                if st.session_state.df.shape[0] == num_new_rows:
                    st.error("Add row limit reached...")
    df = st.session_state.df
    df.to_csv("fleet 2/score7.csv",encoding='utf-8',index=False)
    st.text(df)
    
   with st.expander("Vessel profits"):
    if "df2" not in st.session_state:
        st.session_state.df2 = pd.DataFrame(columns=["Vessel 1: Sansa", 
                                                    "Vessel 2: Olenna", 
                                                    "Vessel 3: Missandei"])
    
    st.markdown("Add Record")
    
    ncol = st.session_state.df2.shape[1]  # col count
    rw = -1
    
    with st.form(key="add profits form", clear_on_submit= True):
        cols = st.columns(ncol)
        rwdta = []
    
        for i in range(ncol):
            rwdta.append(cols[i].text_input(st.session_state.df2.columns[i]))
    
        # you can insert code for a list comprehension here to change the data (rwdta) 
        # values into integer / float, if required
    
        if st.form_submit_button("Add"):
            if st.session_state.df2.shape[0] == num_new_rows:
                st.error("Add row limit reached. Cant add any more records..")
            else:
                rw = st.session_state.df2.shape[0] + 1
                st.info(f"Row: {rw} / {num_new_rows} added")
                st.session_state.df2.loc[rw] = rwdta
    
                if st.session_state.df2.shape[0] == num_new_rows:
                    st.error("Add row limit reached...")
    


    df2 = st.session_state.df2
    df2.to_csv("fleet 2/profits7.csv",encoding='utf-8',index=False)

    st.text(st.session_state.df2)
 
   with st.expander("Vessel fines"):
    if "df3" not in st.session_state:
        st.session_state.df3 = pd.DataFrame(columns=["Vessel 1: Sansa", 
                                                    "Vessel 2: Olenna", 
                                                    "Vessel 3: Missandei"])
    
    st.markdown("Add Record")
    
    ncol = st.session_state.df3.shape[1]  # col count
    rw = -1
    
    with st.form(key="add fines form", clear_on_submit= True):
        cols = st.columns(ncol)
        rwdta = []
    
        for i in range(ncol):
            rwdta.append(cols[i].text_input(st.session_state.df3.columns[i]))
    
        # you can insert code for a list comprehension here to change the data (rwdta) 
        # values into integer / float, if required
    
        if st.form_submit_button("Add"):
            if st.session_state.df3.shape[0] == num_new_rows:
                st.error("Add row limit reached. Cant add any more records..")
            else:
                rw = st.session_state.df3.shape[0] + 1
                st.info(f"Row: {rw} / {num_new_rows} added")
                st.session_state.df3.loc[rw] = rwdta
    
                if st.session_state.df3.shape[0] == num_new_rows:
                    st.error("Add row limit reached...")
    

    df3 = st.session_state.df3
    df3.to_csv("fleet 2/fines7.csv",encoding='utf-8',index=False)

    st.text(st.session_state.df3)

###############################################################################
#2.i Overall reports                                                          #
###############################################################################
   st.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)
   if st.button('Create overall reports'): 
    st.header("**Day 7: Overall reports**")
    st.markdown('''In this subsection the fleet manager can see the current accumulated score and profit of day 7 and the daily fleet scores and profits.''')
    st.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)
    #
    df = pd.read_csv(os.path.join(dir_main,'fleet 2/score7.csv'), sep=',')
    df2 = pd.read_csv(os.path.join(dir_main,'fleet 2/profits7.csv'), sep=',')
    df3 = pd.read_csv(os.path.join(dir_main,'fleet 2/fines7.csv'), sep=',')
    df4 = pd.read_excel(os.path.join(dir_main,'fleet 2/results5.xlsx'))
    df4.index = df4.Vessel
    df5 = pd.read_excel(os.path.join(dir_main,'fleet 2/results6.xlsx'))
    df5.index = df5.Vessel
    
    d = {'Vessel': df.columns[0:3] , 
         'Score':  df.iloc[-1] , 
         'Score t-1' : df4.Score,
         'Score t-2' : df5.Score,
         'Accumulated Profit':df2.iloc[-1] ,
         'Fines':df3.iloc[-1] }
    outcomes = pd.DataFrame(data=d)
    outcomes=outcomes.reset_index(drop=True)
    st.dataframe(outcomes)

    wb = openpyxl.Workbook()
    wb.save(os.path.join(dir_main,'fleet 2/reports7.xlsx')) 
    book = load_workbook(os.path.join(dir_main,'fleet 2/reports7.xlsx'))
    del book['Sheet']   

    writer = pd.ExcelWriter(os.path.join(dir_main,'fleet 2/reports7.xlsx'), engine='openpyxl')
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}
    outcomes.to_excel(writer, sheet_name = "Outcomes 7", na_rep = "", index =   False, startrow = 0)  
    writer.save()
    
    #Plot scores 
    df01 = pd.read_csv(os.path.join(dir_main,'fleet 2/score7.csv'), sep=',')
    df01.index = np.arange(1, len(df01) + 1)
    #Create the daily average 
    df01['Fleet score']=df01.sum(axis=1)/3
    df01['Action'] = df01.index
    df01 = pd.melt(df01, id_vars='Action', value_vars=["Vessel 1: Sansa", 
                                                    "Vessel 2: Olenna", 
                                                    "Vessel 3: Missandei", 
                                                    "Fleet score"])
    st.subheader("*Daily fleet and vessel scores*")
    fig01 = px.line(df01, x='Action', y='value', color='variable', markers=True,range_x=[0,10],width=650,height=650)
    fig01.update_layout(yaxis_type='category')
    fig01.layout.xaxis.fixedrange= True
    fig01.layout.yaxis.fixedrange: True
    fig01.update(layout_showlegend=True)
    fig01.update_traces(patch={"line": {"color": "green", "width": 3, "dash": 'dot'}}, selector={"legendgroup": "Fleet score"}) 
    st.plotly_chart(fig01)
  
    #Plot profits  
    df02 = pd.read_csv(os.path.join(dir_main,'fleet 2/profits7.csv'), sep=',')
    df03 = pd.read_csv(os.path.join(dir_main,'fleet 2/fines7.csv'), sep=',')
    df02.index = np.arange(1, len(df02) + 1)
    df03.index = np.arange(1, len(df03) + 1)

    df02['Action'] = df02.index
    df02['Fleet profit']=df02.sum(axis=1)
    aux = df01.loc[df01.variable=="Fleet score",'value'].reset_index(drop=True)
    aux.index = np.arange(1, len(aux) + 1)
    df02['Fleet score'] = aux
    
    df03['Fleet fines']=df03.sum(axis=1)
    df02['Fleet w. score'] = df02['Fleet profit'] - df03['Fleet fines']
    
    df02 = pd.melt(df02, id_vars='Action', value_vars=["Fleet profit",
                                                       "Fleet score",
                                                       "Fleet w. score"])
    df02_1 = df02.loc[df02.variable == "Fleet score",:]
    df02_2 = df02.loc[df02.variable != "Fleet score",:]
    
    # Create figure with secondary y-axis 
    subfig = make_subplots(specs=[[{"secondary_y": True}]])

    # Defining Custom Colors
    colours = {
      "Fleet profit": "#A1A9A3",
      "Fleet w. score": "#DFDF70",
    }

    # create two independent figures with px.line each containing data from multiple columns
    fig = px.line(df02_1, x='Action', y='value',markers=True,range_x=[1,10],width=650,height=650, text="value")
    fig.layout.xaxis.fixedrange= True
    fig.update_traces(line_color='orange', line_width=2)
    fig2 = px.bar(df02_2, y='value', x='Action', color='variable',range_x=[1,10],width=650,height=650,color_discrete_map=colours)
    fig.update_traces(yaxis="y2")
    fig2.update(layout_showlegend=True)
    
    subfig.add_traces(fig2.data + fig.data)
    subfig.layout.xaxis.fixedrange= True
    
    # Set y-axes titles
    subfig.layout.yaxis.title="Kr."
    subfig.layout.yaxis2.title="Score"
    subfig.update_xaxes(range=[0,10], dtick=1)
    st.subheader("*Fleet profit loss by score increase*")
    st.plotly_chart(subfig)
  
     #Save the data for next day 
    cav=open(dir_main+"fleet 2/reports7.xlsx",'rb')
    b64_01 = base64.b64encode(cav.read()).decode('utf-8')
    href = f'<a href="data:file/cav;base64,{b64_01}" download="results7.xlsx">:arrow_down_small: Overall results day 7 </a>'
    st.markdown(href, unsafe_allow_html=True)
  
###############################################################################
#2.i Day 8                                                                    #
###############################################################################
  if pag_input=="Day 8": 
    #Selection of section in the daily report
   # Title
   st.title(title)  
   st.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)
   st.header("**Day 8: Type in information**")
   st.markdown('''In this subsection the fleet manager types in the score and the profit for every action (this is already
accumulated)''')

   with st.expander("Vessel scores"):
#    if "df" not in st.session_state or score.columns[0]=="Unnamed: 0":
    if "df" not in st.session_state:
        st.session_state.df = pd.DataFrame(columns=["Vessel 1: Sansa", 
                                                    "Vessel 2: Olenna", 
                                                    "Vessel 3: Missandei"])


    st.markdown("Add Record")
    
    num_new_rows = 10
    ncol = st.session_state.df.shape[1]  # col count
    rw = -1
    
    with st.form(key="add form", clear_on_submit= True):
        cols = st.columns(ncol)
        rwdta = []
    
        for i in range(ncol):
            rwdta.append(cols[i].text_input(st.session_state.df.columns[i]))
    
        # you can insert code for a list comprehension here to change the data (rwdta) 
        # values into integer / float, if required
    
        if st.form_submit_button("Add"):
            if st.session_state.df.shape[0] == num_new_rows:
                st.error("Add row limit reached. Cant add any more records..")
            else:
                rw = st.session_state.df.shape[0] + 1
                st.info(f"Row: {rw} / {num_new_rows} added")
                st.session_state.df.loc[rw] = rwdta
    
                if st.session_state.df.shape[0] == num_new_rows:
                    st.error("Add row limit reached...")
    df = st.session_state.df
    df.to_csv("fleet 2/score8.csv",encoding='utf-8',index=False)
    st.text(df)
    
   with st.expander("Vessel profits"):
    if "df2" not in st.session_state:
        st.session_state.df2 = pd.DataFrame(columns=["Vessel 1: Sansa", 
                                                    "Vessel 2: Olenna", 
                                                    "Vessel 3: Missandei"])
    
    st.markdown("Add Record")
    
    ncol = st.session_state.df2.shape[1]  # col count
    rw = -1
    
    with st.form(key="add profits form", clear_on_submit= True):
        cols = st.columns(ncol)
        rwdta = []
    
        for i in range(ncol):
            rwdta.append(cols[i].text_input(st.session_state.df2.columns[i]))
    
        # you can insert code for a list comprehension here to change the data (rwdta) 
        # values into integer / float, if required
    
        if st.form_submit_button("Add"):
            if st.session_state.df2.shape[0] == num_new_rows:
                st.error("Add row limit reached. Cant add any more records..")
            else:
                rw = st.session_state.df2.shape[0] + 1
                st.info(f"Row: {rw} / {num_new_rows} added")
                st.session_state.df2.loc[rw] = rwdta
    
                if st.session_state.df2.shape[0] == num_new_rows:
                    st.error("Add row limit reached...")
    


    df2 = st.session_state.df2
    df2.to_csv("fleet 2/profits8.csv",encoding='utf-8',index=False)

    st.text(st.session_state.df2)
 
   with st.expander("Vessel fines"):
    if "df3" not in st.session_state:
        st.session_state.df3 = pd.DataFrame(columns=["Vessel 1: Sansa", 
                                                    "Vessel 2: Olenna", 
                                                    "Vessel 3: Missandei"])
    
    st.markdown("Add Record")
    
    ncol = st.session_state.df3.shape[1]  # col count
    rw = -1
    
    with st.form(key="add fines form", clear_on_submit= True):
        cols = st.columns(ncol)
        rwdta = []
    
        for i in range(ncol):
            rwdta.append(cols[i].text_input(st.session_state.df3.columns[i]))
    
        # you can insert code for a list comprehension here to change the data (rwdta) 
        # values into integer / float, if required
    
        if st.form_submit_button("Add"):
            if st.session_state.df3.shape[0] == num_new_rows:
                st.error("Add row limit reached. Cant add any more records..")
            else:
                rw = st.session_state.df3.shape[0] + 1
                st.info(f"Row: {rw} / {num_new_rows} added")
                st.session_state.df3.loc[rw] = rwdta
    
                if st.session_state.df3.shape[0] == num_new_rows:
                    st.error("Add row limit reached...")
    

    df3 = st.session_state.df3
    df3.to_csv("fleet 2/fines8.csv",encoding='utf-8',index=False)

    st.text(st.session_state.df3)

###############################################################################
#2.i Overall reports                                                          #
###############################################################################
   st.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)
   if st.button('Create overall reports'): 
    st.header("**Day 8: Overall reports**")
    st.markdown('''In this subsection the fleet manager can see the current accumulated score and profit of day 8 and the daily fleet scores and profits.''')
    st.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)
    #
    df = pd.read_csv(os.path.join(dir_main,'fleet 2/score8.csv'), sep=',')
    df2 = pd.read_csv(os.path.join(dir_main,'fleet 2/profits8.csv'), sep=',')
    df3 = pd.read_csv(os.path.join(dir_main,'fleet 2/fines8.csv'), sep=',')
    df4 = pd.read_excel(os.path.join(dir_main,'fleet 2/results6.xlsx'))
    df4.index = df4.Vessel
    df5 = pd.read_excel(os.path.join(dir_main,'fleet 2/results7.xlsx'))
    df5.index = df5.Vessel
    
    d = {'Vessel': df.columns[0:3] , 
         'Score':  df.iloc[-1] , 
         'Score t-1' : df4.Score,
         'Score t-2' : df5.Score,
         'Accumulated Profit':df2.iloc[-1] ,
         'Fines':df3.iloc[-1] }
    outcomes = pd.DataFrame(data=d)
    outcomes=outcomes.reset_index(drop=True)
    st.dataframe(outcomes)

    wb = openpyxl.Workbook()
    wb.save(os.path.join(dir_main,'fleet 2/reports8.xlsx')) 
    book = load_workbook(os.path.join(dir_main,'fleet 2/reports8.xlsx'))
    del book['Sheet']   

    writer = pd.ExcelWriter(os.path.join(dir_main,'fleet 2/reports8.xlsx'), engine='openpyxl')
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}
    outcomes.to_excel(writer, sheet_name = "Outcomes 8", na_rep = "", index =   False, startrow = 0)  
    writer.save()
    
    #Plot scores 
    df01 = pd.read_csv(os.path.join(dir_main,'fleet 2/score8.csv'), sep=',')
    df01.index = np.arange(1, len(df01) + 1)
    #Create the daily average 
    df01['Fleet score']=df01.sum(axis=1)/3
    df01['Action'] = df01.index
    df01 = pd.melt(df01, id_vars='Action', value_vars=["Vessel 1: Sansa", 
                                                    "Vessel 2: Olenna", 
                                                    "Vessel 3: Missandei", 
                                                    "Fleet score"])
    st.subheader("*Daily fleet and vessel scores*")
    fig01 = px.line(df01, x='Action', y='value', color='variable', markers=True,range_x=[0,10],width=650,height=650)
    fig01.update_layout(yaxis_type='category')
    fig01.layout.xaxis.fixedrange= True
    fig01.layout.yaxis.fixedrange: True
    fig01.update(layout_showlegend=True)
    fig01.update_traces(patch={"line": {"color": "green", "width": 3, "dash": 'dot'}}, selector={"legendgroup": "Fleet score"}) 
    st.plotly_chart(fig01)
  
    #Plot profits  
    df02 = pd.read_csv(os.path.join(dir_main,'fleet 2/profits8.csv'), sep=',')
    df03 = pd.read_csv(os.path.join(dir_main,'fleet 2/fines8.csv'), sep=',')
    df02.index = np.arange(1, len(df02) + 1)
    df03.index = np.arange(1, len(df03) + 1)

    df02['Action'] = df02.index
    df02['Fleet profit']=df02.sum(axis=1)
    aux = df01.loc[df01.variable=="Fleet score",'value'].reset_index(drop=True)
    aux.index = np.arange(1, len(aux) + 1)
    df02['Fleet score'] = aux
    
    df03['Fleet fines']=df03.sum(axis=1)
    df02['Fleet w. score'] = df02['Fleet profit'] - df03['Fleet fines']
    
    df02 = pd.melt(df02, id_vars='Action', value_vars=["Fleet profit",
                                                       "Fleet score",
                                                       "Fleet w. score"])
    df02_1 = df02.loc[df02.variable == "Fleet score",:]
    df02_2 = df02.loc[df02.variable != "Fleet score",:]
    
    # Create figure with secondary y-axis 
    subfig = make_subplots(specs=[[{"secondary_y": True}]])

    # Defining Custom Colors
    colours = {
      "Fleet profit": "#A1A9A3",
      "Fleet w. score": "#DFDF70",
    }

    # create two independent figures with px.line each containing data from multiple columns
    fig = px.line(df02_1, x='Action', y='value',markers=True,range_x=[1,10],width=650,height=650, text="value")
    fig.layout.xaxis.fixedrange= True
    fig.update_traces(line_color='orange', line_width=2)
    fig2 = px.bar(df02_2, y='value', x='Action', color='variable',range_x=[1,10],width=650,height=650,color_discrete_map=colours)
    fig.update_traces(yaxis="y2")
    fig2.update(layout_showlegend=True)
    
    subfig.add_traces(fig2.data + fig.data)
    subfig.layout.xaxis.fixedrange= True
    
    # Set y-axes titles
    subfig.layout.yaxis.title="Kr."
    subfig.layout.yaxis2.title="Score"
    subfig.update_xaxes(range=[0,10], dtick=1)
    st.subheader("*Fleet profit loss by score increase*")
    st.plotly_chart(subfig)
  
     #Save the data for next day 
    cav=open(dir_main+"fleet 2/reports8.xlsx",'rb')
    b64_01 = base64.b64encode(cav.read()).decode('utf-8')
    href = f'<a href="data:file/cav;base64,{b64_01}" download="results8.xlsx">:arrow_down_small: Overall results day 8 </a>'
    st.markdown(href, unsafe_allow_html=True)
  
###############################################################################
#2.i Day 4                                                                    #
###############################################################################
  if pag_input=="Day 9": 
    #Selection of section in the daily report
   # Title
   st.title(title)  
   st.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)
   st.header("**Day 9: Type in information**")
   st.markdown('''In this subsection the fleet manager types in the score and the profit for every action (this is already
accumulated)''')

   with st.expander("Vessel scores"):
#    if "df" not in st.session_state or score.columns[0]=="Unnamed: 0":
    if "df" not in st.session_state:
        st.session_state.df = pd.DataFrame(columns=["Vessel 1: Sansa", 
                                                    "Vessel 2: Olenna", 
                                                    "Vessel 3: Missandei"])


    st.markdown("Add Record")
    
    num_new_rows = 10
    ncol = st.session_state.df.shape[1]  # col count
    rw = -1
    
    with st.form(key="add form", clear_on_submit= True):
        cols = st.columns(ncol)
        rwdta = []
    
        for i in range(ncol):
            rwdta.append(cols[i].text_input(st.session_state.df.columns[i]))
    
        # you can insert code for a list comprehension here to change the data (rwdta) 
        # values into integer / float, if required
    
        if st.form_submit_button("Add"):
            if st.session_state.df.shape[0] == num_new_rows:
                st.error("Add row limit reached. Cant add any more records..")
            else:
                rw = st.session_state.df.shape[0] + 1
                st.info(f"Row: {rw} / {num_new_rows} added")
                st.session_state.df.loc[rw] = rwdta
    
                if st.session_state.df.shape[0] == num_new_rows:
                    st.error("Add row limit reached...")
    df = st.session_state.df
    df.to_csv("fleet 2/score9.csv",encoding='utf-8',index=False)
    st.text(df)
    
   with st.expander("Vessel profits"):
    if "df2" not in st.session_state:
        st.session_state.df2 = pd.DataFrame(columns=["Vessel 1: Sansa", 
                                                    "Vessel 2: Olenna", 
                                                    "Vessel 3: Missandei"])
    
    st.markdown("Add Record")
    
    ncol = st.session_state.df2.shape[1]  # col count
    rw = -1
    
    with st.form(key="add profits form", clear_on_submit= True):
        cols = st.columns(ncol)
        rwdta = []
    
        for i in range(ncol):
            rwdta.append(cols[i].text_input(st.session_state.df2.columns[i]))
    
        # you can insert code for a list comprehension here to change the data (rwdta) 
        # values into integer / float, if required
    
        if st.form_submit_button("Add"):
            if st.session_state.df2.shape[0] == num_new_rows:
                st.error("Add row limit reached. Cant add any more records..")
            else:
                rw = st.session_state.df2.shape[0] + 1
                st.info(f"Row: {rw} / {num_new_rows} added")
                st.session_state.df2.loc[rw] = rwdta
    
                if st.session_state.df2.shape[0] == num_new_rows:
                    st.error("Add row limit reached...")
    


    df2 = st.session_state.df2
    df2.to_csv("fleet 2/profits9.csv",encoding='utf-8',index=False)

    st.text(st.session_state.df2)
 
   with st.expander("Vessel fines"):
    if "df3" not in st.session_state:
        st.session_state.df3 = pd.DataFrame(columns=["Vessel 1: Sansa", 
                                                    "Vessel 2: Olenna", 
                                                    "Vessel 3: Missandei"])
    
    st.markdown("Add Record")
    
    ncol = st.session_state.df3.shape[1]  # col count
    rw = -1
    
    with st.form(key="add fines form", clear_on_submit= True):
        cols = st.columns(ncol)
        rwdta = []
    
        for i in range(ncol):
            rwdta.append(cols[i].text_input(st.session_state.df3.columns[i]))
    
        # you can insert code for a list comprehension here to change the data (rwdta) 
        # values into integer / float, if required
    
        if st.form_submit_button("Add"):
            if st.session_state.df3.shape[0] == num_new_rows:
                st.error("Add row limit reached. Cant add any more records..")
            else:
                rw = st.session_state.df3.shape[0] + 1
                st.info(f"Row: {rw} / {num_new_rows} added")
                st.session_state.df3.loc[rw] = rwdta
    
                if st.session_state.df3.shape[0] == num_new_rows:
                    st.error("Add row limit reached...")
    

    df3 = st.session_state.df3
    df3.to_csv("fleet 2/fines9.csv",encoding='utf-8',index=False)

    st.text(st.session_state.df3)

###############################################################################
#2.i Overall reports                                                          #
###############################################################################
   st.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)
   if st.button('Create overall reports'): 
    st.header("**Day 9: Overall reports**")
    st.markdown('''In this subsection the fleet manager can see the current accumulated score and profit of day 9 and the daily fleet scores and profits.''')
    st.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)
    #
    df = pd.read_csv(os.path.join(dir_main,'fleet 2/score9.csv'), sep=',')
    df2 = pd.read_csv(os.path.join(dir_main,'fleet 2/profits9.csv'), sep=',')
    df3 = pd.read_csv(os.path.join(dir_main,'fleet 2/fines9.csv'), sep=',')
    df4 = pd.read_excel(os.path.join(dir_main,'fleet 2/results7.xlsx'))
    df4.index = df4.Vessel
    df5 = pd.read_excel(os.path.join(dir_main,'fleet 2/results8.xlsx'))
    df5.index = df5.Vessel
    
    d = {'Vessel': df.columns[0:3] , 
         'Score':  df.iloc[-1] , 
         'Score t-1' : df4.Score,
         'Score t-2' : df5.Score,
         'Accumulated Profit':df2.iloc[-1] ,
         'Fines':df3.iloc[-1] }
    outcomes = pd.DataFrame(data=d)
    outcomes=outcomes.reset_index(drop=True)
    st.dataframe(outcomes)

    wb = openpyxl.Workbook()
    wb.save(os.path.join(dir_main,'fleet 2/reports9.xlsx')) 
    book = load_workbook(os.path.join(dir_main,'fleet 2/reports9.xlsx'))
    del book['Sheet']   

    writer = pd.ExcelWriter(os.path.join(dir_main,'fleet 2/reports9.xlsx'), engine='openpyxl')
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}
    outcomes.to_excel(writer, sheet_name = "Outcomes 9", na_rep = "", index =   False, startrow = 0)  
    writer.save()
    
    #Plot scores 
    df01 = pd.read_csv(os.path.join(dir_main,'fleet 2/score9.csv'), sep=',')
    df01.index = np.arange(1, len(df01) + 1)
    #Create the daily average 
    df01['Fleet score']=df01.sum(axis=1)/3
    df01['Action'] = df01.index
    df01 = pd.melt(df01, id_vars='Action', value_vars=["Vessel 1: Sansa", 
                                                    "Vessel 2: Olenna", 
                                                    "Vessel 3: Missandei", 
                                                    "Fleet score"])
    st.subheader("*Daily fleet and vessel scores*")
    fig01 = px.line(df01, x='Action', y='value', color='variable', markers=True,range_x=[0,10],width=650,height=650)
    fig01.update_layout(yaxis_type='category')
    fig01.layout.xaxis.fixedrange= True
    fig01.layout.yaxis.fixedrange: True
    fig01.update(layout_showlegend=True)
    fig01.update_traces(patch={"line": {"color": "green", "width": 3, "dash": 'dot'}}, selector={"legendgroup": "Fleet score"}) 
    st.plotly_chart(fig01)
  
    #Plot profits  
    df02 = pd.read_csv(os.path.join(dir_main,'fleet 2/profits9.csv'), sep=',')
    df03 = pd.read_csv(os.path.join(dir_main,'fleet 2/fines9.csv'), sep=',')
    df02.index = np.arange(1, len(df02) + 1)
    df03.index = np.arange(1, len(df03) + 1)

    df02['Action'] = df02.index
    df02['Fleet profit']=df02.sum(axis=1)
    aux = df01.loc[df01.variable=="Fleet score",'value'].reset_index(drop=True)
    aux.index = np.arange(1, len(aux) + 1)
    df02['Fleet score'] = aux
    
    df03['Fleet fines']=df03.sum(axis=1)
    df02['Fleet w. score'] = df02['Fleet profit'] - df03['Fleet fines']
    
    df02 = pd.melt(df02, id_vars='Action', value_vars=["Fleet profit",
                                                       "Fleet score",
                                                       "Fleet w. score"])
    df02_1 = df02.loc[df02.variable == "Fleet score",:]
    df02_2 = df02.loc[df02.variable != "Fleet score",:]
    
    # Create figure with secondary y-axis 
    subfig = make_subplots(specs=[[{"secondary_y": True}]])

    # Defining Custom Colors
    colours = {
      "Fleet profit": "#A1A9A3",
      "Fleet w. score": "#DFDF70",
    }

    # create two independent figures with px.line each containing data from multiple columns
    fig = px.line(df02_1, x='Action', y='value',markers=True,range_x=[1,10],width=650,height=650, text="value")
    fig.layout.xaxis.fixedrange= True
    fig.update_traces(line_color='orange', line_width=2)
    fig2 = px.bar(df02_2, y='value', x='Action', color='variable',range_x=[1,10],width=650,height=650,color_discrete_map=colours)
    fig.update_traces(yaxis="y2")
    fig2.update(layout_showlegend=True)
    
    subfig.add_traces(fig2.data + fig.data)
    subfig.layout.xaxis.fixedrange= True
    
    # Set y-axes titles
    subfig.layout.yaxis.title="Kr."
    subfig.layout.yaxis2.title="Score"
    subfig.update_xaxes(range=[0,10], dtick=1)
    st.subheader("*Fleet profit loss by score increase*")
    st.plotly_chart(subfig)
  
     #Save the data for next day 
    cav=open(dir_main+"fleet 2/reports9.xlsx",'rb')
    b64_01 = base64.b64encode(cav.read()).decode('utf-8')
    href = f'<a href="data:file/cav;base64,{b64_01}" download="results9.xlsx">:arrow_down_small: Overall results day 9 </a>'
    st.markdown(href, unsafe_allow_html=True)
  
  
###############################################################################
#2.i Day 10                                                                   #
###############################################################################
  if pag_input=="Day 10": 
    #Selection of section in the daily report
   # Title
   st.title(title)  
   st.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)
   st.header("**Day 10: Type in information**")
   st.markdown('''In this subsection the fleet manager types in the score and the profit for every action (this is already
accumulated)''')

   with st.expander("Vessel scores"):
#    if "df" not in st.session_state or score.columns[0]=="Unnamed: 0":
    if "df" not in st.session_state:
        st.session_state.df = pd.DataFrame(columns=["Vessel 1: Sansa", 
                                                    "Vessel 2: Olenna", 
                                                    "Vessel 3: Missandei"])


    st.markdown("Add Record")
    
    num_new_rows = 10
    ncol = st.session_state.df.shape[1]  # col count
    rw = -1
    
    with st.form(key="add form", clear_on_submit= True):
        cols = st.columns(ncol)
        rwdta = []
    
        for i in range(ncol):
            rwdta.append(cols[i].text_input(st.session_state.df.columns[i]))
    
        # you can insert code for a list comprehension here to change the data (rwdta) 
        # values into integer / float, if required
    
        if st.form_submit_button("Add"):
            if st.session_state.df.shape[0] == num_new_rows:
                st.error("Add row limit reached. Cant add any more records..")
            else:
                rw = st.session_state.df.shape[0] + 1
                st.info(f"Row: {rw} / {num_new_rows} added")
                st.session_state.df.loc[rw] = rwdta
    
                if st.session_state.df.shape[0] == num_new_rows:
                    st.error("Add row limit reached...")
    df = st.session_state.df
    df.to_csv("fleet 2/score10.csv",encoding='utf-8',index=False)
    st.text(df)
    
   with st.expander("Vessel profits"):
    if "df2" not in st.session_state:
        st.session_state.df2 = pd.DataFrame(columns=["Vessel 1: Sansa", 
                                                    "Vessel 2: Olenna", 
                                                    "Vessel 3: Missandei"])
    
    st.markdown("Add Record")
    
    ncol = st.session_state.df2.shape[1]  # col count
    rw = -1
    
    with st.form(key="add profits form", clear_on_submit= True):
        cols = st.columns(ncol)
        rwdta = []
    
        for i in range(ncol):
            rwdta.append(cols[i].text_input(st.session_state.df2.columns[i]))
    
        # you can insert code for a list comprehension here to change the data (rwdta) 
        # values into integer / float, if required
    
        if st.form_submit_button("Add"):
            if st.session_state.df2.shape[0] == num_new_rows:
                st.error("Add row limit reached. Cant add any more records..")
            else:
                rw = st.session_state.df2.shape[0] + 1
                st.info(f"Row: {rw} / {num_new_rows} added")
                st.session_state.df2.loc[rw] = rwdta
    
                if st.session_state.df2.shape[0] == num_new_rows:
                    st.error("Add row limit reached...")
    


    df2 = st.session_state.df2
    df2.to_csv("fleet 2/profits10.csv",encoding='utf-8',index=False)

    st.text(st.session_state.df2)
 
   with st.expander("Vessel fines"):
    if "df3" not in st.session_state:
        st.session_state.df3 = pd.DataFrame(columns=["Vessel 1: Sansa", 
                                                    "Vessel 2: Olenna", 
                                                    "Vessel 3: Missandei"])
    
    st.markdown("Add Record")
    
    ncol = st.session_state.df3.shape[1]  # col count
    rw = -1
    
    with st.form(key="add fines form", clear_on_submit= True):
        cols = st.columns(ncol)
        rwdta = []
    
        for i in range(ncol):
            rwdta.append(cols[i].text_input(st.session_state.df3.columns[i]))
    
        # you can insert code for a list comprehension here to change the data (rwdta) 
        # values into integer / float, if required
    
        if st.form_submit_button("Add"):
            if st.session_state.df3.shape[0] == num_new_rows:
                st.error("Add row limit reached. Cant add any more records..")
            else:
                rw = st.session_state.df3.shape[0] + 1
                st.info(f"Row: {rw} / {num_new_rows} added")
                st.session_state.df3.loc[rw] = rwdta
    
                if st.session_state.df3.shape[0] == num_new_rows:
                    st.error("Add row limit reached...")
    

    df3 = st.session_state.df3
    df3.to_csv("fleet 2/fines10.csv",encoding='utf-8',index=False)

    st.text(st.session_state.df3)

###############################################################################
#2.i Overall reports                                                          #
###############################################################################
   st.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)
   if st.button('Create overall reports'): 
    st.header("**Day 10: Overall reports**")
    st.markdown('''In this subsection the fleet manager can see the current accumulated score and profit of day 10 and the daily fleet scores and profits.''')
    st.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)
    #
    df = pd.read_csv(os.path.join(dir_main,'fleet 2/score10.csv'), sep=',')
    df2 = pd.read_csv(os.path.join(dir_main,'fleet 2/profits10.csv'), sep=',')
    df3 = pd.read_csv(os.path.join(dir_main,'fleet 2/fines10.csv'), sep=',')
    df4 = pd.read_excel(os.path.join(dir_main,'fleet 2/results8.xlsx'))
    df4.index = df4.Vessel
    df5 = pd.read_excel(os.path.join(dir_main,'fleet 2/results9.xlsx'))
    df5.index = df5.Vessel
    
    d = {'Vessel': df.columns[0:3] , 
         'Score':  df.iloc[-1] , 
         'Score t-1' : df4.Score,
         'Score t-2' : df5.Score,
         'Accumulated Profit':df2.iloc[-1] ,
         'Fines':df3.iloc[-1] }
    outcomes = pd.DataFrame(data=d)
    outcomes=outcomes.reset_index(drop=True)
    st.dataframe(outcomes)

    wb = openpyxl.Workbook()
    wb.save(os.path.join(dir_main,'fleet 2/reports10.xlsx')) 
    book = load_workbook(os.path.join(dir_main,'fleet 2/reports10.xlsx'))
    del book['Sheet']   

    writer = pd.ExcelWriter(os.path.join(dir_main,'fleet 2/reports10.xlsx'), engine='openpyxl')
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}
    outcomes.to_excel(writer, sheet_name = "Outcomes 10", na_rep = "", index =   False, startrow = 0)  
    writer.save()
    
    #Plot scores 
    df01 = pd.read_csv(os.path.join(dir_main,'fleet 2/score10.csv'), sep=',')
    df01.index = np.arange(1, len(df01) + 1)
    #Create the daily average 
    df01['Fleet score']=df01.sum(axis=1)/3
    df01['Action'] = df01.index
    df01 = pd.melt(df01, id_vars='Action', value_vars=["Vessel 1: Sansa", 
                                                    "Vessel 2: Olenna", 
                                                    "Vessel 3: Missandei", 
                                                    "Fleet score"])
    st.subheader("*Daily fleet and vessel scores*")
    fig01 = px.line(df01, x='Action', y='value', color='variable', markers=True,range_x=[0,10],width=650,height=650)
    fig01.update_layout(yaxis_type='category')
    fig01.layout.xaxis.fixedrange= True
    fig01.layout.yaxis.fixedrange: True
    fig01.update(layout_showlegend=True)
    fig01.update_traces(patch={"line": {"color": "green", "width": 3, "dash": 'dot'}}, selector={"legendgroup": "Fleet score"}) 
    st.plotly_chart(fig01)
  
    #Plot profits  
    df02 = pd.read_csv(os.path.join(dir_main,'fleet 2/profits10.csv'), sep=',')
    df03 = pd.read_csv(os.path.join(dir_main,'fleet 2/fines10.csv'), sep=',')
    df02.index = np.arange(1, len(df02) + 1)
    df03.index = np.arange(1, len(df03) + 1)

    df02['Action'] = df02.index
    df02['Fleet profit']=df02.sum(axis=1)
    aux = df01.loc[df01.variable=="Fleet score",'value'].reset_index(drop=True)
    aux.index = np.arange(1, len(aux) + 1)
    df02['Fleet score'] = aux
    
    df03['Fleet fines']=df03.sum(axis=1)
    df02['Fleet w. score'] = df02['Fleet profit'] - df03['Fleet fines']
    
    df02 = pd.melt(df02, id_vars='Action', value_vars=["Fleet profit",
                                                       "Fleet score",
                                                       "Fleet w. score"])
    df02_1 = df02.loc[df02.variable == "Fleet score",:]
    df02_2 = df02.loc[df02.variable != "Fleet score",:]
    
    # Create figure with secondary y-axis 
    subfig = make_subplots(specs=[[{"secondary_y": True}]])

    # Defining Custom Colors
    colours = {
      "Fleet profit": "#A1A9A3",
      "Fleet w. score": "#DFDF70",
    }

    # create two independent figures with px.line each containing data from multiple columns
    fig = px.line(df02_1, x='Action', y='value',markers=True,range_x=[1,10],width=650,height=650, text="value")
    fig.layout.xaxis.fixedrange= True
    fig.update_traces(line_color='orange', line_width=2)
    fig2 = px.bar(df02_2, y='value', x='Action', color='variable',range_x=[1,10],width=650,height=650,color_discrete_map=colours)
    fig.update_traces(yaxis="y2")
    fig2.update(layout_showlegend=True)
    
    subfig.add_traces(fig2.data + fig.data)
    subfig.layout.xaxis.fixedrange= True
    
    # Set y-axes titles
    subfig.layout.yaxis.title="Kr."
    subfig.layout.yaxis2.title="Score"
    subfig.update_xaxes(range=[0,10], dtick=1)
    st.subheader("*Fleet profit loss by score increase*")
    st.plotly_chart(subfig)
  
     #Save the data for next day 
    cav=open(dir_main+"fleet 2/reports10.xlsx",'rb')
    b64_01 = base64.b64encode(cav.read()).decode('utf-8')
    href = f'<a href="data:file/cav;base64,{b64_01}" download="results10.xlsx">:arrow_down_small: Overall results day 10 </a>'
    st.markdown(href, unsafe_allow_html=True)
   
###############################################################################
#A. Fleets                                                                    #
###############################################################################
if user == "fleet3": 
    
###############################################################################
#2. Section selection                                                         #
###############################################################################
  section= ["Day 1", "Day 2", 'Day 3','Day 4','Day 5','Day 6','Day 7','Day 8','Day 9','Day 10']
  
  st.sidebar.title("Section selection")
  st.sidebar.markdown('''Select the day you wish to see.''')
  
  #Selection of daily report 
  pag_input = st.sidebar.selectbox('Select day:', section)

  st.sidebar.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)

###############################################################################
#2.i Day 1                                                                    #
###############################################################################
  if pag_input=="Day 1": 
    #Selection of section in the daily report
   # Title
   st.title(title)  
   st.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)
   st.header("**Day 1: Type in information**")
   st.markdown('''In this subsection the fleet manager types in the score and the profit for every action (this is already
accumulated)''')

   with st.expander("Vessel scores"):
#    if "df" not in st.session_state or score.columns[0]=="Unnamed: 0":
    if "df" not in st.session_state:
        st.session_state.df = pd.DataFrame(columns=["Vessel 1: Margaery", 
                                                    "Vessel 2: Brienne", 
                                                    "Vessel 3: Lysa"])


    st.markdown("Add Record")
    
    num_new_rows = 10
    ncol = st.session_state.df.shape[1]  # col count
    rw = -1
    
    with st.form(key="add form", clear_on_submit= True):
        cols = st.columns(ncol)
        rwdta = []
    
        for i in range(ncol):
            rwdta.append(cols[i].text_input(st.session_state.df.columns[i]))
    
        # you can insert code for a list comprehension here to change the data (rwdta) 
        # values into integer / float, if required
    
        if st.form_submit_button("Add"):
            if st.session_state.df.shape[0] == num_new_rows:
                st.error("Add row limit reached. Cant add any more records..")
            else:
                rw = st.session_state.df.shape[0] + 1
                st.info(f"Row: {rw} / {num_new_rows} added")
                st.session_state.df.loc[rw] = rwdta
    
                if st.session_state.df.shape[0] == num_new_rows:
                    st.error("Add row limit reached...")
    df = st.session_state.df
    df.to_csv("fleet 3/score1.csv",encoding='utf-8',index=False)
    st.text(df)
    
   with st.expander("Vessel profits"):
    if "df2" not in st.session_state:
        st.session_state.df2 = pd.DataFrame(columns=["Vessel 1: Margaery", 
                                                    "Vessel 2: Brienne", 
                                                    "Vessel 3: Lysa"])
    
    st.markdown("Add Record")
    
    ncol = st.session_state.df2.shape[1]  # col count
    rw = -1
    
    with st.form(key="add profits form", clear_on_submit= True):
        cols = st.columns(ncol)
        rwdta = []
    
        for i in range(ncol):
            rwdta.append(cols[i].text_input(st.session_state.df2.columns[i]))
    
        # you can insert code for a list comprehension here to change the data (rwdta) 
        # values into integer / float, if required
    
        if st.form_submit_button("Add"):
            if st.session_state.df2.shape[0] == num_new_rows:
                st.error("Add row limit reached. Cant add any more records..")
            else:
                rw = st.session_state.df2.shape[0] + 1
                st.info(f"Row: {rw} / {num_new_rows} added")
                st.session_state.df2.loc[rw] = rwdta
    
                if st.session_state.df2.shape[0] == num_new_rows:
                    st.error("Add row limit reached...")
    


    df2 = st.session_state.df2
    df2.to_csv("fleet 3/profits1.csv",encoding='utf-8',index=False)

    st.text(st.session_state.df2)
 
   with st.expander("Vessel fines"):
    if "df3" not in st.session_state:
        st.session_state.df3 = pd.DataFrame(columns=["Vessel 1: Margaery", 
                                                    "Vessel 2: Brienne", 
                                                    "Vessel 3: Lysa"])
    
    st.markdown("Add Record")
    
    ncol = st.session_state.df3.shape[1]  # col count
    rw = -1
    
    with st.form(key="add fines form", clear_on_submit= True):
        cols = st.columns(ncol)
        rwdta = []
    
        for i in range(ncol):
            rwdta.append(cols[i].text_input(st.session_state.df3.columns[i]))
    
        # you can insert code for a list comprehension here to change the data (rwdta) 
        # values into integer / float, if required
    
        if st.form_submit_button("Add"):
            if st.session_state.df3.shape[0] == num_new_rows:
                st.error("Add row limit reached. Cant add any more records..")
            else:
                rw = st.session_state.df3.shape[0] + 1
                st.info(f"Row: {rw} / {num_new_rows} added")
                st.session_state.df3.loc[rw] = rwdta
    
                if st.session_state.df3.shape[0] == num_new_rows:
                    st.error("Add row limit reached...")
    

    df3 = st.session_state.df3
    df3.to_csv("fleet 3/fines1.csv",encoding='utf-8',index=False)

    st.text(st.session_state.df3)

###############################################################################
#2.i Overall reports                                                          #
###############################################################################
   st.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)
   if st.button('Create overall reports'): 
    st.header("**Day 1: Overall reports**")
    st.markdown('''In this subsection the fleet manager can see the current accumulted score and profit of day 1 and the daily fleet scores and profits.''')
    st.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)
    #
    df = pd.read_csv(os.path.join(dir_main,'fleet 3/score1.csv'), sep=',')
    df2 = pd.read_csv(os.path.join(dir_main,'fleet 3/profits1.csv'), sep=',')
    df3 = pd.read_csv(os.path.join(dir_main,'fleet 3/fines1.csv'), sep=',')
    d = {'Vessel': df.columns[0:3] , 'Score':  df.iloc[-1] , 'Accumulated Profit':df2.iloc[-1] ,'Fines':df3.iloc[-1] }
    outcomes = pd.DataFrame(data=d)
    outcomes=outcomes.reset_index(drop=True)
    st.dataframe(outcomes)
    
    wb = openpyxl.Workbook()
    wb.save(os.path.join(dir_main,'fleet 3/reports1.xlsx')) 
    book = load_workbook(os.path.join(dir_main,'fleet 3/reports1.xlsx'))
    del book['Sheet']   

    writer = pd.ExcelWriter(os.path.join(dir_main,'fleet 3/reports1.xlsx'), engine='openpyxl')
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}
    outcomes.to_excel(writer, sheet_name = "Outcomes 1", na_rep = "", index =   False, startrow = 0)  
    writer.save()

    #Plot scores 
    df01 = pd.read_csv(os.path.join(dir_main,'fleet 3/score1.csv'), sep=',')
    df01.index = np.arange(1, len(df01) + 1)
    #Create the daily average 
    df01['Fleet score']=df01.sum(axis=1)/3
    df01['Action'] = df01.index
    df01 = pd.melt(df01, id_vars='Action', value_vars=["Vessel 1: Margaery", 
                                                    "Vessel 2: Brienne", 
                                                    "Vessel 3: Lysa", 
                                                    "Fleet score"])
    st.subheader("*Daily fleet and vessel scores*")
    fig01 = px.line(df01, x='Action', y='value', color='variable', markers=True,range_x=[0,10],width=650,height=650)
    fig01.update_layout(yaxis_type='category')
    fig01.layout.xaxis.fixedrange= True
    fig01.layout.yaxis.fixedrange: True
    fig01.update(layout_showlegend=True)
    fig01.update_traces(patch={"line": {"color": "green", "width": 3, "dash": 'dot'}}, selector={"legendgroup": "Fleet score"}) 
    st.plotly_chart(fig01)
  
    #Plot profits  
    df02 = pd.read_csv(os.path.join(dir_main,'fleet 3/profits1.csv'), sep=',')
    df03 = pd.read_csv(os.path.join(dir_main,'fleet 3/fines1.csv'), sep=',')
    df02.index = np.arange(1, len(df02) + 1)
    df03.index = np.arange(1, len(df03) + 1)

    df02['Action'] = df02.index
    df02['Fleet profit']=df02.sum(axis=1)
    aux = df01.loc[df01.variable=="Fleet score",'value'].reset_index(drop=True)
    aux.index = np.arange(1, len(aux) + 1)
    df02['Fleet score'] = aux
    
    df03['Fleet fines']=df03.sum(axis=1)
    df02['Fleet w. score'] = df02['Fleet profit'] - df03['Fleet fines']
    
    df02 = pd.melt(df02, id_vars='Action', value_vars=["Fleet profit",
                                                       "Fleet score",
                                                       "Fleet w. score"])
    df02_1 = df02.loc[df02.variable == "Fleet score",:]
    df02_2 = df02.loc[df02.variable != "Fleet score",:]
    
    # Create figure with secondary y-axis 
    subfig = make_subplots(specs=[[{"secondary_y": True}]])

    # Defining Custom Colors
    colours = {
      "Fleet profit": "#A1A9A3",
      "Fleet w. score": "#DFDF70",
    }

    # create two independent figures with px.line each containing data from multiple columns
    fig = px.line(df02_1, x='Action', y='value',markers=True,range_x=[1,10],width=650,height=650, text="value")
    fig.layout.xaxis.fixedrange= True
    fig.update_traces(line_color='orange', line_width=2)
    fig2 = px.bar(df02_2, y='value', x='Action', color='variable',range_x=[1,10],width=650,height=650,color_discrete_map=colours)
    fig.update_traces(yaxis="y2")
    fig2.update(layout_showlegend=True)
    
    subfig.add_traces(fig2.data + fig.data)
    subfig.layout.xaxis.fixedrange= True
    
    # Set y-axes titles
    subfig.layout.yaxis.title="Kr."
    subfig.layout.yaxis2.title="Score"
    subfig.update_xaxes(range=[0,10], dtick=1)
    st.subheader("*Fleet profit loss by score increase*")
    st.plotly_chart(subfig)
  
    #Save the data for next day 
    cav=open(dir_main+"fleet 3/reports1.xlsx",'rb')
    b64_01 = base64.b64encode(cav.read()).decode('utf-8')
    href = f'<a href="data:file/cav;base64,{b64_01}" download="results1.xlsx">:arrow_down_small: Overall results day 1 </a>'
    st.markdown(href, unsafe_allow_html=True)

###############################################################################
#2.i Day 2                                                                    #
###############################################################################
  if pag_input=="Day 2": 
    #Selection of section in the daily report
   # Title
   st.title(title)  
   st.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)
   st.header("**Day 2: Type in information**")
   st.markdown('''In this subsection the fleet manager types in the score and the profit for every action (this is already
accumulated)''')

   with st.expander("Vessel scores"):
#    if "df" not in st.session_state or score.columns[0]=="Unnamed: 0":
    if "df" not in st.session_state:
        st.session_state.df = pd.DataFrame(columns=["Vessel 1: Margaery", 
                                                    "Vessel 2: Brienne", 
                                                    "Vessel 3: Lysa"])


    st.markdown("Add Record")
    
    num_new_rows = 10
    ncol = st.session_state.df.shape[1]  # col count
    rw = -1
    
    with st.form(key="add form", clear_on_submit= True):
        cols = st.columns(ncol)
        rwdta = []
    
        for i in range(ncol):
            rwdta.append(cols[i].text_input(st.session_state.df.columns[i]))
    
        # you can insert code for a list comprehension here to change the data (rwdta) 
        # values into integer / float, if required
    
        if st.form_submit_button("Add"):
            if st.session_state.df.shape[0] == num_new_rows:
                st.error("Add row limit reached. Cant add any more records..")
            else:
                rw = st.session_state.df.shape[0] + 1
                st.info(f"Row: {rw} / {num_new_rows} added")
                st.session_state.df.loc[rw] = rwdta
    
                if st.session_state.df.shape[0] == num_new_rows:
                    st.error("Add row limit reached...")
    df = st.session_state.df
    df.to_csv("fleet 3/score2.csv",encoding='utf-8',index=False)
    st.text(df)
    
   with st.expander("Vessel profits"):
    if "df2" not in st.session_state:
        st.session_state.df2 = pd.DataFrame(columns=["Vessel 1: Margaery", 
                                                    "Vessel 2: Brienne", 
                                                    "Vessel 3: Lysa"])
    
    st.markdown("Add Record")
    
    ncol = st.session_state.df2.shape[1]  # col count
    rw = -1
    
    with st.form(key="add profits form", clear_on_submit= True):
        cols = st.columns(ncol)
        rwdta = []
    
        for i in range(ncol):
            rwdta.append(cols[i].text_input(st.session_state.df2.columns[i]))
    
        # you can insert code for a list comprehension here to change the data (rwdta) 
        # values into integer / float, if required
    
        if st.form_submit_button("Add"):
            if st.session_state.df2.shape[0] == num_new_rows:
                st.error("Add row limit reached. Cant add any more records..")
            else:
                rw = st.session_state.df2.shape[0] + 1
                st.info(f"Row: {rw} / {num_new_rows} added")
                st.session_state.df2.loc[rw] = rwdta
    
                if st.session_state.df2.shape[0] == num_new_rows:
                    st.error("Add row limit reached...")
    


    df2 = st.session_state.df2
    df2.to_csv("fleet 3/profits2.csv",encoding='utf-8',index=False)

    st.text(st.session_state.df2)
 
   with st.expander("Vessel fines"):
    if "df3" not in st.session_state:
        st.session_state.df3 = pd.DataFrame(columns=["Vessel 1: Margaery", 
                                                    "Vessel 2: Brienne", 
                                                    "Vessel 3: Lysa"])
    
    st.markdown("Add Record")
    
    ncol = st.session_state.df3.shape[1]  # col count
    rw = -1
    
    with st.form(key="add fines form", clear_on_submit= True):
        cols = st.columns(ncol)
        rwdta = []
    
        for i in range(ncol):
            rwdta.append(cols[i].text_input(st.session_state.df3.columns[i]))
    
        # you can insert code for a list comprehension here to change the data (rwdta) 
        # values into integer / float, if required
    
        if st.form_submit_button("Add"):
            if st.session_state.df3.shape[0] == num_new_rows:
                st.error("Add row limit reached. Cant add any more records..")
            else:
                rw = st.session_state.df3.shape[0] + 1
                st.info(f"Row: {rw} / {num_new_rows} added")
                st.session_state.df3.loc[rw] = rwdta
    
                if st.session_state.df3.shape[0] == num_new_rows:
                    st.error("Add row limit reached...")
    

    df3 = st.session_state.df3
    df3.to_csv("fleet 3/fines2.csv",encoding='utf-8',index=False)

    st.text(st.session_state.df3)

###############################################################################
#2.i Overall reports                                                          #
###############################################################################
   st.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)
   if st.button('Create overall reports'): 
    st.header("**Day 2: Overall reports**")
    st.markdown('''In this subsection the fleet manager can see the current accumulated score and profit of day 2 and the daily fleet scores and profits.''')
    st.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)
    #
    df = pd.read_csv(os.path.join(dir_main,'fleet 3/score2.csv'), sep=',')
    df2 = pd.read_csv(os.path.join(dir_main,'fleet 3/profits2.csv'), sep=',')
    df3 = pd.read_csv(os.path.join(dir_main,'fleet 3/fines2.csv'), sep=',')
    df4 = pd.read_excel(os.path.join(dir_main,'fleet 3/results1.xlsx'))
    df4.index = df4.Vessel
    
    d = {'Vessel': df.columns[0:3] , 
         'Score':  df.iloc[-1] , 
         'Score t-1' : df4.Score,
         'Accumulated Profit':df2.iloc[-1] ,
         'Fines':df3.iloc[-1] }
    outcomes = pd.DataFrame(data=d)
    outcomes=outcomes.reset_index(drop=True)
    st.dataframe(outcomes)

    wb = openpyxl.Workbook()
    wb.save(os.path.join(dir_main,'fleet 3/reports2.xlsx')) 
    book = load_workbook(os.path.join(dir_main,'fleet 3/reports2.xlsx'))
    del book['Sheet']   

    writer = pd.ExcelWriter(os.path.join(dir_main,'fleet 3/reports2.xlsx'), engine='openpyxl')
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}
    outcomes.to_excel(writer, sheet_name = "Outcomes 2", na_rep = "", index =   False, startrow = 0)  
    writer.save()
    
    #Plot scores 
    df01 = pd.read_csv(os.path.join(dir_main,'fleet 3/score2.csv'), sep=',')
    df01.index = np.arange(1, len(df01) + 1)
    #Create the daily average 
    df01['Fleet score']=df01.sum(axis=1)/3
    df01['Action'] = df01.index
    df01 = pd.melt(df01, id_vars='Action', value_vars=["Vessel 1: Margaery", 
                                                    "Vessel 2: Brienne", 
                                                    "Vessel 3: Lysa", 
                                                    "Fleet score"])
    st.subheader("*Daily fleet and vessel scores*")
    fig01 = px.line(df01, x='Action', y='value', color='variable', markers=True,range_x=[0,10],width=650,height=650)
    fig01.update_layout(yaxis_type='category')
    fig01.layout.xaxis.fixedrange= True
    fig01.layout.yaxis.fixedrange: True
    fig01.update(layout_showlegend=True)
    fig01.update_traces(patch={"line": {"color": "green", "width": 3, "dash": 'dot'}}, selector={"legendgroup": "Fleet score"}) 
    st.plotly_chart(fig01)
  
    #Plot profits  
    df02 = pd.read_csv(os.path.join(dir_main,'fleet 3/profits2.csv'), sep=',')
    df03 = pd.read_csv(os.path.join(dir_main,'fleet 3/fines2.csv'), sep=',')
    df02.index = np.arange(1, len(df02) + 1)
    df03.index = np.arange(1, len(df03) + 1)

    df02['Action'] = df02.index
    df02['Fleet profit']=df02.sum(axis=1)
    aux = df01.loc[df01.variable=="Fleet score",'value'].reset_index(drop=True)
    aux.index = np.arange(1, len(aux) + 1)
    df02['Fleet score'] = aux
    
    df03['Fleet fines']=df03.sum(axis=1)
    df02['Fleet w. score'] = df02['Fleet profit'] - df03['Fleet fines']
    
    df02 = pd.melt(df02, id_vars='Action', value_vars=["Fleet profit",
                                                       "Fleet score",
                                                       "Fleet w. score"])
    df02_1 = df02.loc[df02.variable == "Fleet score",:]
    df02_2 = df02.loc[df02.variable != "Fleet score",:]
    
    # Create figure with secondary y-axis 
    subfig = make_subplots(specs=[[{"secondary_y": True}]])

    # Defining Custom Colors
    colours = {
      "Fleet profit": "#A1A9A3",
      "Fleet w. score": "#DFDF70",
    }

    # create two independent figures with px.line each containing data from multiple columns
    fig = px.line(df02_1, x='Action', y='value',markers=True,range_x=[1,10],width=650,height=650, text="value")
    fig.layout.xaxis.fixedrange= True
    fig.update_traces(line_color='orange', line_width=2)
    fig2 = px.bar(df02_2, y='value', x='Action', color='variable',range_x=[1,10],width=650,height=650,color_discrete_map=colours)
    fig.update_traces(yaxis="y2")
    fig2.update(layout_showlegend=True)
    
    subfig.add_traces(fig2.data + fig.data)
    subfig.layout.xaxis.fixedrange= True
    
    # Set y-axes titles
    subfig.layout.yaxis.title="Kr."
    subfig.layout.yaxis2.title="Score"
    subfig.update_xaxes(range=[0,10], dtick=1)
    st.subheader("*Fleet profit loss by score increase*")
    st.plotly_chart(subfig)
  
     #Save the data for next day 
    cav=open(dir_main+"fleet 3/reports2.xlsx",'rb')
    b64_01 = base64.b64encode(cav.read()).decode('utf-8')
    href = f'<a href="data:file/cav;base64,{b64_01}" download="results2.xlsx">:arrow_down_small: Overall results day 2 </a>'
    st.markdown(href, unsafe_allow_html=True)
 
###############################################################################
#2.i Day 3                                                                    #
###############################################################################
  if pag_input=="Day 3": 
    #Selection of section in the daily report
   # Title
   st.title(title)  
   st.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)
   st.header("**Day 3: Type in information**")
   st.markdown('''In this subsection the fleet manager types in the score and the profit for every action (this is already
accumulated)''')

   with st.expander("Vessel scores"):
#    if "df" not in st.session_state or score.columns[0]=="Unnamed: 0":
    if "df" not in st.session_state:
        st.session_state.df = pd.DataFrame(columns=["Vessel 1: Margaery", 
                                                    "Vessel 2: Brienne", 
                                                    "Vessel 3: Lysa"])


    st.markdown("Add Record")
    
    num_new_rows = 10
    ncol = st.session_state.df.shape[1]  # col count
    rw = -1
    
    with st.form(key="add form", clear_on_submit= True):
        cols = st.columns(ncol)
        rwdta = []
    
        for i in range(ncol):
            rwdta.append(cols[i].text_input(st.session_state.df.columns[i]))
    
        # you can insert code for a list comprehension here to change the data (rwdta) 
        # values into integer / float, if required
    
        if st.form_submit_button("Add"):
            if st.session_state.df.shape[0] == num_new_rows:
                st.error("Add row limit reached. Cant add any more records..")
            else:
                rw = st.session_state.df.shape[0] + 1
                st.info(f"Row: {rw} / {num_new_rows} added")
                st.session_state.df.loc[rw] = rwdta
    
                if st.session_state.df.shape[0] == num_new_rows:
                    st.error("Add row limit reached...")
    df = st.session_state.df
    df.to_csv("fleet 3/score3.csv",encoding='utf-8',index=False)
    st.text(df)
    
   with st.expander("Vessel profits"):
    if "df2" not in st.session_state:
        st.session_state.df2 = pd.DataFrame(columns=["Vessel 1: Margaery", 
                                                    "Vessel 2: Brienne", 
                                                    "Vessel 3: Lysa"])
    
    st.markdown("Add Record")
    
    ncol = st.session_state.df2.shape[1]  # col count
    rw = -1
    
    with st.form(key="add profits form", clear_on_submit= True):
        cols = st.columns(ncol)
        rwdta = []
    
        for i in range(ncol):
            rwdta.append(cols[i].text_input(st.session_state.df2.columns[i]))
    
        # you can insert code for a list comprehension here to change the data (rwdta) 
        # values into integer / float, if required
    
        if st.form_submit_button("Add"):
            if st.session_state.df2.shape[0] == num_new_rows:
                st.error("Add row limit reached. Cant add any more records..")
            else:
                rw = st.session_state.df2.shape[0] + 1
                st.info(f"Row: {rw} / {num_new_rows} added")
                st.session_state.df2.loc[rw] = rwdta
    
                if st.session_state.df2.shape[0] == num_new_rows:
                    st.error("Add row limit reached...")
    


    df2 = st.session_state.df2
    df2.to_csv("fleet 3/profits3.csv",encoding='utf-8',index=False)

    st.text(st.session_state.df2)
 
   with st.expander("Vessel fines"):
    if "df3" not in st.session_state:
        st.session_state.df3 = pd.DataFrame(columns=["Vessel 1: Margaery", 
                                                    "Vessel 2: Brienne", 
                                                    "Vessel 3: Lysa"])
    
    st.markdown("Add Record")
    
    ncol = st.session_state.df3.shape[1]  # col count
    rw = -1
    
    with st.form(key="add fines form", clear_on_submit= True):
        cols = st.columns(ncol)
        rwdta = []
    
        for i in range(ncol):
            rwdta.append(cols[i].text_input(st.session_state.df3.columns[i]))
    
        # you can insert code for a list comprehension here to change the data (rwdta) 
        # values into integer / float, if required
    
        if st.form_submit_button("Add"):
            if st.session_state.df3.shape[0] == num_new_rows:
                st.error("Add row limit reached. Cant add any more records..")
            else:
                rw = st.session_state.df3.shape[0] + 1
                st.info(f"Row: {rw} / {num_new_rows} added")
                st.session_state.df3.loc[rw] = rwdta
    
                if st.session_state.df3.shape[0] == num_new_rows:
                    st.error("Add row limit reached...")
    

    df3 = st.session_state.df3
    df3.to_csv("fleet 3/fines3.csv",encoding='utf-8',index=False)

    st.text(st.session_state.df3)

###############################################################################
#2.i Overall reports                                                          #
###############################################################################
   st.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)
   if st.button('Create overall reports'): 
    st.header("**Day 3: Overall reports**")
    st.markdown('''In this subsection the fleet manager can see the current accumulated score and profit of day 3 and the daily fleet scores and profits.''')
    st.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)
    #
    df = pd.read_csv(os.path.join(dir_main,'fleet 3/score3.csv'), sep=',')
    df2 = pd.read_csv(os.path.join(dir_main,'fleet 3/profits3.csv'), sep=',')
    df3 = pd.read_csv(os.path.join(dir_main,'fleet 3/fines3.csv'), sep=',')
    df4 = pd.read_excel(os.path.join(dir_main,'fleet 3/results2.xlsx'))
    df4.index = df4.Vessel
    df5 = pd.read_excel(os.path.join(dir_main,'fleet 3/results1.xlsx'))
    df5.index = df5.Vessel
    
    d = {'Vessel': df.columns[0:3] , 
         'Score':  df.iloc[-1] , 
         'Score t-1' : df4.Score,
         'Score t-2' : df5.Score,
         'Accumulated Profit':df2.iloc[-1] ,
         'Fines':df3.iloc[-1] }
    outcomes = pd.DataFrame(data=d)
    outcomes=outcomes.reset_index(drop=True)
    st.dataframe(outcomes)

    wb = openpyxl.Workbook()
    wb.save(os.path.join(dir_main,'fleet 3/reports3.xlsx')) 
    book = load_workbook(os.path.join(dir_main,'fleet 3/reports3.xlsx'))
    del book['Sheet']   

    writer = pd.ExcelWriter(os.path.join(dir_main,'fleet 3/reports3.xlsx'), engine='openpyxl')
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}
    outcomes.to_excel(writer, sheet_name = "Outcomes 3", na_rep = "", index =   False, startrow = 0)  
    writer.save()
    
    #Plot scores 
    df01 = pd.read_csv(os.path.join(dir_main,'fleet 3/score3.csv'), sep=',')
    df01.index = np.arange(1, len(df01) + 1)
    #Create the daily average 
    df01['Fleet score']=df01.sum(axis=1)/3
    df01['Action'] = df01.index
    df01 = pd.melt(df01, id_vars='Action', value_vars=["Vessel 1: Margaery", 
                                                    "Vessel 2: Brienne", 
                                                    "Vessel 3: Lysa", 
                                                    "Fleet score"])
    st.subheader("*Daily fleet and vessel scores*")
    fig01 = px.line(df01, x='Action', y='value', color='variable', markers=True,range_x=[0,10],width=650,height=650)
    fig01.update_layout(yaxis_type='category')
    fig01.layout.xaxis.fixedrange= True
    fig01.layout.yaxis.fixedrange: True
    fig01.update(layout_showlegend=True)
    fig01.update_traces(patch={"line": {"color": "green", "width": 3, "dash": 'dot'}}, selector={"legendgroup": "Fleet score"}) 
    st.plotly_chart(fig01)
  
    #Plot profits  
    df02 = pd.read_csv(os.path.join(dir_main,'fleet 3/profits3.csv'), sep=',')
    df03 = pd.read_csv(os.path.join(dir_main,'fleet 3/fines3.csv'), sep=',')
    df02.index = np.arange(1, len(df02) + 1)
    df03.index = np.arange(1, len(df03) + 1)

    df02['Action'] = df02.index
    df02['Fleet profit']=df02.sum(axis=1)
    aux = df01.loc[df01.variable=="Fleet score",'value'].reset_index(drop=True)
    aux.index = np.arange(1, len(aux) + 1)
    df02['Fleet score'] = aux
    
    df03['Fleet fines']=df03.sum(axis=1)
    df02['Fleet w. score'] = df02['Fleet profit'] - df03['Fleet fines']
    
    df02 = pd.melt(df02, id_vars='Action', value_vars=["Fleet profit",
                                                       "Fleet score",
                                                       "Fleet w. score"])
    df02_1 = df02.loc[df02.variable == "Fleet score",:]
    df02_2 = df02.loc[df02.variable != "Fleet score",:]
    
    # Create figure with secondary y-axis 
    subfig = make_subplots(specs=[[{"secondary_y": True}]])

    # Defining Custom Colors
    colours = {
      "Fleet profit": "#A1A9A3",
      "Fleet w. score": "#DFDF70",
    }

    # create two independent figures with px.line each containing data from multiple columns
    fig = px.line(df02_1, x='Action', y='value',markers=True,range_x=[1,10],width=650,height=650, text="value")
    fig.layout.xaxis.fixedrange= True
    fig.update_traces(line_color='orange', line_width=2)
    fig2 = px.bar(df02_2, y='value', x='Action', color='variable',range_x=[1,10],width=650,height=650,color_discrete_map=colours)
    fig.update_traces(yaxis="y2")
    fig2.update(layout_showlegend=True)
    
    subfig.add_traces(fig2.data + fig.data)
    subfig.layout.xaxis.fixedrange= True
    
    # Set y-axes titles
    subfig.layout.yaxis.title="Kr."
    subfig.layout.yaxis2.title="Score"
    subfig.update_xaxes(range=[0,10], dtick=1)
    st.subheader("*Fleet profit loss by score increase*")
    st.plotly_chart(subfig)
  
     #Save the data for next day 
    cav=open(dir_main+"fleet 3/reports3.xlsx",'rb')
    b64_01 = base64.b64encode(cav.read()).decode('utf-8')
    href = f'<a href="data:file/cav;base64,{b64_01}" download="results3.xlsx">:arrow_down_small: Overall results day 3 </a>'
    st.markdown(href, unsafe_allow_html=True)
  
###############################################################################
#2.i Day 4                                                                    #
###############################################################################
  if pag_input=="Day 4": 
    #Selection of section in the daily report
   # Title
   st.title(title)  
   st.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)
   st.header("**Day 4: Type in information**")
   st.markdown('''In this subsection the fleet manager types in the score and the profit for every action (this is already
accumulated)''')

   with st.expander("Vessel scores"):
#    if "df" not in st.session_state or score.columns[0]=="Unnamed: 0":
    if "df" not in st.session_state:
        st.session_state.df = pd.DataFrame(columns=["Vessel 1: Margaery", 
                                                    "Vessel 2: Brienne", 
                                                    "Vessel 3: Lysa"])


    st.markdown("Add Record")
    
    num_new_rows = 10
    ncol = st.session_state.df.shape[1]  # col count
    rw = -1
    
    with st.form(key="add form", clear_on_submit= True):
        cols = st.columns(ncol)
        rwdta = []
    
        for i in range(ncol):
            rwdta.append(cols[i].text_input(st.session_state.df.columns[i]))
    
        # you can insert code for a list comprehension here to change the data (rwdta) 
        # values into integer / float, if required
    
        if st.form_submit_button("Add"):
            if st.session_state.df.shape[0] == num_new_rows:
                st.error("Add row limit reached. Cant add any more records..")
            else:
                rw = st.session_state.df.shape[0] + 1
                st.info(f"Row: {rw} / {num_new_rows} added")
                st.session_state.df.loc[rw] = rwdta
    
                if st.session_state.df.shape[0] == num_new_rows:
                    st.error("Add row limit reached...")
    df = st.session_state.df
    df.to_csv("fleet 3/score4.csv",encoding='utf-8',index=False)
    st.text(df)
    
   with st.expander("Vessel profits"):
    if "df2" not in st.session_state:
        st.session_state.df2 = pd.DataFrame(columns=["Vessel 1: Margaery", 
                                                    "Vessel 2: Brienne", 
                                                    "Vessel 3: Lysa"])
    
    st.markdown("Add Record")
    
    ncol = st.session_state.df2.shape[1]  # col count
    rw = -1
    
    with st.form(key="add profits form", clear_on_submit= True):
        cols = st.columns(ncol)
        rwdta = []
    
        for i in range(ncol):
            rwdta.append(cols[i].text_input(st.session_state.df2.columns[i]))
    
        # you can insert code for a list comprehension here to change the data (rwdta) 
        # values into integer / float, if required
    
        if st.form_submit_button("Add"):
            if st.session_state.df2.shape[0] == num_new_rows:
                st.error("Add row limit reached. Cant add any more records..")
            else:
                rw = st.session_state.df2.shape[0] + 1
                st.info(f"Row: {rw} / {num_new_rows} added")
                st.session_state.df2.loc[rw] = rwdta
    
                if st.session_state.df2.shape[0] == num_new_rows:
                    st.error("Add row limit reached...")
    


    df2 = st.session_state.df2
    df2.to_csv("fleet 3/profits4.csv",encoding='utf-8',index=False)

    st.text(st.session_state.df2)
 
   with st.expander("Vessel fines"):
    if "df3" not in st.session_state:
        st.session_state.df3 = pd.DataFrame(columns=["Vessel 1: Margaery", 
                                                    "Vessel 2: Brienne", 
                                                    "Vessel 3: Lysa"])
    
    st.markdown("Add Record")
    
    ncol = st.session_state.df3.shape[1]  # col count
    rw = -1
    
    with st.form(key="add fines form", clear_on_submit= True):
        cols = st.columns(ncol)
        rwdta = []
    
        for i in range(ncol):
            rwdta.append(cols[i].text_input(st.session_state.df3.columns[i]))
    
        # you can insert code for a list comprehension here to change the data (rwdta) 
        # values into integer / float, if required
    
        if st.form_submit_button("Add"):
            if st.session_state.df3.shape[0] == num_new_rows:
                st.error("Add row limit reached. Cant add any more records..")
            else:
                rw = st.session_state.df3.shape[0] + 1
                st.info(f"Row: {rw} / {num_new_rows} added")
                st.session_state.df3.loc[rw] = rwdta
    
                if st.session_state.df3.shape[0] == num_new_rows:
                    st.error("Add row limit reached...")
    

    df3 = st.session_state.df3
    df3.to_csv("fleet 3/fines4.csv",encoding='utf-8',index=False)

    st.text(st.session_state.df3)

###############################################################################
#2.i Overall reports                                                          #
###############################################################################
   st.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)
   if st.button('Create overall reports'): 
    st.header("**Day 4: Overall reports**")
    st.markdown('''In this subsection the fleet manager can see the current accumulated score and profit of day 4 and the daily fleet scores and profits.''')
    st.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)
    #
    df = pd.read_csv(os.path.join(dir_main,'fleet 3/score4.csv'), sep=',')
    df2 = pd.read_csv(os.path.join(dir_main,'fleet 3/profits4.csv'), sep=',')
    df3 = pd.read_csv(os.path.join(dir_main,'fleet 3/fines4.csv'), sep=',')
    df4 = pd.read_excel(os.path.join(dir_main,'fleet 3/results2.xlsx'))
    df4.index = df4.Vessel
    df5 = pd.read_excel(os.path.join(dir_main,'fleet 3/results3.xlsx'))
    df5.index = df5.Vessel
    
    d = {'Vessel': df.columns[0:3] , 
         'Score':  df.iloc[-1] , 
         'Score t-1' : df4.Score,
         'Score t-2' : df5.Score,
         'Accumulated Profit':df2.iloc[-1] ,
         'Fines':df3.iloc[-1] }
    outcomes = pd.DataFrame(data=d)
    outcomes=outcomes.reset_index(drop=True)
    st.dataframe(outcomes)

    wb = openpyxl.Workbook()
    wb.save(os.path.join(dir_main,'fleet 3/reports4.xlsx')) 
    book = load_workbook(os.path.join(dir_main,'fleet 3/reports4.xlsx'))
    del book['Sheet']   

    writer = pd.ExcelWriter(os.path.join(dir_main,'fleet 3/reports4.xlsx'), engine='openpyxl')
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}
    outcomes.to_excel(writer, sheet_name = "Outcomes 4", na_rep = "", index =   False, startrow = 0)  
    writer.save()
    
    #Plot scores 
    df01 = pd.read_csv(os.path.join(dir_main,'fleet 3/score4.csv'), sep=',')
    df01.index = np.arange(1, len(df01) + 1)
    #Create the daily average 
    df01['Fleet score']=df01.sum(axis=1)/3
    df01['Action'] = df01.index
    df01 = pd.melt(df01, id_vars='Action', value_vars=["Vessel 1: Margaery", 
                                                    "Vessel 2: Brienne", 
                                                    "Vessel 3: Lysa", 
                                                    "Fleet score"])
    st.subheader("*Daily fleet and vessel scores*")
    fig01 = px.line(df01, x='Action', y='value', color='variable', markers=True,range_x=[0,10],width=650,height=650)
    fig01.update_layout(yaxis_type='category')
    fig01.layout.xaxis.fixedrange= True
    fig01.layout.yaxis.fixedrange: True
    fig01.update(layout_showlegend=True)
    fig01.update_traces(patch={"line": {"color": "green", "width": 3, "dash": 'dot'}}, selector={"legendgroup": "Fleet score"}) 
    st.plotly_chart(fig01)
  
    #Plot profits  
    df02 = pd.read_csv(os.path.join(dir_main,'fleet 3/profits4.csv'), sep=',')
    df03 = pd.read_csv(os.path.join(dir_main,'fleet 3/fines4.csv'), sep=',')
    df02.index = np.arange(1, len(df02) + 1)
    df03.index = np.arange(1, len(df03) + 1)

    df02['Action'] = df02.index
    df02['Fleet profit']=df02.sum(axis=1)
    aux = df01.loc[df01.variable=="Fleet score",'value'].reset_index(drop=True)
    aux.index = np.arange(1, len(aux) + 1)
    df02['Fleet score'] = aux
    
    df03['Fleet fines']=df03.sum(axis=1)
    df02['Fleet w. score'] = df02['Fleet profit'] - df03['Fleet fines']
    
    df02 = pd.melt(df02, id_vars='Action', value_vars=["Fleet profit",
                                                       "Fleet score",
                                                       "Fleet w. score"])
    df02_1 = df02.loc[df02.variable == "Fleet score",:]
    df02_2 = df02.loc[df02.variable != "Fleet score",:]
    
    # Create figure with secondary y-axis 
    subfig = make_subplots(specs=[[{"secondary_y": True}]])

    # Defining Custom Colors
    colours = {
      "Fleet profit": "#A1A9A3",
      "Fleet w. score": "#DFDF70",
    }

    # create two independent figures with px.line each containing data from multiple columns
    fig = px.line(df02_1, x='Action', y='value',markers=True,range_x=[1,10],width=650,height=650, text="value")
    fig.layout.xaxis.fixedrange= True
    fig.update_traces(line_color='orange', line_width=2)
    fig2 = px.bar(df02_2, y='value', x='Action', color='variable',range_x=[1,10],width=650,height=650,color_discrete_map=colours)
    fig.update_traces(yaxis="y2")
    fig2.update(layout_showlegend=True)
    
    subfig.add_traces(fig2.data + fig.data)
    subfig.layout.xaxis.fixedrange= True
    
    # Set y-axes titles
    subfig.layout.yaxis.title="Kr."
    subfig.layout.yaxis2.title="Score"
    subfig.update_xaxes(range=[0,10], dtick=1)
    st.subheader("*Fleet profit loss by score increase*")
    st.plotly_chart(subfig)
  
     #Save the data for next day 
    cav=open(dir_main+"fleet 3/reports4.xlsx",'rb')
    b64_01 = base64.b64encode(cav.read()).decode('utf-8')
    href = f'<a href="data:file/cav;base64,{b64_01}" download="results4.xlsx">:arrow_down_small: Overall results day 4 </a>'
    st.markdown(href, unsafe_allow_html=True)
  
###############################################################################
#2.i Day 5                                                                    #
###############################################################################
  if pag_input=="Day 5": 
    #Selection of section in the daily report
   # Title
   st.title(title)  
   st.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)
   st.header("**Day 5: Type in information**")
   st.markdown('''In this subsection the fleet manager types in the score and the profit for every action (this is already
accumulated)''')

   with st.expander("Vessel scores"):
#    if "df" not in st.session_state or score.columns[0]=="Unnamed: 0":
    if "df" not in st.session_state:
        st.session_state.df = pd.DataFrame(columns=["Vessel 1: Margaery", 
                                                    "Vessel 2: Brienne", 
                                                    "Vessel 3: Lysa"])


    st.markdown("Add Record")
    
    num_new_rows = 10
    ncol = st.session_state.df.shape[1]  # col count
    rw = -1
    
    with st.form(key="add form", clear_on_submit= True):
        cols = st.columns(ncol)
        rwdta = []
    
        for i in range(ncol):
            rwdta.append(cols[i].text_input(st.session_state.df.columns[i]))
    
        # you can insert code for a list comprehension here to change the data (rwdta) 
        # values into integer / float, if required
    
        if st.form_submit_button("Add"):
            if st.session_state.df.shape[0] == num_new_rows:
                st.error("Add row limit reached. Cant add any more records..")
            else:
                rw = st.session_state.df.shape[0] + 1
                st.info(f"Row: {rw} / {num_new_rows} added")
                st.session_state.df.loc[rw] = rwdta
    
                if st.session_state.df.shape[0] == num_new_rows:
                    st.error("Add row limit reached...")
    df = st.session_state.df
    df.to_csv("fleet 3/score5.csv",encoding='utf-8',index=False)
    st.text(df)
    
   with st.expander("Vessel profits"):
    if "df2" not in st.session_state:
        st.session_state.df2 = pd.DataFrame(columns=["Vessel 1: Margaery", 
                                                    "Vessel 2: Brienne", 
                                                    "Vessel 3: Lysa"])
    
    st.markdown("Add Record")
    
    ncol = st.session_state.df2.shape[1]  # col count
    rw = -1
    
    with st.form(key="add profits form", clear_on_submit= True):
        cols = st.columns(ncol)
        rwdta = []
    
        for i in range(ncol):
            rwdta.append(cols[i].text_input(st.session_state.df2.columns[i]))
    
        # you can insert code for a list comprehension here to change the data (rwdta) 
        # values into integer / float, if required
    
        if st.form_submit_button("Add"):
            if st.session_state.df2.shape[0] == num_new_rows:
                st.error("Add row limit reached. Cant add any more records..")
            else:
                rw = st.session_state.df2.shape[0] + 1
                st.info(f"Row: {rw} / {num_new_rows} added")
                st.session_state.df2.loc[rw] = rwdta
    
                if st.session_state.df2.shape[0] == num_new_rows:
                    st.error("Add row limit reached...")
    


    df2 = st.session_state.df2
    df2.to_csv("fleet 3/profits5.csv",encoding='utf-8',index=False)

    st.text(st.session_state.df2)
 
   with st.expander("Vessel fines"):
    if "df3" not in st.session_state:
        st.session_state.df3 = pd.DataFrame(columns=["Vessel 1: Margaery", 
                                                    "Vessel 2: Brienne", 
                                                    "Vessel 3: Lysa"])
    
    st.markdown("Add Record")
    
    ncol = st.session_state.df3.shape[1]  # col count
    rw = -1
    
    with st.form(key="add fines form", clear_on_submit= True):
        cols = st.columns(ncol)
        rwdta = []
    
        for i in range(ncol):
            rwdta.append(cols[i].text_input(st.session_state.df3.columns[i]))
    
        # you can insert code for a list comprehension here to change the data (rwdta) 
        # values into integer / float, if required
    
        if st.form_submit_button("Add"):
            if st.session_state.df3.shape[0] == num_new_rows:
                st.error("Add row limit reached. Cant add any more records..")
            else:
                rw = st.session_state.df3.shape[0] + 1
                st.info(f"Row: {rw} / {num_new_rows} added")
                st.session_state.df3.loc[rw] = rwdta
    
                if st.session_state.df3.shape[0] == num_new_rows:
                    st.error("Add row limit reached...")
    

    df3 = st.session_state.df3
    df3.to_csv("fleet 3/fines5.csv",encoding='utf-8',index=False)

    st.text(st.session_state.df3)

###############################################################################
#2.i Overall reports                                                          #
###############################################################################
   st.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)
   if st.button('Create overall reports'): 
    st.header("**Day 5: Overall reports**")
    st.markdown('''In this subsection the fleet manager can see the current accumulated score and profit of day 5 and the daily fleet scores and profits.''')
    st.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)
    #
    df = pd.read_csv(os.path.join(dir_main,'fleet 3/score5.csv'), sep=',')
    df2 = pd.read_csv(os.path.join(dir_main,'fleet 3/profits5.csv'), sep=',')
    df3 = pd.read_csv(os.path.join(dir_main,'fleet 3/fines5.csv'), sep=',')
    df4 = pd.read_excel(os.path.join(dir_main,'fleet 3/results3.xlsx'))
    df4.index = df4.Vessel
    df5 = pd.read_excel(os.path.join(dir_main,'fleet 3/results4.xlsx'))
    df5.index = df5.Vessel
    
    d = {'Vessel': df.columns[0:3] , 
         'Score':  df.iloc[-1] , 
         'Score t-1' : df4.Score,
         'Score t-2' : df5.Score,
         'Accumulated Profit':df2.iloc[-1] ,
         'Fines':df3.iloc[-1] }
    outcomes = pd.DataFrame(data=d)
    outcomes=outcomes.reset_index(drop=True)
    st.dataframe(outcomes)

    wb = openpyxl.Workbook()
    wb.save(os.path.join(dir_main,'fleet 3/reports5.xlsx')) 
    book = load_workbook(os.path.join(dir_main,'fleet 3/reports5.xlsx'))
    del book['Sheet']   

    writer = pd.ExcelWriter(os.path.join(dir_main,'fleet 3/reports5.xlsx'), engine='openpyxl')
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}
    outcomes.to_excel(writer, sheet_name = "Outcomes 5", na_rep = "", index =   False, startrow = 0)  
    writer.save()
    
    #Plot scores 
    df01 = pd.read_csv(os.path.join(dir_main,'fleet 3/score5.csv'), sep=',')
    df01.index = np.arange(1, len(df01) + 1)
    #Create the daily average 
    df01['Fleet score']=df01.sum(axis=1)/3
    df01['Action'] = df01.index
    df01 = pd.melt(df01, id_vars='Action', value_vars=["Vessel 1: Margaery", 
                                                    "Vessel 2: Brienne", 
                                                    "Vessel 3: Lysa", 
                                                    "Fleet score"])
    st.subheader("*Daily fleet and vessel scores*")
    fig01 = px.line(df01, x='Action', y='value', color='variable', markers=True,range_x=[0,10],width=650,height=650)
    fig01.update_layout(yaxis_type='category')
    fig01.layout.xaxis.fixedrange= True
    fig01.layout.yaxis.fixedrange: True
    fig01.update(layout_showlegend=True)
    fig01.update_traces(patch={"line": {"color": "green", "width": 3, "dash": 'dot'}}, selector={"legendgroup": "Fleet score"}) 
    st.plotly_chart(fig01)
  
    #Plot profits  
    df02 = pd.read_csv(os.path.join(dir_main,'fleet 3/profits5.csv'), sep=',')
    df03 = pd.read_csv(os.path.join(dir_main,'fleet 3/fines5.csv'), sep=',')
    df02.index = np.arange(1, len(df02) + 1)
    df03.index = np.arange(1, len(df03) + 1)

    df02['Action'] = df02.index
    df02['Fleet profit']=df02.sum(axis=1)
    aux = df01.loc[df01.variable=="Fleet score",'value'].reset_index(drop=True)
    aux.index = np.arange(1, len(aux) + 1)
    df02['Fleet score'] = aux
    
    df03['Fleet fines']=df03.sum(axis=1)
    df02['Fleet w. score'] = df02['Fleet profit'] - df03['Fleet fines']
    
    df02 = pd.melt(df02, id_vars='Action', value_vars=["Fleet profit",
                                                       "Fleet score",
                                                       "Fleet w. score"])
    df02_1 = df02.loc[df02.variable == "Fleet score",:]
    df02_2 = df02.loc[df02.variable != "Fleet score",:]
    
    # Create figure with secondary y-axis 
    subfig = make_subplots(specs=[[{"secondary_y": True}]])

    # Defining Custom Colors
    colours = {
      "Fleet profit": "#A1A9A3",
      "Fleet w. score": "#DFDF70",
    }

    # create two independent figures with px.line each containing data from multiple columns
    fig = px.line(df02_1, x='Action', y='value',markers=True,range_x=[1,10],width=650,height=650, text="value")
    fig.layout.xaxis.fixedrange= True
    fig.update_traces(line_color='orange', line_width=2)
    fig2 = px.bar(df02_2, y='value', x='Action', color='variable',range_x=[1,10],width=650,height=650,color_discrete_map=colours)
    fig.update_traces(yaxis="y2")
    fig2.update(layout_showlegend=True)
    
    subfig.add_traces(fig2.data + fig.data)
    subfig.layout.xaxis.fixedrange= True
    
    # Set y-axes titles
    subfig.layout.yaxis.title="Kr."
    subfig.layout.yaxis2.title="Score"
    subfig.update_xaxes(range=[0,10], dtick=1)
    st.subheader("*Fleet profit loss by score increase*")
    st.plotly_chart(subfig)
  
     #Save the data for next day 
    cav=open(dir_main+"fleet 3/reports5.xlsx",'rb')
    b64_01 = base64.b64encode(cav.read()).decode('utf-8')
    href = f'<a href="data:file/cav;base64,{b64_01}" download="results5.xlsx">:arrow_down_small: Overall results day 5 </a>'
    st.markdown(href, unsafe_allow_html=True)
 
###############################################################################
#2.i Day 6                                                                    #
###############################################################################
  if pag_input=="Day 6": 
    #Selection of section in the daily report
   # Title
   st.title(title)  
   st.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)
   st.header("**Day 6: Type in information**")
   st.markdown('''In this subsection the fleet manager types in the score and the profit for every action (this is already
accumulated)''')

   with st.expander("Vessel scores"):
#    if "df" not in st.session_state or score.columns[0]=="Unnamed: 0":
    if "df" not in st.session_state:
        st.session_state.df = pd.DataFrame(columns=["Vessel 1: Margaery", 
                                                    "Vessel 2: Brienne", 
                                                    "Vessel 3: Lysa"])


    st.markdown("Add Record")
    
    num_new_rows = 10
    ncol = st.session_state.df.shape[1]  # col count
    rw = -1
    
    with st.form(key="add form", clear_on_submit= True):
        cols = st.columns(ncol)
        rwdta = []
    
        for i in range(ncol):
            rwdta.append(cols[i].text_input(st.session_state.df.columns[i]))
    
        # you can insert code for a list comprehension here to change the data (rwdta) 
        # values into integer / float, if required
    
        if st.form_submit_button("Add"):
            if st.session_state.df.shape[0] == num_new_rows:
                st.error("Add row limit reached. Cant add any more records..")
            else:
                rw = st.session_state.df.shape[0] + 1
                st.info(f"Row: {rw} / {num_new_rows} added")
                st.session_state.df.loc[rw] = rwdta
    
                if st.session_state.df.shape[0] == num_new_rows:
                    st.error("Add row limit reached...")
    df = st.session_state.df
    df.to_csv("fleet 3/score6.csv",encoding='utf-8',index=False)
    st.text(df)
    
   with st.expander("Vessel profits"):
    if "df2" not in st.session_state:
        st.session_state.df2 = pd.DataFrame(columns=["Vessel 1: Margaery", 
                                                    "Vessel 2: Brienne", 
                                                    "Vessel 3: Lysa"])
    
    st.markdown("Add Record")
    
    ncol = st.session_state.df2.shape[1]  # col count
    rw = -1
    
    with st.form(key="add profits form", clear_on_submit= True):
        cols = st.columns(ncol)
        rwdta = []
    
        for i in range(ncol):
            rwdta.append(cols[i].text_input(st.session_state.df2.columns[i]))
    
        # you can insert code for a list comprehension here to change the data (rwdta) 
        # values into integer / float, if required
    
        if st.form_submit_button("Add"):
            if st.session_state.df2.shape[0] == num_new_rows:
                st.error("Add row limit reached. Cant add any more records..")
            else:
                rw = st.session_state.df2.shape[0] + 1
                st.info(f"Row: {rw} / {num_new_rows} added")
                st.session_state.df2.loc[rw] = rwdta
    
                if st.session_state.df2.shape[0] == num_new_rows:
                    st.error("Add row limit reached...")
    


    df2 = st.session_state.df2
    df2.to_csv("fleet 3/profits6.csv",encoding='utf-8',index=False)

    st.text(st.session_state.df2)
 
   with st.expander("Vessel fines"):
    if "df3" not in st.session_state:
        st.session_state.df3 = pd.DataFrame(columns=["Vessel 1: Margaery", 
                                                    "Vessel 2: Brienne", 
                                                    "Vessel 3: Lysa"])
    
    st.markdown("Add Record")
    
    ncol = st.session_state.df3.shape[1]  # col count
    rw = -1
    
    with st.form(key="add fines form", clear_on_submit= True):
        cols = st.columns(ncol)
        rwdta = []
    
        for i in range(ncol):
            rwdta.append(cols[i].text_input(st.session_state.df3.columns[i]))
    
        # you can insert code for a list comprehension here to change the data (rwdta) 
        # values into integer / float, if required
    
        if st.form_submit_button("Add"):
            if st.session_state.df3.shape[0] == num_new_rows:
                st.error("Add row limit reached. Cant add any more records..")
            else:
                rw = st.session_state.df3.shape[0] + 1
                st.info(f"Row: {rw} / {num_new_rows} added")
                st.session_state.df3.loc[rw] = rwdta
    
                if st.session_state.df3.shape[0] == num_new_rows:
                    st.error("Add row limit reached...")
    

    df3 = st.session_state.df3
    df3.to_csv("fleet 3/fines6.csv",encoding='utf-8',index=False)

    st.text(st.session_state.df3)

###############################################################################
#2.i Overall reports                                                          #
###############################################################################
   st.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)
   if st.button('Create overall reports'): 
    st.header("**Day 6: Overall reports**")
    st.markdown('''In this subsection the fleet manager can see the current accumulated score and profit of day 6 and the daily fleet scores and profits.''')
    st.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)
    #
    df = pd.read_csv(os.path.join(dir_main,'fleet 3/score6.csv'), sep=',')
    df2 = pd.read_csv(os.path.join(dir_main,'fleet 3/profits6.csv'), sep=',')
    df3 = pd.read_csv(os.path.join(dir_main,'fleet 3/fines6.csv'), sep=',')
    df4 = pd.read_excel(os.path.join(dir_main,'fleet 3/results4.xlsx'))
    df4.index = df4.Vessel
    df5 = pd.read_excel(os.path.join(dir_main,'fleet 3/results5.xlsx'))
    df5.index = df5.Vessel
    
    d = {'Vessel': df.columns[0:3] , 
         'Score':  df.iloc[-1] , 
         'Score t-1' : df4.Score,
         'Score t-2' : df5.Score,
         'Accumulated Profit':df2.iloc[-1] ,
         'Fines':df3.iloc[-1] }
    outcomes = pd.DataFrame(data=d)
    outcomes=outcomes.reset_index(drop=True)
    st.dataframe(outcomes)

    wb = openpyxl.Workbook()
    wb.save(os.path.join(dir_main,'fleet 3/reports6.xlsx')) 
    book = load_workbook(os.path.join(dir_main,'fleet 3/reports6.xlsx'))
    del book['Sheet']   

    writer = pd.ExcelWriter(os.path.join(dir_main,'fleet 3/reports6.xlsx'), engine='openpyxl')
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}
    outcomes.to_excel(writer, sheet_name = "Outcomes 6", na_rep = "", index =   False, startrow = 0)  
    writer.save()
    
    #Plot scores 
    df01 = pd.read_csv(os.path.join(dir_main,'fleet 3/score6.csv'), sep=',')
    df01.index = np.arange(1, len(df01) + 1)
    #Create the daily average 
    df01['Fleet score']=df01.sum(axis=1)/3
    df01['Action'] = df01.index
    df01 = pd.melt(df01, id_vars='Action', value_vars=["Vessel 1: Margaery", 
                                                    "Vessel 2: Brienne", 
                                                    "Vessel 3: Lysa", 
                                                    "Fleet score"])
    st.subheader("*Daily fleet and vessel scores*")
    fig01 = px.line(df01, x='Action', y='value', color='variable', markers=True,range_x=[0,10],width=650,height=650)
    fig01.update_layout(yaxis_type='category')
    fig01.layout.xaxis.fixedrange= True
    fig01.layout.yaxis.fixedrange: True
    fig01.update(layout_showlegend=True)
    fig01.update_traces(patch={"line": {"color": "green", "width": 3, "dash": 'dot'}}, selector={"legendgroup": "Fleet score"}) 
    st.plotly_chart(fig01)
  
    #Plot profits  
    df02 = pd.read_csv(os.path.join(dir_main,'fleet 3/profits6.csv'), sep=',')
    df03 = pd.read_csv(os.path.join(dir_main,'fleet 3/fines6.csv'), sep=',')
    df02.index = np.arange(1, len(df02) + 1)
    df03.index = np.arange(1, len(df03) + 1)

    df02['Action'] = df02.index
    df02['Fleet profit']=df02.sum(axis=1)
    aux = df01.loc[df01.variable=="Fleet score",'value'].reset_index(drop=True)
    aux.index = np.arange(1, len(aux) + 1)
    df02['Fleet score'] = aux
    
    df03['Fleet fines']=df03.sum(axis=1)
    df02['Fleet w. score'] = df02['Fleet profit'] - df03['Fleet fines']
    
    df02 = pd.melt(df02, id_vars='Action', value_vars=["Fleet profit",
                                                       "Fleet score",
                                                       "Fleet w. score"])
    df02_1 = df02.loc[df02.variable == "Fleet score",:]
    df02_2 = df02.loc[df02.variable != "Fleet score",:]
    
    # Create figure with secondary y-axis 
    subfig = make_subplots(specs=[[{"secondary_y": True}]])

    # Defining Custom Colors
    colours = {
      "Fleet profit": "#A1A9A3",
      "Fleet w. score": "#DFDF70",
    }

    # create two independent figures with px.line each containing data from multiple columns
    fig = px.line(df02_1, x='Action', y='value',markers=True,range_x=[1,10],width=650,height=650, text="value")
    fig.layout.xaxis.fixedrange= True
    fig.update_traces(line_color='orange', line_width=2)
    fig2 = px.bar(df02_2, y='value', x='Action', color='variable',range_x=[1,10],width=650,height=650,color_discrete_map=colours)
    fig.update_traces(yaxis="y2")
    fig2.update(layout_showlegend=True)
    
    subfig.add_traces(fig2.data + fig.data)
    subfig.layout.xaxis.fixedrange= True
    
    # Set y-axes titles
    subfig.layout.yaxis.title="Kr."
    subfig.layout.yaxis2.title="Score"
    subfig.update_xaxes(range=[0,10], dtick=1)
    st.subheader("*Fleet profit loss by score increase*")
    st.plotly_chart(subfig)
  
     #Save the data for next day 
    cav=open(dir_main+"fleet 3/reports6.xlsx",'rb')
    b64_01 = base64.b64encode(cav.read()).decode('utf-8')
    href = f'<a href="data:file/cav;base64,{b64_01}" download="results6.xlsx">:arrow_down_small: Overall results day 6 </a>'
    st.markdown(href, unsafe_allow_html=True)
  
###############################################################################
#2.i Day 7                                                                    #
###############################################################################
  if pag_input=="Day 7": 
    #Selection of section in the daily report
   # Title
   st.title(title)  
   st.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)
   st.header("**Day 7: Type in information**")
   st.markdown('''In this subsection the fleet manager types in the score and the profit for every action (this is already
accumulated)''')

   with st.expander("Vessel scores"):
#    if "df" not in st.session_state or score.columns[0]=="Unnamed: 0":
    if "df" not in st.session_state:
        st.session_state.df = pd.DataFrame(columns=["Vessel 1: Margaery", 
                                                    "Vessel 2: Brienne", 
                                                    "Vessel 3: Lysa"])


    st.markdown("Add Record")
    
    num_new_rows = 10
    ncol = st.session_state.df.shape[1]  # col count
    rw = -1
    
    with st.form(key="add form", clear_on_submit= True):
        cols = st.columns(ncol)
        rwdta = []
    
        for i in range(ncol):
            rwdta.append(cols[i].text_input(st.session_state.df.columns[i]))
    
        # you can insert code for a list comprehension here to change the data (rwdta) 
        # values into integer / float, if required
    
        if st.form_submit_button("Add"):
            if st.session_state.df.shape[0] == num_new_rows:
                st.error("Add row limit reached. Cant add any more records..")
            else:
                rw = st.session_state.df.shape[0] + 1
                st.info(f"Row: {rw} / {num_new_rows} added")
                st.session_state.df.loc[rw] = rwdta
    
                if st.session_state.df.shape[0] == num_new_rows:
                    st.error("Add row limit reached...")
    df = st.session_state.df
    df.to_csv("fleet 3/score7.csv",encoding='utf-8',index=False)
    st.text(df)
    
   with st.expander("Vessel profits"):
    if "df2" not in st.session_state:
        st.session_state.df2 = pd.DataFrame(columns=["Vessel 1: Margaery", 
                                                    "Vessel 2: Brienne", 
                                                    "Vessel 3: Lysa"])
    
    st.markdown("Add Record")
    
    ncol = st.session_state.df2.shape[1]  # col count
    rw = -1
    
    with st.form(key="add profits form", clear_on_submit= True):
        cols = st.columns(ncol)
        rwdta = []
    
        for i in range(ncol):
            rwdta.append(cols[i].text_input(st.session_state.df2.columns[i]))
    
        # you can insert code for a list comprehension here to change the data (rwdta) 
        # values into integer / float, if required
    
        if st.form_submit_button("Add"):
            if st.session_state.df2.shape[0] == num_new_rows:
                st.error("Add row limit reached. Cant add any more records..")
            else:
                rw = st.session_state.df2.shape[0] + 1
                st.info(f"Row: {rw} / {num_new_rows} added")
                st.session_state.df2.loc[rw] = rwdta
    
                if st.session_state.df2.shape[0] == num_new_rows:
                    st.error("Add row limit reached...")
    


    df2 = st.session_state.df2
    df2.to_csv("fleet 3/profits7.csv",encoding='utf-8',index=False)

    st.text(st.session_state.df2)
 
   with st.expander("Vessel fines"):
    if "df3" not in st.session_state:
        st.session_state.df3 = pd.DataFrame(columns=["Vessel 1: Margaery", 
                                                    "Vessel 2: Brienne", 
                                                    "Vessel 3: Lysa"])
    
    st.markdown("Add Record")
    
    ncol = st.session_state.df3.shape[1]  # col count
    rw = -1
    
    with st.form(key="add fines form", clear_on_submit= True):
        cols = st.columns(ncol)
        rwdta = []
    
        for i in range(ncol):
            rwdta.append(cols[i].text_input(st.session_state.df3.columns[i]))
    
        # you can insert code for a list comprehension here to change the data (rwdta) 
        # values into integer / float, if required
    
        if st.form_submit_button("Add"):
            if st.session_state.df3.shape[0] == num_new_rows:
                st.error("Add row limit reached. Cant add any more records..")
            else:
                rw = st.session_state.df3.shape[0] + 1
                st.info(f"Row: {rw} / {num_new_rows} added")
                st.session_state.df3.loc[rw] = rwdta
    
                if st.session_state.df3.shape[0] == num_new_rows:
                    st.error("Add row limit reached...")
    

    df3 = st.session_state.df3
    df3.to_csv("fleet 3/fines7.csv",encoding='utf-8',index=False)

    st.text(st.session_state.df3)

###############################################################################
#2.i Overall reports                                                          #
###############################################################################
   st.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)
   if st.button('Create overall reports'): 
    st.header("**Day 7: Overall reports**")
    st.markdown('''In this subsection the fleet manager can see the current accumulated score and profit of day 7 and the daily fleet scores and profits.''')
    st.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)
    #
    df = pd.read_csv(os.path.join(dir_main,'fleet 3/score7.csv'), sep=',')
    df2 = pd.read_csv(os.path.join(dir_main,'fleet 3/profits7.csv'), sep=',')
    df3 = pd.read_csv(os.path.join(dir_main,'fleet 3/fines7.csv'), sep=',')
    df4 = pd.read_excel(os.path.join(dir_main,'fleet 3/results5.xlsx'))
    df4.index = df4.Vessel
    df5 = pd.read_excel(os.path.join(dir_main,'fleet 3/results6.xlsx'))
    df5.index = df5.Vessel
    
    d = {'Vessel': df.columns[0:3] , 
         'Score':  df.iloc[-1] , 
         'Score t-1' : df4.Score,
         'Score t-2' : df5.Score,
         'Accumulated Profit':df2.iloc[-1] ,
         'Fines':df3.iloc[-1] }
    outcomes = pd.DataFrame(data=d)
    outcomes=outcomes.reset_index(drop=True)
    st.dataframe(outcomes)

    wb = openpyxl.Workbook()
    wb.save(os.path.join(dir_main,'fleet 3/reports7.xlsx')) 
    book = load_workbook(os.path.join(dir_main,'fleet 3/reports7.xlsx'))
    del book['Sheet']   

    writer = pd.ExcelWriter(os.path.join(dir_main,'fleet 3/reports7.xlsx'), engine='openpyxl')
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}
    outcomes.to_excel(writer, sheet_name = "Outcomes 7", na_rep = "", index =   False, startrow = 0)  
    writer.save()
    
    #Plot scores 
    df01 = pd.read_csv(os.path.join(dir_main,'fleet 3/score7.csv'), sep=',')
    df01.index = np.arange(1, len(df01) + 1)
    #Create the daily average 
    df01['Fleet score']=df01.sum(axis=1)/3
    df01['Action'] = df01.index
    df01 = pd.melt(df01, id_vars='Action', value_vars=["Vessel 1: Margaery", 
                                                    "Vessel 2: Brienne", 
                                                    "Vessel 3: Lysa", 
                                                    "Fleet score"])
    st.subheader("*Daily fleet and vessel scores*")
    fig01 = px.line(df01, x='Action', y='value', color='variable', markers=True,range_x=[0,10],width=650,height=650)
    fig01.update_layout(yaxis_type='category')
    fig01.layout.xaxis.fixedrange= True
    fig01.layout.yaxis.fixedrange: True
    fig01.update(layout_showlegend=True)
    fig01.update_traces(patch={"line": {"color": "green", "width": 3, "dash": 'dot'}}, selector={"legendgroup": "Fleet score"}) 
    st.plotly_chart(fig01)
  
    #Plot profits  
    df02 = pd.read_csv(os.path.join(dir_main,'fleet 3/profits7.csv'), sep=',')
    df03 = pd.read_csv(os.path.join(dir_main,'fleet 3/fines7.csv'), sep=',')
    df02.index = np.arange(1, len(df02) + 1)
    df03.index = np.arange(1, len(df03) + 1)

    df02['Action'] = df02.index
    df02['Fleet profit']=df02.sum(axis=1)
    aux = df01.loc[df01.variable=="Fleet score",'value'].reset_index(drop=True)
    aux.index = np.arange(1, len(aux) + 1)
    df02['Fleet score'] = aux
    
    df03['Fleet fines']=df03.sum(axis=1)
    df02['Fleet w. score'] = df02['Fleet profit'] - df03['Fleet fines']
    
    df02 = pd.melt(df02, id_vars='Action', value_vars=["Fleet profit",
                                                       "Fleet score",
                                                       "Fleet w. score"])
    df02_1 = df02.loc[df02.variable == "Fleet score",:]
    df02_2 = df02.loc[df02.variable != "Fleet score",:]
    
    # Create figure with secondary y-axis 
    subfig = make_subplots(specs=[[{"secondary_y": True}]])

    # Defining Custom Colors
    colours = {
      "Fleet profit": "#A1A9A3",
      "Fleet w. score": "#DFDF70",
    }

    # create two independent figures with px.line each containing data from multiple columns
    fig = px.line(df02_1, x='Action', y='value',markers=True,range_x=[1,10],width=650,height=650, text="value")
    fig.layout.xaxis.fixedrange= True
    fig.update_traces(line_color='orange', line_width=2)
    fig2 = px.bar(df02_2, y='value', x='Action', color='variable',range_x=[1,10],width=650,height=650,color_discrete_map=colours)
    fig.update_traces(yaxis="y2")
    fig2.update(layout_showlegend=True)
    
    subfig.add_traces(fig2.data + fig.data)
    subfig.layout.xaxis.fixedrange= True
    
    # Set y-axes titles
    subfig.layout.yaxis.title="Kr."
    subfig.layout.yaxis2.title="Score"
    subfig.update_xaxes(range=[0,10], dtick=1)
    st.subheader("*Fleet profit loss by score increase*")
    st.plotly_chart(subfig)
  
     #Save the data for next day 
    cav=open(dir_main+"fleet 3/reports7.xlsx",'rb')
    b64_01 = base64.b64encode(cav.read()).decode('utf-8')
    href = f'<a href="data:file/cav;base64,{b64_01}" download="results7.xlsx">:arrow_down_small: Overall results day 7 </a>'
    st.markdown(href, unsafe_allow_html=True)
  
###############################################################################
#2.i Day 8                                                                    #
###############################################################################
  if pag_input=="Day 8": 
    #Selection of section in the daily report
   # Title
   st.title(title)  
   st.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)
   st.header("**Day 8: Type in information**")
   st.markdown('''In this subsection the fleet manager types in the score and the profit for every action (this is already
accumulated)''')

   with st.expander("Vessel scores"):
#    if "df" not in st.session_state or score.columns[0]=="Unnamed: 0":
    if "df" not in st.session_state:
        st.session_state.df = pd.DataFrame(columns=["Vessel 1: Margaery", 
                                                    "Vessel 2: Brienne", 
                                                    "Vessel 3: Lysa"])


    st.markdown("Add Record")
    
    num_new_rows = 10
    ncol = st.session_state.df.shape[1]  # col count
    rw = -1
    
    with st.form(key="add form", clear_on_submit= True):
        cols = st.columns(ncol)
        rwdta = []
    
        for i in range(ncol):
            rwdta.append(cols[i].text_input(st.session_state.df.columns[i]))
    
        # you can insert code for a list comprehension here to change the data (rwdta) 
        # values into integer / float, if required
    
        if st.form_submit_button("Add"):
            if st.session_state.df.shape[0] == num_new_rows:
                st.error("Add row limit reached. Cant add any more records..")
            else:
                rw = st.session_state.df.shape[0] + 1
                st.info(f"Row: {rw} / {num_new_rows} added")
                st.session_state.df.loc[rw] = rwdta
    
                if st.session_state.df.shape[0] == num_new_rows:
                    st.error("Add row limit reached...")
    df = st.session_state.df
    df.to_csv("fleet 3/score8.csv",encoding='utf-8',index=False)
    st.text(df)
    
   with st.expander("Vessel profits"):
    if "df2" not in st.session_state:
        st.session_state.df2 = pd.DataFrame(columns=["Vessel 1: Margaery", 
                                                    "Vessel 2: Brienne", 
                                                    "Vessel 3: Lysa"])
    
    st.markdown("Add Record")
    
    ncol = st.session_state.df2.shape[1]  # col count
    rw = -1
    
    with st.form(key="add profits form", clear_on_submit= True):
        cols = st.columns(ncol)
        rwdta = []
    
        for i in range(ncol):
            rwdta.append(cols[i].text_input(st.session_state.df2.columns[i]))
    
        # you can insert code for a list comprehension here to change the data (rwdta) 
        # values into integer / float, if required
    
        if st.form_submit_button("Add"):
            if st.session_state.df2.shape[0] == num_new_rows:
                st.error("Add row limit reached. Cant add any more records..")
            else:
                rw = st.session_state.df2.shape[0] + 1
                st.info(f"Row: {rw} / {num_new_rows} added")
                st.session_state.df2.loc[rw] = rwdta
    
                if st.session_state.df2.shape[0] == num_new_rows:
                    st.error("Add row limit reached...")
    


    df2 = st.session_state.df2
    df2.to_csv("fleet 3/profits8.csv",encoding='utf-8',index=False)

    st.text(st.session_state.df2)
 
   with st.expander("Vessel fines"):
    if "df3" not in st.session_state:
        st.session_state.df3 = pd.DataFrame(columns=["Vessel 1: Margaery", 
                                                    "Vessel 2: Brienne", 
                                                    "Vessel 3: Lysa"])
    
    st.markdown("Add Record")
    
    ncol = st.session_state.df3.shape[1]  # col count
    rw = -1
    
    with st.form(key="add fines form", clear_on_submit= True):
        cols = st.columns(ncol)
        rwdta = []
    
        for i in range(ncol):
            rwdta.append(cols[i].text_input(st.session_state.df3.columns[i]))
    
        # you can insert code for a list comprehension here to change the data (rwdta) 
        # values into integer / float, if required
    
        if st.form_submit_button("Add"):
            if st.session_state.df3.shape[0] == num_new_rows:
                st.error("Add row limit reached. Cant add any more records..")
            else:
                rw = st.session_state.df3.shape[0] + 1
                st.info(f"Row: {rw} / {num_new_rows} added")
                st.session_state.df3.loc[rw] = rwdta
    
                if st.session_state.df3.shape[0] == num_new_rows:
                    st.error("Add row limit reached...")
    

    df3 = st.session_state.df3
    df3.to_csv("fleet 3/fines8.csv",encoding='utf-8',index=False)

    st.text(st.session_state.df3)

###############################################################################
#2.i Overall reports                                                          #
###############################################################################
   st.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)
   if st.button('Create overall reports'): 
    st.header("**Day 8: Overall reports**")
    st.markdown('''In this subsection the fleet manager can see the current accumulated score and profit of day 8 and the daily fleet scores and profits.''')
    st.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)
    #
    df = pd.read_csv(os.path.join(dir_main,'fleet 3/score8.csv'), sep=',')
    df2 = pd.read_csv(os.path.join(dir_main,'fleet 3/profits8.csv'), sep=',')
    df3 = pd.read_csv(os.path.join(dir_main,'fleet 3/fines8.csv'), sep=',')
    df4 = pd.read_excel(os.path.join(dir_main,'fleet 3/results6.xlsx'))
    df4.index = df4.Vessel
    df5 = pd.read_excel(os.path.join(dir_main,'fleet 3/results7.xlsx'))
    df5.index = df5.Vessel
    
    d = {'Vessel': df.columns[0:3] , 
         'Score':  df.iloc[-1] , 
         'Score t-1' : df4.Score,
         'Score t-2' : df5.Score,
         'Accumulated Profit':df2.iloc[-1] ,
         'Fines':df3.iloc[-1] }
    outcomes = pd.DataFrame(data=d)
    outcomes=outcomes.reset_index(drop=True)
    st.dataframe(outcomes)

    wb = openpyxl.Workbook()
    wb.save(os.path.join(dir_main,'fleet 3/reports8.xlsx')) 
    book = load_workbook(os.path.join(dir_main,'fleet 3/reports8.xlsx'))
    del book['Sheet']   

    writer = pd.ExcelWriter(os.path.join(dir_main,'fleet 3/reports8.xlsx'), engine='openpyxl')
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}
    outcomes.to_excel(writer, sheet_name = "Outcomes 8", na_rep = "", index =   False, startrow = 0)  
    writer.save()
    
    #Plot scores 
    df01 = pd.read_csv(os.path.join(dir_main,'fleet 3/score8.csv'), sep=',')
    df01.index = np.arange(1, len(df01) + 1)
    #Create the daily average 
    df01['Fleet score']=df01.sum(axis=1)/3
    df01['Action'] = df01.index
    df01 = pd.melt(df01, id_vars='Action', value_vars=["Vessel 1: Margaery", 
                                                    "Vessel 2: Brienne", 
                                                    "Vessel 3: Lysa", 
                                                    "Fleet score"])
    st.subheader("*Daily fleet and vessel scores*")
    fig01 = px.line(df01, x='Action', y='value', color='variable', markers=True,range_x=[0,10],width=650,height=650)
    fig01.update_layout(yaxis_type='category')
    fig01.layout.xaxis.fixedrange= True
    fig01.layout.yaxis.fixedrange: True
    fig01.update(layout_showlegend=True)
    fig01.update_traces(patch={"line": {"color": "green", "width": 3, "dash": 'dot'}}, selector={"legendgroup": "Fleet score"}) 
    st.plotly_chart(fig01)
  
    #Plot profits  
    df02 = pd.read_csv(os.path.join(dir_main,'fleet 3/profits8.csv'), sep=',')
    df03 = pd.read_csv(os.path.join(dir_main,'fleet 3/fines8.csv'), sep=',')
    df02.index = np.arange(1, len(df02) + 1)
    df03.index = np.arange(1, len(df03) + 1)

    df02['Action'] = df02.index
    df02['Fleet profit']=df02.sum(axis=1)
    aux = df01.loc[df01.variable=="Fleet score",'value'].reset_index(drop=True)
    aux.index = np.arange(1, len(aux) + 1)
    df02['Fleet score'] = aux
    
    df03['Fleet fines']=df03.sum(axis=1)
    df02['Fleet w. score'] = df02['Fleet profit'] - df03['Fleet fines']
    
    df02 = pd.melt(df02, id_vars='Action', value_vars=["Fleet profit",
                                                       "Fleet score",
                                                       "Fleet w. score"])
    df02_1 = df02.loc[df02.variable == "Fleet score",:]
    df02_2 = df02.loc[df02.variable != "Fleet score",:]
    
    # Create figure with secondary y-axis 
    subfig = make_subplots(specs=[[{"secondary_y": True}]])

    # Defining Custom Colors
    colours = {
      "Fleet profit": "#A1A9A3",
      "Fleet w. score": "#DFDF70",
    }

    # create two independent figures with px.line each containing data from multiple columns
    fig = px.line(df02_1, x='Action', y='value',markers=True,range_x=[1,10],width=650,height=650, text="value")
    fig.layout.xaxis.fixedrange= True
    fig.update_traces(line_color='orange', line_width=2)
    fig2 = px.bar(df02_2, y='value', x='Action', color='variable',range_x=[1,10],width=650,height=650,color_discrete_map=colours)
    fig.update_traces(yaxis="y2")
    fig2.update(layout_showlegend=True)
    
    subfig.add_traces(fig2.data + fig.data)
    subfig.layout.xaxis.fixedrange= True
    
    # Set y-axes titles
    subfig.layout.yaxis.title="Kr."
    subfig.layout.yaxis2.title="Score"
    subfig.update_xaxes(range=[0,10], dtick=1)
    st.subheader("*Fleet profit loss by score increase*")
    st.plotly_chart(subfig)
  
     #Save the data for next day 
    cav=open(dir_main+"fleet 3/reports8.xlsx",'rb')
    b64_01 = base64.b64encode(cav.read()).decode('utf-8')
    href = f'<a href="data:file/cav;base64,{b64_01}" download="results8.xlsx">:arrow_down_small: Overall results day 8 </a>'
    st.markdown(href, unsafe_allow_html=True)
  
###############################################################################
#2.i Day 4                                                                    #
###############################################################################
  if pag_input=="Day 9": 
    #Selection of section in the daily report
   # Title
   st.title(title)  
   st.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)
   st.header("**Day 9: Type in information**")
   st.markdown('''In this subsection the fleet manager types in the score and the profit for every action (this is already
accumulated)''')

   with st.expander("Vessel scores"):
#    if "df" not in st.session_state or score.columns[0]=="Unnamed: 0":
    if "df" not in st.session_state:
        st.session_state.df = pd.DataFrame(columns=["Vessel 1: Margaery", 
                                                    "Vessel 2: Brienne", 
                                                    "Vessel 3: Lysa"])


    st.markdown("Add Record")
    
    num_new_rows = 10
    ncol = st.session_state.df.shape[1]  # col count
    rw = -1
    
    with st.form(key="add form", clear_on_submit= True):
        cols = st.columns(ncol)
        rwdta = []
    
        for i in range(ncol):
            rwdta.append(cols[i].text_input(st.session_state.df.columns[i]))
    
        # you can insert code for a list comprehension here to change the data (rwdta) 
        # values into integer / float, if required
    
        if st.form_submit_button("Add"):
            if st.session_state.df.shape[0] == num_new_rows:
                st.error("Add row limit reached. Cant add any more records..")
            else:
                rw = st.session_state.df.shape[0] + 1
                st.info(f"Row: {rw} / {num_new_rows} added")
                st.session_state.df.loc[rw] = rwdta
    
                if st.session_state.df.shape[0] == num_new_rows:
                    st.error("Add row limit reached...")
    df = st.session_state.df
    df.to_csv("fleet 3/score9.csv",encoding='utf-8',index=False)
    st.text(df)
    
   with st.expander("Vessel profits"):
    if "df2" not in st.session_state:
        st.session_state.df2 = pd.DataFrame(columns=["Vessel 1: Margaery", 
                                                    "Vessel 2: Brienne", 
                                                    "Vessel 3: Lysa"])
    
    st.markdown("Add Record")
    
    ncol = st.session_state.df2.shape[1]  # col count
    rw = -1
    
    with st.form(key="add profits form", clear_on_submit= True):
        cols = st.columns(ncol)
        rwdta = []
    
        for i in range(ncol):
            rwdta.append(cols[i].text_input(st.session_state.df2.columns[i]))
    
        # you can insert code for a list comprehension here to change the data (rwdta) 
        # values into integer / float, if required
    
        if st.form_submit_button("Add"):
            if st.session_state.df2.shape[0] == num_new_rows:
                st.error("Add row limit reached. Cant add any more records..")
            else:
                rw = st.session_state.df2.shape[0] + 1
                st.info(f"Row: {rw} / {num_new_rows} added")
                st.session_state.df2.loc[rw] = rwdta
    
                if st.session_state.df2.shape[0] == num_new_rows:
                    st.error("Add row limit reached...")
    


    df2 = st.session_state.df2
    df2.to_csv("fleet 3/profits9.csv",encoding='utf-8',index=False)

    st.text(st.session_state.df2)
 
   with st.expander("Vessel fines"):
    if "df3" not in st.session_state:
        st.session_state.df3 = pd.DataFrame(columns=["Vessel 1: Margaery", 
                                                    "Vessel 2: Brienne", 
                                                    "Vessel 3: Lysa"])
    
    st.markdown("Add Record")
    
    ncol = st.session_state.df3.shape[1]  # col count
    rw = -1
    
    with st.form(key="add fines form", clear_on_submit= True):
        cols = st.columns(ncol)
        rwdta = []
    
        for i in range(ncol):
            rwdta.append(cols[i].text_input(st.session_state.df3.columns[i]))
    
        # you can insert code for a list comprehension here to change the data (rwdta) 
        # values into integer / float, if required
    
        if st.form_submit_button("Add"):
            if st.session_state.df3.shape[0] == num_new_rows:
                st.error("Add row limit reached. Cant add any more records..")
            else:
                rw = st.session_state.df3.shape[0] + 1
                st.info(f"Row: {rw} / {num_new_rows} added")
                st.session_state.df3.loc[rw] = rwdta
    
                if st.session_state.df3.shape[0] == num_new_rows:
                    st.error("Add row limit reached...")
    

    df3 = st.session_state.df3
    df3.to_csv("fleet 3/fines9.csv",encoding='utf-8',index=False)

    st.text(st.session_state.df3)

###############################################################################
#2.i Overall reports                                                          #
###############################################################################
   st.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)
   if st.button('Create overall reports'): 
    st.header("**Day 9: Overall reports**")
    st.markdown('''In this subsection the fleet manager can see the current accumulated score and profit of day 9 and the daily fleet scores and profits.''')
    st.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)
    #
    df = pd.read_csv(os.path.join(dir_main,'fleet 3/score9.csv'), sep=',')
    df2 = pd.read_csv(os.path.join(dir_main,'fleet 3/profits9.csv'), sep=',')
    df3 = pd.read_csv(os.path.join(dir_main,'fleet 3/fines9.csv'), sep=',')
    df4 = pd.read_excel(os.path.join(dir_main,'fleet 3/results7.xlsx'))
    df4.index = df4.Vessel
    df5 = pd.read_excel(os.path.join(dir_main,'fleet 3/results8.xlsx'))
    df5.index = df5.Vessel
    
    d = {'Vessel': df.columns[0:3] , 
         'Score':  df.iloc[-1] , 
         'Score t-1' : df4.Score,
         'Score t-2' : df5.Score,
         'Accumulated Profit':df2.iloc[-1] ,
         'Fines':df3.iloc[-1] }
    outcomes = pd.DataFrame(data=d)
    outcomes=outcomes.reset_index(drop=True)
    st.dataframe(outcomes)

    wb = openpyxl.Workbook()
    wb.save(os.path.join(dir_main,'fleet 3/reports9.xlsx')) 
    book = load_workbook(os.path.join(dir_main,'fleet 3/reports9.xlsx'))
    del book['Sheet']   

    writer = pd.ExcelWriter(os.path.join(dir_main,'fleet 3/reports9.xlsx'), engine='openpyxl')
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}
    outcomes.to_excel(writer, sheet_name = "Outcomes 9", na_rep = "", index =   False, startrow = 0)  
    writer.save()
    
    #Plot scores 
    df01 = pd.read_csv(os.path.join(dir_main,'fleet 3/score9.csv'), sep=',')
    df01.index = np.arange(1, len(df01) + 1)
    #Create the daily average 
    df01['Fleet score']=df01.sum(axis=1)/3
    df01['Action'] = df01.index
    df01 = pd.melt(df01, id_vars='Action', value_vars=["Vessel 1: Margaery", 
                                                    "Vessel 2: Brienne", 
                                                    "Vessel 3: Lysa", 
                                                    "Fleet score"])
    st.subheader("*Daily fleet and vessel scores*")
    fig01 = px.line(df01, x='Action', y='value', color='variable', markers=True,range_x=[0,10],width=650,height=650)
    fig01.update_layout(yaxis_type='category')
    fig01.layout.xaxis.fixedrange= True
    fig01.layout.yaxis.fixedrange: True
    fig01.update(layout_showlegend=True)
    fig01.update_traces(patch={"line": {"color": "green", "width": 3, "dash": 'dot'}}, selector={"legendgroup": "Fleet score"}) 
    st.plotly_chart(fig01)
  
    #Plot profits  
    df02 = pd.read_csv(os.path.join(dir_main,'fleet 3/profits9.csv'), sep=',')
    df03 = pd.read_csv(os.path.join(dir_main,'fleet 3/fines9.csv'), sep=',')
    df02.index = np.arange(1, len(df02) + 1)
    df03.index = np.arange(1, len(df03) + 1)

    df02['Action'] = df02.index
    df02['Fleet profit']=df02.sum(axis=1)
    aux = df01.loc[df01.variable=="Fleet score",'value'].reset_index(drop=True)
    aux.index = np.arange(1, len(aux) + 1)
    df02['Fleet score'] = aux
    
    df03['Fleet fines']=df03.sum(axis=1)
    df02['Fleet w. score'] = df02['Fleet profit'] - df03['Fleet fines']
    
    df02 = pd.melt(df02, id_vars='Action', value_vars=["Fleet profit",
                                                       "Fleet score",
                                                       "Fleet w. score"])
    df02_1 = df02.loc[df02.variable == "Fleet score",:]
    df02_2 = df02.loc[df02.variable != "Fleet score",:]
    
    # Create figure with secondary y-axis 
    subfig = make_subplots(specs=[[{"secondary_y": True}]])

    # Defining Custom Colors
    colours = {
      "Fleet profit": "#A1A9A3",
      "Fleet w. score": "#DFDF70",
    }

    # create two independent figures with px.line each containing data from multiple columns
    fig = px.line(df02_1, x='Action', y='value',markers=True,range_x=[1,10],width=650,height=650, text="value")
    fig.layout.xaxis.fixedrange= True
    fig.update_traces(line_color='orange', line_width=2)
    fig2 = px.bar(df02_2, y='value', x='Action', color='variable',range_x=[1,10],width=650,height=650,color_discrete_map=colours)
    fig.update_traces(yaxis="y2")
    fig2.update(layout_showlegend=True)
    
    subfig.add_traces(fig2.data + fig.data)
    subfig.layout.xaxis.fixedrange= True
    
    # Set y-axes titles
    subfig.layout.yaxis.title="Kr."
    subfig.layout.yaxis2.title="Score"
    subfig.update_xaxes(range=[0,10], dtick=1)
    st.subheader("*Fleet profit loss by score increase*")
    st.plotly_chart(subfig)
  
     #Save the data for next day 
    cav=open(dir_main+"fleet 3/reports9.xlsx",'rb')
    b64_01 = base64.b64encode(cav.read()).decode('utf-8')
    href = f'<a href="data:file/cav;base64,{b64_01}" download="results9.xlsx">:arrow_down_small: Overall results day 9 </a>'
    st.markdown(href, unsafe_allow_html=True)
  
  
###############################################################################
#2.i Day 10                                                                   #
###############################################################################
  if pag_input=="Day 10": 
    #Selection of section in the daily report
   # Title
   st.title(title)  
   st.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)
   st.header("**Day 10: Type in information**")
   st.markdown('''In this subsection the fleet manager types in the score and the profit for every action (this is already
accumulated)''')

   with st.expander("Vessel scores"):
#    if "df" not in st.session_state or score.columns[0]=="Unnamed: 0":
    if "df" not in st.session_state:
        st.session_state.df = pd.DataFrame(columns=["Vessel 1: Margaery", 
                                                    "Vessel 2: Brienne", 
                                                    "Vessel 3: Lysa"])


    st.markdown("Add Record")
    
    num_new_rows = 10
    ncol = st.session_state.df.shape[1]  # col count
    rw = -1
    
    with st.form(key="add form", clear_on_submit= True):
        cols = st.columns(ncol)
        rwdta = []
    
        for i in range(ncol):
            rwdta.append(cols[i].text_input(st.session_state.df.columns[i]))
    
        # you can insert code for a list comprehension here to change the data (rwdta) 
        # values into integer / float, if required
    
        if st.form_submit_button("Add"):
            if st.session_state.df.shape[0] == num_new_rows:
                st.error("Add row limit reached. Cant add any more records..")
            else:
                rw = st.session_state.df.shape[0] + 1
                st.info(f"Row: {rw} / {num_new_rows} added")
                st.session_state.df.loc[rw] = rwdta
    
                if st.session_state.df.shape[0] == num_new_rows:
                    st.error("Add row limit reached...")
    df = st.session_state.df
    df.to_csv("fleet 3/score10.csv",encoding='utf-8',index=False)
    st.text(df)
    
   with st.expander("Vessel profits"):
    if "df2" not in st.session_state:
        st.session_state.df2 = pd.DataFrame(columns=["Vessel 1: Margaery", 
                                                    "Vessel 2: Brienne", 
                                                    "Vessel 3: Lysa"])
    
    st.markdown("Add Record")
    
    ncol = st.session_state.df2.shape[1]  # col count
    rw = -1
    
    with st.form(key="add profits form", clear_on_submit= True):
        cols = st.columns(ncol)
        rwdta = []
    
        for i in range(ncol):
            rwdta.append(cols[i].text_input(st.session_state.df2.columns[i]))
    
        # you can insert code for a list comprehension here to change the data (rwdta) 
        # values into integer / float, if required
    
        if st.form_submit_button("Add"):
            if st.session_state.df2.shape[0] == num_new_rows:
                st.error("Add row limit reached. Cant add any more records..")
            else:
                rw = st.session_state.df2.shape[0] + 1
                st.info(f"Row: {rw} / {num_new_rows} added")
                st.session_state.df2.loc[rw] = rwdta
    
                if st.session_state.df2.shape[0] == num_new_rows:
                    st.error("Add row limit reached...")
    


    df2 = st.session_state.df2
    df2.to_csv("fleet 3/profits10.csv",encoding='utf-8',index=False)

    st.text(st.session_state.df2)
 
   with st.expander("Vessel fines"):
    if "df3" not in st.session_state:
        st.session_state.df3 = pd.DataFrame(columns=["Vessel 1: Margaery", 
                                                    "Vessel 2: Brienne", 
                                                    "Vessel 3: Lysa"])
    
    st.markdown("Add Record")
    
    ncol = st.session_state.df3.shape[1]  # col count
    rw = -1
    
    with st.form(key="add fines form", clear_on_submit= True):
        cols = st.columns(ncol)
        rwdta = []
    
        for i in range(ncol):
            rwdta.append(cols[i].text_input(st.session_state.df3.columns[i]))
    
        # you can insert code for a list comprehension here to change the data (rwdta) 
        # values into integer / float, if required
    
        if st.form_submit_button("Add"):
            if st.session_state.df3.shape[0] == num_new_rows:
                st.error("Add row limit reached. Cant add any more records..")
            else:
                rw = st.session_state.df3.shape[0] + 1
                st.info(f"Row: {rw} / {num_new_rows} added")
                st.session_state.df3.loc[rw] = rwdta
    
                if st.session_state.df3.shape[0] == num_new_rows:
                    st.error("Add row limit reached...")
    

    df3 = st.session_state.df3
    df3.to_csv("fleet 3/fines10.csv",encoding='utf-8',index=False)

    st.text(st.session_state.df3)

###############################################################################
#2.i Overall reports                                                          #
###############################################################################
   st.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)
   if st.button('Create overall reports'): 
    st.header("**Day 10: Overall reports**")
    st.markdown('''In this subsection the fleet manager can see the current accumulated score and profit of day 10 and the daily fleet scores and profits.''')
    st.markdown('''
              <hr style="border:1.75px solid black"> </hr>
              ''', unsafe_allow_html=True)
    #
    df = pd.read_csv(os.path.join(dir_main,'fleet 3/score10.csv'), sep=',')
    df2 = pd.read_csv(os.path.join(dir_main,'fleet 3/profits10.csv'), sep=',')
    df3 = pd.read_csv(os.path.join(dir_main,'fleet 3/fines10.csv'), sep=',')
    df4 = pd.read_excel(os.path.join(dir_main,'fleet 3/results8.xlsx'))
    df4.index = df4.Vessel
    df5 = pd.read_excel(os.path.join(dir_main,'fleet 3/results9.xlsx'))
    df5.index = df5.Vessel
    
    d = {'Vessel': df.columns[0:3] , 
         'Score':  df.iloc[-1] , 
         'Score t-1' : df4.Score,
         'Score t-2' : df5.Score,
         'Accumulated Profit':df2.iloc[-1] ,
         'Fines':df3.iloc[-1] }
    outcomes = pd.DataFrame(data=d)
    outcomes=outcomes.reset_index(drop=True)
    st.dataframe(outcomes)

    wb = openpyxl.Workbook()
    wb.save(os.path.join(dir_main,'fleet 3/reports10.xlsx')) 
    book = load_workbook(os.path.join(dir_main,'fleet 3/reports10.xlsx'))
    del book['Sheet']   

    writer = pd.ExcelWriter(os.path.join(dir_main,'fleet 3/reports10.xlsx'), engine='openpyxl')
    writer.book = book
    writer.sheets = {ws.title: ws for ws in book.worksheets}
    outcomes.to_excel(writer, sheet_name = "Outcomes 10", na_rep = "", index =   False, startrow = 0)  
    writer.save()
    
    #Plot scores 
    df01 = pd.read_csv(os.path.join(dir_main,'fleet 3/score10.csv'), sep=',')
    df01.index = np.arange(1, len(df01) + 1)
    #Create the daily average 
    df01['Fleet score']=df01.sum(axis=1)/3
    df01['Action'] = df01.index
    df01 = pd.melt(df01, id_vars='Action', value_vars=["Vessel 1: Margaery", 
                                                    "Vessel 2: Brienne", 
                                                    "Vessel 3: Lysa", 
                                                    "Fleet score"])
    st.subheader("*Daily fleet and vessel scores*")
    fig01 = px.line(df01, x='Action', y='value', color='variable', markers=True,range_x=[0,10],width=650,height=650)
    fig01.update_layout(yaxis_type='category')
    fig01.layout.xaxis.fixedrange= True
    fig01.layout.yaxis.fixedrange: True
    fig01.update(layout_showlegend=True)
    fig01.update_traces(patch={"line": {"color": "green", "width": 3, "dash": 'dot'}}, selector={"legendgroup": "Fleet score"}) 
    st.plotly_chart(fig01)
  
    #Plot profits  
    df02 = pd.read_csv(os.path.join(dir_main,'fleet 3/profits10.csv'), sep=',')
    df03 = pd.read_csv(os.path.join(dir_main,'fleet 3/fines10.csv'), sep=',')
    df02.index = np.arange(1, len(df02) + 1)
    df03.index = np.arange(1, len(df03) + 1)

    df02['Action'] = df02.index
    df02['Fleet profit']=df02.sum(axis=1)
    aux = df01.loc[df01.variable=="Fleet score",'value'].reset_index(drop=True)
    aux.index = np.arange(1, len(aux) + 1)
    df02['Fleet score'] = aux
    
    df03['Fleet fines']=df03.sum(axis=1)
    df02['Fleet w. score'] = df02['Fleet profit'] - df03['Fleet fines']
    
    df02 = pd.melt(df02, id_vars='Action', value_vars=["Fleet profit",
                                                       "Fleet score",
                                                       "Fleet w. score"])
    df02_1 = df02.loc[df02.variable == "Fleet score",:]
    df02_2 = df02.loc[df02.variable != "Fleet score",:]
    
    # Create figure with secondary y-axis 
    subfig = make_subplots(specs=[[{"secondary_y": True}]])

    # Defining Custom Colors
    colours = {
      "Fleet profit": "#A1A9A3",
      "Fleet w. score": "#DFDF70",
    }

    # create two independent figures with px.line each containing data from multiple columns
    fig = px.line(df02_1, x='Action', y='value',markers=True,range_x=[1,10],width=650,height=650, text="value")
    fig.layout.xaxis.fixedrange= True
    fig.update_traces(line_color='orange', line_width=2)
    fig2 = px.bar(df02_2, y='value', x='Action', color='variable',range_x=[1,10],width=650,height=650,color_discrete_map=colours)
    fig.update_traces(yaxis="y2")
    fig2.update(layout_showlegend=True)
    
    subfig.add_traces(fig2.data + fig.data)
    subfig.layout.xaxis.fixedrange= True
    
    # Set y-axes titles
    subfig.layout.yaxis.title="Kr."
    subfig.layout.yaxis2.title="Score"
    subfig.update_xaxes(range=[0,10], dtick=1)
    st.subheader("*Fleet profit loss by score increase*")
    st.plotly_chart(subfig)
  
     #Save the data for next day 
    cav=open(dir_main+"fleet 3/reports10.xlsx",'rb')
    b64_01 = base64.b64encode(cav.read()).decode('utf-8')
    href = f'<a href="data:file/cav;base64,{b64_01}" download="results10.xlsx">:arrow_down_small: Overall results day 10 </a>'
    st.markdown(href, unsafe_allow_html=True)
   
    
